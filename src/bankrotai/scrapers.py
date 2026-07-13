from __future__ import annotations

import hashlib
import json
import logging
import math
import random
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, replace
from io import BytesIO
from urllib.parse import urlencode, urljoin
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from datetime import datetime

from bankrotai.domain import NormalizedLot
from typing import Callable, Optional
from bankrotai.logic import classify_category, persist_lot, upsert_lot_events_from_raw

logger = logging.getLogger(__name__)

from bankrotai.extractors import (
    extract_price, extract_area, extract_cadastral, extract_address,
    extract_land_area, extract_floors, extract_legal_status,
    extract_cadastral_numbers, extract_building_area, extract_room_area,
    extract_year_built, extract_commissioning_year, extract_cultural_heritage
)

# --- Parsers ---

@dataclass
class ParserConfig:
    container_selector: str = "div.lot"
    title_selector: str = "a"
    base_url: str = "https://tbankrot.ru"

@dataclass
class ParsedLotData:
    external_id: str
    title: str
    url: str

    # Цена
    price_text: str = ""
    current_price: float | None = None

    # Текстовые данные
    description: str = ""
    status: str = "unknown"
    address: str = ""

    # Кадастр
    cadastral_number: str = ""
    cadastral_numbers: list[str] = field(default_factory=list)

    # Площади
    area: float | None = None
    building_area: float | None = None
    room_area: float | None = None
    land_area: float | None = None

    # Тех. параметры
    floors: int | None = None
    year_built: int | None = None
    year_commissioning: int | None = None
    is_cultural_heritage: bool = False
    raw_payload: dict[str, Any] = field(default_factory=dict)


def parse_money(text: str | None) -> float | None:
    if not text:
        return None

    clean = (
        text.replace("\xa0", " ")
            .replace("\u202f", " ")
            .replace("₽", "")
            .replace("руб.", "")
            .replace("руб", "")
            .strip()
    )

    # Берём первое денежное число из конкретного контейнера цены
    m = re.search(r'\d[\d\s]*(?:[,.]\d+)?', clean)
    if not m:
        return None

    raw = m.group(0).replace(" ", "").replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


@dataclass
class TorgiGovSearchFilters:
    search_text: str = ""
    type_transaction: str | None = None
    price_min: float | None = None
    price_max: float | None = None
    subject_rf: str | None = None
    fias: str | None = None
    ownership_form: str | None = None
    category_code: str | None = None
    lot_status: str | None = None
    currency_code: str | None = None
    publish_date_from: str | None = None
    publish_date_to: str | None = None
    bidd_end_time_from: str | None = None
    bidd_end_time_to: str | None = None
    auction_start_date_from: str | None = None
    auction_start_date_to: str | None = None
    notice_number: str | None = None
    etp_code: str | None = None
    bidd_type: str | None = None
    bidd_form: str | None = None
    notice_status: str | None = None
    organizer_name: str | None = None
    organizer_inn: str | None = None
    right_holder_name: str | None = None
    right_holder_inn: str | None = None
    attachment_text: str | None = None
    match_phrase: bool = False
    is_msp: bool = False
    page: int = 1
    page_size: int = 20


@dataclass
class TBankrotSearchFilters:
    search_text: str = ""
    region: str | None = None
    price_min: float | None = None
    price_max: float | None = None
    lot_number: str | None = None
    trade_type: str | None = None
    photo_only: bool = False
    debtor: str | None = None
    auction_manager: str | None = None
    organizer: str | None = None
    stop_words: str | None = None
    show_closed: bool = False
    show_paused: bool = False
    page: int = 1
    page_size: int = 20


class TorgiGovClientError(RuntimeError):
    """User-facing error raised when torgi.gov.ru cannot be queried."""


class TorgiGovClient:
    BASE_URL = "https://torgi.gov.ru"
    SEARCH_ENDPOINT = f"{BASE_URL}/new/api/public/lotcards/search"
    EXCEL_EXPORT_ENDPOINT = f"{BASE_URL}/new/api/public/lotcards/export/excel"
    FALLBACK_LIST_URL = f"{BASE_URL}/new/public/lots/reg"
    DEFAULT_LOT_STATUS = "PUBLISHED,APPLICATIONS_SUBMISSION"

    CATEGORY_CODE_LABELS = {
        "903": "Земельный участок со зданием",
        "22": "Транспорт",
        "7": "Недвижимость",
        "8": "Здания",
        "9": "Жилые помещения",
        "10": "Нежилые помещения",
        "11": "Сооружения",
        "12": "Объекты незавершенного строительства",
        "2": "Земельные участки",
        "5": "Акции и Доли",
        "6": "Права пользования и лицензии",
        "23": "Строительство и развитие территорий",
        "17": "Государственно-частное партнерство",
        "900": "ЖКХ",
        "13": "Прочее",
    }
    CATEGORY_LABEL_TO_CODE = {v.lower(): k for k, v in CATEGORY_CODE_LABELS.items()}
    CATEGORY_GROUP_CODE_MAP = {
        # On torgi.gov.ru the visible "Недвижимость" node is a parent category.
        # Search/export endpoints return real results for its child category codes.
        "7": "8,9,10,11,12",
    }
    CATEGORY_CODE_TO_INTERNAL = {
        "903": "commercial_building_with_land",
        "22": "transport",
        "7": "real_estate",
        "8": "real_estate",
        "9": "real_estate",
        "10": "real_estate",
        "11": "real_estate",
        "12": "real_estate",
        "2": "land",
        "5": "other",
        "6": "other",
        "23": "other",
        "17": "other",
        "900": "other",
        "13": "other",
    }
    SUBJECT_RF_CODES = {
        "Алтайский край": "22",
        "Краснодарский край": "23",
        "Красноярский край": "24",
        "Москва": "77",
        "Московская область": "50",
        "Омская область": "55",
        "Пермский край": "59",
        "Республика Башкортостан": "02",
        "Республика Татарстан": "16",
        "Ростовская область": "61",
        "Самарская область": "63",
        "Санкт-Петербург": "78",
        "Свердловская область": "66",
        "Тульская область": "71",
        "Тюменская область": "72",
        "Челябинская область": "74",
        "Ярославская область": "76",
    }

    SUBJECT_RF_CODES.update({
        "Алтайский край": "22", "Амурская область": "28", "Архангельская область": "29",
        "Астраханская область": "30", "Белгородская область": "31", "Брянская область": "32",
        "Владимирская область": "33", "Волгоградская область": "34", "Вологодская область": "35",
        "Воронежская область": "36", "Еврейская автономная область": "79", "Забайкальский край": "75",
        "Ивановская область": "37", "Иркутская область": "38", "Кабардино-Балкарская Республика": "07",
        "Калининградская область": "39", "Калужская область": "40", "Камчатский край": "41",
        "Карачаево-Черкесская Республика": "09", "Кемеровская область": "42", "Кировская область": "43",
        "Костромская область": "44", "Краснодарский край": "23", "Красноярский край": "24",
        "Курганская область": "45", "Курская область": "46", "Ленинградская область": "47",
        "Липецкая область": "48", "Магаданская область": "49", "Москва": "77",
        "Московская область": "50", "Мурманская область": "51", "Ненецкий автономный округ": "83",
        "Нижегородская область": "52", "Новгородская область": "53", "Новосибирская область": "54",
        "Омская область": "55", "Оренбургская область": "56", "Орловская область": "57",
        "Пензенская область": "58", "Пермский край": "59", "Приморский край": "25",
        "Псковская область": "60", "Республика Адыгея": "01", "Республика Алтай": "04",
        "Республика Башкортостан": "02", "Республика Бурятия": "03", "Республика Дагестан": "05",
        "Республика Ингушетия": "06", "Республика Калмыкия": "08", "Республика Карелия": "10",
        "Республика Коми": "11", "Республика Крым": "91", "Республика Марий Эл": "12",
        "Республика Мордовия": "13", "Республика Саха (Якутия)": "14",
        "Республика Северная Осетия - Алания": "15", "Республика Татарстан": "16",
        "Республика Тыва": "17", "Республика Хакасия": "19", "Ростовская область": "61",
        "Рязанская область": "62", "Самарская область": "63", "Санкт-Петербург": "78",
        "Саратовская область": "64", "Сахалинская область": "65", "Свердловская область": "66",
        "Севастополь": "92", "Смоленская область": "67", "Ставропольский край": "26",
        "Тамбовская область": "68", "Тверская область": "69", "Томская область": "70",
        "Тульская область": "71", "Тюменская область": "72", "Удмуртская Республика": "18",
        "Ульяновская область": "73", "Хабаровский край": "27",
        "Ханты-Мансийский автономный округ - Югра": "86", "Челябинская область": "74",
        "Чеченская Республика": "20", "Чувашская Республика": "21",
        "Чукотский автономный округ": "87", "Ямало-Ненецкий автономный округ": "89",
        "Ярославская область": "76",
    })
    SUBJECT_CODE_TO_NAME = {code: name for name, code in SUBJECT_RF_CODES.items()}
    SUBJECT_RF_DISTRICTS = {
        "Центральный федеральный округ": [
            "Белгородская область", "Брянская область", "Владимирская область", "Воронежская область",
            "Ивановская область", "Калужская область", "Костромская область", "Курская область",
            "Липецкая область", "Москва", "Московская область", "Орловская область",
            "Рязанская область", "Смоленская область", "Тамбовская область", "Тверская область",
            "Тульская область", "Ярославская область",
        ],
        "Северо-Западный федеральный округ": [
            "Архангельская область", "Вологодская область", "Калининградская область",
            "Ленинградская область", "Мурманская область", "Ненецкий автономный округ",
            "Новгородская область", "Псковская область", "Республика Карелия", "Республика Коми",
            "Санкт-Петербург",
        ],
        "Южный федеральный округ": [
            "Астраханская область", "Волгоградская область", "Краснодарский край",
            "Республика Адыгея", "Республика Калмыкия", "Республика Крым", "Ростовская область",
            "Севастополь",
        ],
        "Северо-Кавказский федеральный округ": [
            "Кабардино-Балкарская Республика", "Карачаево-Черкесская Республика",
            "Республика Дагестан", "Республика Ингушетия", "Республика Северная Осетия - Алания",
            "Ставропольский край", "Чеченская Республика",
        ],
        "Приволжский федеральный округ": [
            "Кировская область", "Нижегородская область", "Оренбургская область", "Пензенская область",
            "Пермский край", "Республика Башкортостан", "Республика Марий Эл", "Республика Мордовия",
            "Республика Татарстан", "Самарская область", "Саратовская область",
            "Удмуртская Республика", "Ульяновская область", "Чувашская Республика",
        ],
        "Уральский федеральный округ": [
            "Курганская область", "Свердловская область", "Тюменская область",
            "Ханты-Мансийский автономный округ - Югра", "Челябинская область",
            "Ямало-Ненецкий автономный округ",
        ],
        "Сибирский федеральный округ": [
            "Алтайский край", "Иркутская область", "Кемеровская область", "Красноярский край",
            "Новосибирская область", "Омская область", "Республика Алтай", "Республика Тыва",
            "Республика Хакасия", "Томская область",
        ],
        "Дальневосточный федеральный округ": [
            "Амурская область", "Еврейская автономная область", "Забайкальский край", "Камчатский край",
            "Магаданская область", "Приморский край", "Республика Бурятия",
            "Республика Саха (Якутия)", "Сахалинская область", "Хабаровский край",
            "Чукотский автономный округ",
        ],
    }

    _STATUS_MAP = {
        "PUBLISHED": "active",
        "APPLICATIONS_SUBMISSION": "active",
        "APPLICATIONS_ACCEPTANCE": "active",
        "BIDDING": "scheduled",
        "DETERMINING_WINNER": "scheduled",
        "COMPLETED": "closed",
        "CANCELED": "closed",
        "CANCELLED": "closed",
        "FAILED": "closed",
        "ANNULLED": "closed",
        "ARCHIVE": "closed",
    }

    _QUERY_FIELD_MAP = {
        "search_text": "text",
        "type_transaction": "typeTransaction",
        "subject_rf": "dynSubjRF",
        "fias": "fias",
        "ownership_form": "ownershipForm",
        "category_code": "catCode",
        "lot_status": "lotStatus",
        "currency_code": "currCode",
        "publish_date_from": "pubFrom",
        "publish_date_to": "pubTo",
        "bidd_end_time_from": "biddEndFrom",
        "bidd_end_time_to": "biddEndTo",
        "auction_start_date_from": "aucStartFrom",
        "auction_start_date_to": "aucStartTo",
        "notice_number": "noticeNumber",
        "etp_code": "etpCode",
        "bidd_type": "biddType",
        "bidd_form": "biddForm",
        "notice_status": "noticeStatus",
        "organizer_name": "organizerName",
        "organizer_inn": "organizerInn",
        "right_holder_name": "rightHolderName",
        "right_holder_inn": "rightHolderInn",
        "attachment_text": "attachmentText",
    }
    _FIAS_GUID_RE = re.compile(
        r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$"
    )
    _LOCAL_TEXT_LOCATION_TERMS = {
        "москва",
        "санкт петербург",
        "санкт-петербург",
        "севастополь",
        "ярославль",
        "липецк",
        "омск",
        "тула",
        "краснодар",
        "ростов",
        "самара",
        "пермь",
        "киров",
        "псков",
        "екатеринбург",
        "челябинск",
        "тюмень",
        "воронеж",
        "калуга",
        "владимир",
        "иваново",
        "кострома",
        "рязань",
        "тверь",
        "смоленск",
    }

    def __init__(
        self,
        *,
        timeout: tuple[float, float] | float = (10, 30),
        user_agent: str | None = None,
        rate_limit: tuple[float, float] = (0.5, 1.5),
        session: requests.Session | None = None,
        diagnostics: bool = False,
    ):
        self.timeout = timeout
        self.rate_limit = rate_limit
        self.diagnostics = diagnostics
        self._last_request_at: float | None = None
        self.session = session or requests.Session()
        self.session.headers.update({
            "User-Agent": user_agent
            or "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "Chrome/134.0 Safari/537.36 BankrotAI/1.0",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Referer": self.FALLBACK_LIST_URL,
        })

    def search_lots(self, filters: TorgiGovSearchFilters) -> tuple[list[NormalizedLot], dict]:
        warnings: list[str] = []
        params, param_warnings = self._build_query_params(filters)
        warnings.extend(param_warnings)
        raw_endpoint = self._prepare_url(self.SEARCH_ENDPOINT, params)

        try:
            payload = self._request_json(params)
        except requests.RequestException as exc:
            message = self._request_error_message(exc)
            logger.warning("TorgiGov JSON API request failed: %s", exc)
            warnings.append(f"JSON API torgi.gov.ru недоступен: {message}")
            return self._search_lots_html_fallback(filters, warnings)
        except ValueError as exc:
            logger.warning("TorgiGov JSON API returned non-JSON response: %s", exc)
            warnings.append("JSON API torgi.gov.ru вернул ответ не в формате JSON.")
            return self._search_lots_html_fallback(filters, warnings)

        items, total, structure_warning = self._extract_items(payload)
        if structure_warning:
            warnings.append(structure_warning)
        total_pages = self._extract_total_pages(payload)
        page_info = self._extract_page_info(payload)

        lots: list[NormalizedLot] = []
        seen: set[str] = set()
        plain_location_filter = bool(filters.fias and not self._looks_like_fias_identifier(filters.fias))
        duplicates = 0
        duplicate_external_ids: list[str] = []
        skipped_without_id = 0
        normalization_errors = 0
        region_filtered = 0
        location_filtered = 0
        text_filtered = 0
        for item in items:
            if not self._external_base_from_payload(item):
                skipped_without_id += 1
                continue
            try:
                lot = self._normalize_api_lot(item)
            except Exception as exc:
                logger.exception("TorgiGov: failed to normalize lot payload")
                normalization_errors += 1
                warnings.append(f"Один лот пропущен из-за изменения структуры данных: {exc}")
                continue
            if filters.subject_rf and not self._lot_matches_subject(lot, filters.subject_rf):
                region_filtered += 1
                continue
            if plain_location_filter and not self._lot_matches_location_text(lot, filters.fias):
                location_filtered += 1
                continue
            if self._should_apply_local_text_filter(filters.search_text, params) and not self._lot_matches_location_text(lot, filters.search_text):
                text_filtered += 1
                continue
            if lot.external_id in seen:
                duplicates += 1
                duplicate_external_ids.append(lot.external_id)
                continue
            seen.add(lot.external_id)
            lots.append(lot)

        page = max(1, int(filters.page or 1))
        page_size = max(1, min(int(filters.page_size or 100), 100))
        has_more = self._has_more(payload, page, page_size, len(lots), total)
        if skipped_without_id:
            warnings.append(f"{skipped_without_id} лотов пропущено: API не вернул стабильный ID.")
        if normalization_errors:
            warnings.append(f"{normalization_errors} лотов пропущено из-за ошибок нормализации.")
        if duplicates:
            warnings.append(f"{duplicates} лотов на странице были дублями по external_id.")
        if region_filtered:
            warnings.append(f"{region_filtered} лотов скрыто: регион в карточке не совпал с выбранным фильтром.")
        if location_filtered:
            warnings.append(f"{location_filtered} lots hidden by local property-location filter.")
        if text_filtered:
            warnings.append(f"{text_filtered} lots hidden by strict local text filter.")
        meta = {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "has_more": has_more,
            "loaded": len(lots),
            "items_on_page": len(items),
            "duplicates": duplicates,
            "duplicate_external_ids": duplicate_external_ids[:50],
            "skipped_without_id": skipped_without_id,
            "normalization_errors": normalization_errors,
            "region_filtered": region_filtered,
            "location_filtered": location_filtered,
            "text_filtered": text_filtered,
            "page_info": page_info,
            "page_number": page_info.get("number"),
            "page_first": page_info.get("first"),
            "page_last": page_info.get("last"),
            "source": "torgi.gov.ru",
            "raw_endpoint": raw_endpoint,
            "raw_params": dict(params),
            "warnings": warnings,
        }
        if self.diagnostics:
            logger.info(
                "TorgiGov page diagnostics: endpoint=%s params=%s page=%s size=%s total=%s "
                "total_pages=%s items=%s loaded=%s duplicates=%s skipped_without_id=%s page_info=%s",
                self.SEARCH_ENDPOINT,
                params,
                page,
                page_size,
                total,
                total_pages,
                len(items),
                len(lots),
                duplicates,
                skipped_without_id,
                page_info,
            )
        return lots, meta

    def search_all_lots(
        self,
        filters: TorgiGovSearchFilters,
        *,
        max_items: int | None = 5000,
        max_pages: int = 500,
        progress_cb: Callable[[int, int | None, int], None] | None = None,
        page_cb: Callable[[list[NormalizedLot], dict[str, Any]], None] | None = None,
        stop_cb: Callable[[], bool] | None = None,
    ) -> tuple[list[NormalizedLot], dict]:
        all_lots: list[NormalizedLot] = []
        seen: set[str] = set()
        warnings: list[str] = []
        duplicate_external_ids: list[str] = []
        duplicates = 0
        skipped_without_id = 0
        raw_items_loaded = 0
        pages_loaded = 0
        page_diagnostics: list[dict[str, Any]] = []
        stop_reason = "unknown"

        page_size = max(1, min(int(filters.page_size or 100), 100))
        total: int | None = None
        total_pages: int | None = None
        first_meta: dict[str, Any] = {}
        last_meta: dict = {}

        for page in range(1, max_pages + 1):
            if stop_cb and stop_cb():
                stop_reason = "user_stopped"
                break
            page_filters = replace(filters, page=page, page_size=page_size)
            try:
                lots, meta = self.search_lots(page_filters)
            except Exception as exc:
                pages_loaded += 1
                stop_reason = f"page_{page}_error"
                message = f"Страница {page} вернула ошибку: {exc}"
                warnings.append(message)
                logger.exception("TorgiGov all-pages page %s failed", page)
                page_diagnostics.append({
                    "page": page,
                    "error": str(exc),
                    "new_unique": 0,
                    "duplicates": 0,
                    "skipped_without_id": 0,
                })
                break

            pages_loaded += 1
            last_meta = dict(meta or {})
            if not first_meta:
                first_meta = dict(last_meta)

            warnings.extend(last_meta.get("warnings") or [])
            page_items = int(last_meta.get("items_on_page") or len(lots))
            page_duplicates = int(last_meta.get("duplicates") or 0)
            page_skipped = int(last_meta.get("skipped_without_id") or 0)
            page_duplicate_external_ids = list(last_meta.get("duplicate_external_ids") or [])
            raw_items_loaded += page_items
            duplicates += page_duplicates
            skipped_without_id += page_skipped
            duplicate_external_ids.extend(last_meta.get("duplicate_external_ids") or [])

            if total is None and last_meta.get("total") is not None:
                try:
                    total = int(last_meta["total"])
                except Exception:
                    total = None

            if total_pages is None:
                raw_total_pages = last_meta.get("total_pages") or last_meta.get("totalPages")
                if raw_total_pages is not None:
                    try:
                        total_pages = int(raw_total_pages)
                    except Exception:
                        total_pages = None
                elif total is not None:
                    total_pages = max(1, math.ceil(total / page_size))

            new_on_page = 0
            new_lots_on_page: list[NormalizedLot] = []
            cross_page_duplicates = 0
            missing_id_on_page = 0
            for lot in lots:
                if not getattr(lot, "external_id", None):
                    missing_id_on_page += 1
                    skipped_without_id += 1
                    continue
                if lot.external_id in seen:
                    cross_page_duplicates += 1
                    duplicates += 1
                    duplicate_external_ids.append(lot.external_id)
                    page_duplicate_external_ids.append(lot.external_id)
                    continue
                seen.add(lot.external_id)
                all_lots.append(lot)
                new_lots_on_page.append(lot)
                new_on_page += 1

                if max_items is not None and len(all_lots) >= max_items:
                    stop_reason = f"max_items={max_items}"
                    break

            page_diag = {
                "page": page,
                "api_page": page - 1,
                "page_size": page_size,
                "total": last_meta.get("total"),
                "total_pages": last_meta.get("total_pages") or last_meta.get("totalPages"),
                "items_on_page": page_items,
                "normalized_on_page": len(lots),
                "new_unique": new_on_page,
                "duplicates": page_duplicates + cross_page_duplicates,
                "skipped_without_id": page_skipped + missing_id_on_page,
                "duplicate_external_ids": page_duplicate_external_ids[:50],
                "page_info": last_meta.get("page_info") or {},
                "raw_endpoint": last_meta.get("raw_endpoint"),
                "raw_params": last_meta.get("raw_params") or {},
            }
            page_diagnostics.append(page_diag)
            if self.diagnostics:
                logger.info("TorgiGov all-pages page diagnostics: %s", page_diag)

            if progress_cb:
                progress_cb(page, total, len(all_lots))
            if page_cb and new_lots_on_page:
                page_cb(new_lots_on_page, page_diag)

            if max_items is not None and len(all_lots) >= max_items:
                break

            if stop_cb and stop_cb():
                stop_reason = "user_stopped"
                break

            if total_pages is not None:
                if page >= total_pages:
                    stop_reason = "reached_total_pages"
                    break
                continue

            if not lots:
                stop_reason = "empty_page"
                break

            if last_meta.get("has_more") is False and len(lots) < page_size:
                stop_reason = "has_more_false"
                break

        if stop_reason == "unknown":
            stop_reason = "max_pages"

        dedup_warnings = list(dict.fromkeys(warnings))
        if max_items is not None and len(all_lots) >= max_items:
            dedup_warnings.append(f"Загрузка остановлена по лимиту max_items={max_items}.")
        if total is not None and len(all_lots) != total:
            diff = total - len(all_lots)
            dedup_warnings.append(
                f"API сообщил total={total}, но после загрузки получено уникальных лотов={len(all_lots)}. "
                f"Разница={diff}. Дубли={duplicates}, пропущено без ID={skipped_without_id}, "
                f"сырых элементов просмотрено={raw_items_loaded}, страниц загружено={pages_loaded}."
            )
            if diff > 0:
                explained = 0
                if duplicates:
                    explained += min(diff, duplicates)
                    dedup_warnings.append(f"{duplicates} элементов были дублями по external_id.")
                if skipped_without_id:
                    explained += min(max(diff - explained, 0), skipped_without_id)
                    dedup_warnings.append(f"{skipped_without_id} элементов пропущено из-за отсутствия стабильного ID.")
                if diff - explained > 0:
                    dedup_warnings.append(
                        f"API вернул меньше уникальных элементов, чем total; необъясненный остаток={diff - explained}."
                    )
            else:
                dedup_warnings.append(
                    f"Уникальных лотов больше, чем API total; проверьте raw_params и page_diagnostics."
                )
        dedup_warnings = list(dict.fromkeys(dedup_warnings))

        raw_endpoint = first_meta.get("raw_endpoint") or last_meta.get("raw_endpoint")
        raw_params = first_meta.get("raw_params") or last_meta.get("raw_params") or {}
        final_meta = {
            **last_meta,
            "page": 1,
            "page_size": page_size,
            "total": total if total is not None else last_meta.get("total"),
            "total_pages": total_pages,
            "loaded": len(all_lots),
            "raw_items_loaded": raw_items_loaded,
            "duplicates": duplicates,
            "duplicate_external_ids": duplicate_external_ids[:100],
            "skipped_without_id": skipped_without_id,
            "pages_loaded": pages_loaded,
            "stop_reason": stop_reason,
            "has_more": bool(total is not None and len(all_lots) < total),
            "source": "torgi.gov.ru",
            "raw_endpoint": raw_endpoint,
            "raw_params": dict(raw_params),
            "last_raw_endpoint": last_meta.get("raw_endpoint"),
            "last_raw_params": dict(last_meta.get("raw_params") or {}),
            "page_diagnostics": page_diagnostics,
            "warnings": dedup_warnings,
            "mode": "all_pages",
        }
        if self.diagnostics:
            logger.info("TorgiGov all-pages diagnostics: %s", final_meta)
        return all_lots, final_meta

    def enrich_lot_details(self, lot: NormalizedLot) -> NormalizedLot:
        logger.info("TorgiGov enrich_lot_details is reserved for future detail API support: %s", lot.external_id)
        return lot

    def _subject_code(self, subject_rf: str | None) -> str:
        value = self._clean_text(subject_rf)
        return self.SUBJECT_RF_CODES.get(value, value)

    def _subject_name(self, subject_rf: str | None) -> str:
        value = self._clean_text(subject_rf)
        code = self._subject_code(value)
        return self.SUBJECT_CODE_TO_NAME.get(code, value)

    def _lot_matches_subject(self, lot: NormalizedLot, subject_rf: str | None) -> bool:
        expected_code = self._subject_code(subject_rf)
        expected_name = self._subject_name(subject_rf).lower()
        expected_slug = self._region_slug(expected_name)
        text = " ".join(
            part for part in (
                lot.region_name,
                lot.region_slug,
                lot.address,
                lot.title,
                lot.description,
            )
            if part
        ).lower()
        cadastral_prefix = (lot.cadastral_number or "").split(":", 1)[0].zfill(2)

        return (
            bool(expected_name and expected_name in text)
            or bool(expected_slug and expected_slug in text)
            or bool(expected_code and cadastral_prefix == expected_code.zfill(2))
        )

    def _looks_like_fias_identifier(self, value: str | None) -> bool:
        cleaned = self._clean_text(value)
        if not cleaned:
            return False
        return bool(self._FIAS_GUID_RE.match(cleaned))

    def _normalize_filter_text(self, value: Any) -> str:
        return self._clean_text(value).casefold().replace("ё", "е")

    def _lot_searchable_text(self, lot: NormalizedLot) -> str:
        raw = ""
        try:
            raw = json.dumps(lot.raw_data or {}, ensure_ascii=False, default=str)
        except Exception:
            raw = str(lot.raw_data or "")
        return self._normalize_filter_text(
            " ".join(
                part
                for part in (
                    lot.title,
                    lot.description,
                    lot.region_name,
                    lot.region_slug,
                    lot.address,
                    lot.cadastral_number,
                    raw,
                )
                if part
            )
        )

    def _text_contains_query_terms(self, text: str, query: str | None) -> bool:
        normalized_query = self._normalize_filter_text(query)
        if not normalized_query:
            return True
        normalized_text = self._normalize_filter_text(text)
        tokens = re.findall(r"[0-9a-zа-я]+", normalized_query, flags=re.IGNORECASE)
        if not tokens:
            return normalized_query in normalized_text
        for token in tokens:
            if len(token) <= 2:
                if token not in normalized_text:
                    return False
                continue
            if not re.search(rf"(?<![0-9a-zа-я]){re.escape(token)}(?![0-9a-zа-я])", normalized_text):
                return False
        return True

    def _lot_matches_location_text(self, lot: NormalizedLot, query: str | None) -> bool:
        return self._text_contains_query_terms(self._lot_searchable_text(lot), query)

    def _should_apply_local_text_filter(self, search_text: str | None, params: dict[str, str]) -> bool:
        cleaned = self._normalize_filter_text(search_text)
        if not cleaned:
            return False
        if not params.get("text"):
            return True
        compact = " ".join(re.findall(r"[0-9a-zа-я]+", cleaned, flags=re.IGNORECASE))
        return compact in self._LOCAL_TEXT_LOCATION_TERMS

    def _build_query_params(self, filters: TorgiGovSearchFilters) -> tuple[dict[str, str], list[str]]:
        warnings: list[str] = []
        ui_page = max(1, int(filters.page or 1))
        api_page = ui_page - 1
        page_size = max(1, min(int(filters.page_size or 100), 100))
        if filters.page_size and int(filters.page_size) != page_size:
            warnings.append("Размер страницы ограничен диапазоном 1-100.")

        params: dict[str, str] = {
            "byFirstVersion": "true",
            "withFacets": "false",
            "page": str(api_page),
            "size": str(page_size),
            "sort": "firstVersionPublicationDate,desc",
            "matchPhrase": "true" if filters.match_phrase else "false",
        }
        lot_status = filters.lot_status or self.DEFAULT_LOT_STATUS
        if lot_status:
            params["lotStatus"] = str(lot_status).strip()

        for attr, query_name in self._QUERY_FIELD_MAP.items():
            if attr == "lot_status":
                continue
            value = getattr(filters, attr, None)
            if value in (None, "", []):
                continue
            if attr == "fias" and not self._looks_like_fias_identifier(str(value).strip()):
                warnings.append(
                    "Property location is a free-text value; it will be applied locally after loading lots."
                )
                continue
            if attr == "subject_rf":
                value = self._subject_code(str(value).strip())
            if attr == "category_code":
                value = self.CATEGORY_GROUP_CODE_MAP.get(str(value).strip(), value)
            params[query_name] = str(value).strip()

        if filters.price_min is not None:
            params["priceMin"] = self._format_number(filters.price_min)
        if filters.price_max is not None:
            params["priceMax"] = self._format_number(filters.price_max)
        if filters.is_msp:
            params["isMsp"] = "true"

        return params, warnings

    def _request_json(self, params: dict[str, str]) -> Any:
        self._respect_rate_limit()
        response = self.session.get(self.SEARCH_ENDPOINT, params=params, timeout=self.timeout)
        response.raise_for_status()
        return response.json()

    def _build_excel_export_params(self, filters: TorgiGovSearchFilters) -> tuple[dict[str, str], list[str]]:
        params, warnings = self._build_query_params(filters)
        for key in ("page", "size", "withFacets"):
            params.pop(key, None)
        return params, warnings

    def search_lots_excel(self, filters: TorgiGovSearchFilters) -> tuple[list[NormalizedLot], dict]:
        params, warnings = self._build_excel_export_params(filters)
        raw_endpoint = self._prepare_url(self.EXCEL_EXPORT_ENDPOINT, params)
        try:
            self._respect_rate_limit()
            response = self.session.get(self.EXCEL_EXPORT_ENDPOINT, params=params, timeout=self.timeout)
            response.raise_for_status()
        except requests.RequestException as exc:
            message = self._request_error_message(exc)
            logger.warning("TorgiGov Excel export failed: %s", exc)
            raise TorgiGovClientError(
                "Не удалось скачать Excel-выгрузку torgi.gov.ru. "
                f"Проверьте сеть или попробуйте позже. Детали: {message}"
            ) from exc

        lots = self._parse_excel_export(response.content)
        if filters.subject_rf:
            before = len(lots)
            lots = [lot for lot in lots if self._lot_matches_subject(lot, filters.subject_rf)]
            filtered = before - len(lots)
        else:
            filtered = 0
        if filters.fias and not self._looks_like_fias_identifier(filters.fias):
            before = len(lots)
            lots = [lot for lot in lots if self._lot_matches_location_text(lot, filters.fias)]
            location_filtered = before - len(lots)
        else:
            location_filtered = 0
        if self._should_apply_local_text_filter(filters.search_text, params):
            before = len(lots)
            lots = [lot for lot in lots if self._lot_matches_location_text(lot, filters.search_text)]
            text_filtered = before - len(lots)
        else:
            text_filtered = 0
        if location_filtered:
            warnings.append(f"{location_filtered} lots hidden by local property-location filter.")
        if text_filtered:
            warnings.append(f"{text_filtered} lots hidden by strict local text filter.")
        return lots, {
            "page": 1,
            "page_size": len(lots),
            "total": len(lots),
            "total_pages": 1,
            "loaded": len(lots),
            "duplicates": 0,
            "skipped_without_id": 0,
            "pages_loaded": 1,
            "stop_reason": "excel_export",
            "has_more": False,
            "source": "torgi.gov.ru Excel",
            "raw_endpoint": raw_endpoint,
            "raw_params": dict(params),
            "region_filtered": filtered,
            "location_filtered": location_filtered,
            "text_filtered": text_filtered,
            "warnings": warnings,
            "mode": "excel",
        }

    def _parse_excel_export(self, content: bytes) -> list[NormalizedLot]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            next(rows)
            header_row = next(rows)
        except StopIteration:
            return []

        headers = [self._clean_text(value) for value in header_row]
        lots: list[NormalizedLot] = []
        seen: set[str] = set()
        for values in rows:
            row = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }
            lot = self._normalize_excel_row(row)
            if not lot or lot.external_id in seen:
                continue
            seen.add(lot.external_id)
            lots.append(lot)
        return lots

    def _normalize_excel_row(self, row: dict[str, Any]) -> NormalizedLot | None:
        def get(name: str) -> Any:
            return row.get(name)

        lot_url = self._clean_text(get("Ссылка на лот в ОЧ Реестра лотов"))
        notice_url = self._clean_text(get("Ссылка на извещение в ОЧ Реестра извещений"))
        notice_number = self._clean_text(get("Номер извещения"))
        lot_number = self._clean_text(get("Номер лота"))
        external_base = self._extract_lot_id_from_url(lot_url) if lot_url else None
        if not external_base:
            external_base = self._join_id_parts(notice_number, lot_number)
        if not external_base:
            return None

        title = self._clean_text(get("Предмет торгов") or get("Описание лота"))
        description = self._clean_text(get("Описание лота") or title)
        category_name = self._clean_text(get("Категория имущества"))
        characteristics = self._clean_text(get("Характеристики имущества"))
        attributes = self._clean_text(get("Дополнительные атрибуты"))
        address = self._clean_text(get("Местонахождение имущества")) or extract_address(description)
        region_name = self._clean_text(get("Субъект РФ")) or self._extract_region_from_text(" ".join([address, description]))
        status = self._clean_text(get("Статус лота") or get("Статус извещения"))
        start_price = self._coerce_price(get("Начальная цена"))
        final_price = self._coerce_price(get("Итоговая цена"))
        combined_text = " ".join(
            part for part in (title, description, address, category_name, characteristics, attributes) if part
        )
        cadastral_numbers = extract_cadastral_numbers(combined_text)
        published_at = self._parse_date(get("Дата публикации") or get("Дата и время начала подачи заявок"))
        bidd_end = self._clean_text(get("Дата и время окончания подачи заявок"))
        auction_start = self._clean_text(get("Дата и время начала аукциона"))
        raw_data = {
            "source": "torgi_gov_excel",
            "notice_number": notice_number,
            "lot_number": lot_number,
            "notice_url": notice_url,
            "trade_type": self._clean_text(get("Вид торгов")),
            "bidding_form": self._clean_text(get("Форма проведения")),
            "etp": self._clean_text(get("Электронная площадка")),
            "deposit": self._coerce_price(get("Размер задатка")),
            "auction_step": self._coerce_price(get("Шаг аукциона")),
            "bidd_end_time": bidd_end,
            "auction_start_date": auction_start,
            "category_display": category_name,
            "characteristics": characteristics,
            "additional_attributes": attributes,
            "cadastral_numbers": cadastral_numbers,
            "raw": row,
        }

        return NormalizedLot(
            external_id=f"torgi_gov:{external_base}",
            source="torgi_gov",
            source_system="torgi.gov.ru",
            title=(title or f"Лот torgi.gov.ru {external_base}")[:500],
            description=(description or title or "")[:5000],
            category=self._category_from_payload("", category_name, title, combined_text),
            region_slug=self._region_slug(region_name),
            region_name=region_name or None,
            address=address or None,
            cadastral_number=cadastral_numbers[0] if cadastral_numbers else None,
            vin=None,
            area=extract_area(combined_text) or extract_land_area(combined_text),
            start_price=start_price,
            current_price=final_price or start_price,
            auction_status=self._normalize_status(status),
            lot_url=lot_url or None,
            source_url=lot_url or notice_url or None,
            detail_level="search_excel",
            raw_data=raw_data,
            published_at=published_at,
            land_area=extract_land_area(combined_text),
            total_area_gba=extract_building_area(combined_text) or extract_room_area(combined_text),
        )

    def _search_lots_html_fallback(
        self,
        filters: TorgiGovSearchFilters,
        warnings: list[str],
    ) -> tuple[list[NormalizedLot], dict]:
        params, _ = self._build_query_params(filters)
        raw_endpoint = self._prepare_url(self.FALLBACK_LIST_URL, params)
        warnings.append("Использован HTML fallback; основной источник должен быть JSON API.")
        try:
            self._respect_rate_limit()
            response = self.session.get(self.FALLBACK_LIST_URL, params=params, timeout=self.timeout)
            response.raise_for_status()
        except requests.RequestException as exc:
            message = self._request_error_message(exc)
            logger.warning("TorgiGov HTML fallback failed: %s", exc)
            raise TorgiGovClientError(
                "Не удалось подключиться к torgi.gov.ru. "
                f"Проверьте сеть или попробуйте позже. Детали: {message}"
            ) from exc

        lots = self._parse_fallback_html(response.text)
        if not lots:
            warnings.append("HTML fallback не нашел карточек лотов; возможно, сайт отдает только SPA без данных.")
        if filters.fias and not self._looks_like_fias_identifier(filters.fias):
            before = len(lots)
            lots = [lot for lot in lots if self._lot_matches_location_text(lot, filters.fias)]
            location_filtered = before - len(lots)
        else:
            location_filtered = 0
        if self._should_apply_local_text_filter(filters.search_text, params):
            before = len(lots)
            lots = [lot for lot in lots if self._lot_matches_location_text(lot, filters.search_text)]
            text_filtered = before - len(lots)
        else:
            text_filtered = 0
        if location_filtered:
            warnings.append(f"{location_filtered} lots hidden by local property-location filter.")
        if text_filtered:
            warnings.append(f"{text_filtered} lots hidden by strict local text filter.")
        page = max(1, int(filters.page or 1))
        page_size = max(1, min(int(filters.page_size or 100), 100))
        return lots, {
            "page": page,
            "page_size": page_size,
            "total": None,
            "total_pages": None,
            "loaded": len(lots),
            "duplicates": 0,
            "skipped_without_id": 0,
            "pages_loaded": 1,
            "stop_reason": "html_fallback",
            "location_filtered": location_filtered,
            "text_filtered": text_filtered,
            "has_more": False,
            "source": "torgi.gov.ru",
            "raw_endpoint": raw_endpoint,
            "raw_params": dict(params),
            "warnings": warnings,
        }

    def _parse_fallback_html(self, html: str) -> list[NormalizedLot]:
        soup = BeautifulSoup(html, "lxml")
        lots: list[NormalizedLot] = []
        seen: set[str] = set()
        for link in soup.find_all("a", href=True):
            href = link.get("href") or ""
            if "/new/public/lots/lot/" not in href:
                continue
            title = self._clean_text(link.get_text(" ", strip=True))
            if not title or len(title) < 20:
                continue
            lot_id = self._extract_lot_id_from_url(href)
            if not lot_id or lot_id in seen:
                continue
            seen.add(lot_id)
            url = urljoin(self.BASE_URL, href)
            text = self._clean_text(link.parent.get_text(" ", strip=True) if link.parent else title)
            cadastral_numbers = extract_cadastral_numbers(text)
            region_name = self._extract_region_from_text(text)
            raw = {
                "source": "html_fallback",
                "title": title,
                "url": url,
                "text": text[:5000],
                "cadastral_numbers": cadastral_numbers,
            }
            lots.append(NormalizedLot(
                external_id=f"torgi_gov:{lot_id}",
                source="torgi_gov",
                source_system="torgi.gov.ru",
                title=title[:500],
                description=text[:5000],
                category=classify_category(title, text),
                region_slug=self._region_slug(region_name),
                region_name=region_name,
                address=extract_address(text),
                cadastral_number=cadastral_numbers[0] if cadastral_numbers else None,
                vin=None,
                area=extract_area(text) or extract_land_area(text),
                start_price=extract_price(text),
                current_price=extract_price(text),
                auction_status=self._normalize_status(text),
                lot_url=url,
                source_url=url,
                detail_level="search",
                raw_data=raw,
            ))
        return lots

    def _normalize_api_lot(self, payload: dict[str, Any]) -> NormalizedLot:
        notice_number = self._pick(payload, "noticeNumber", "noticeNo", "noticeId", "noticeRegistryNumber")
        lot_number = self._pick(payload, "lotNumber", "lotNum", "number", "lotNo")
        external_base = self._external_base_from_payload(payload)
        if not external_base:
            external_base = hashlib.sha256(
                json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
            ).hexdigest()[:20]
        external_id = f"torgi_gov:{external_base}"

        title = self._clean_text(self._pick(
            payload,
            "lotName", "name", "title", "lotTitle", "subject", "objectName", "noticeName",
        ))
        description = self._clean_text(self._pick(
            payload,
            "lotDescription", "description", "fullDescription", "objectInfo", "assetDescription", "comment",
        ))
        if not title:
            title = description[:200] if description else f"Лот torgi.gov.ru {external_base}"
        if not description:
            description = title

        category_code = self._clean_text(self._pick(payload, "categoryCode", "catCode", "category.code"))
        category_name = self._clean_text(self._pick(
            payload,
            "categoryName", "category.name", "category", "lotCategory", "lotCategoryName",
        ))
        category = self._category_from_payload(category_code, category_name, title, description)
        region_name = self._clean_text(self._pick(
            payload,
            "subjectRFName", "subjectName", "subjectRF", "subject.label", "regionName",
            "rfSubject", "locationRegion", "lotAddress.subjectName",
        )) or self._extract_region_from_text(description)
        address = self._clean_text(self._pick(
            payload,
            "address", "location", "locationAddress", "objectAddress", "estateAddress",
            "fiasAddress", "lotAddress", "addr",
        )) or extract_address(description)

        combined_text = " ".join(part for part in (title, description, address or "", region_name or "") if part)
        cadastral_numbers = self._extract_cadastral_numbers(payload, combined_text)
        start_price = self._pick_price(payload, "priceMin", "initialPrice", "startPrice", "price", "lotPrice")
        current_price = self._pick_price(payload, "currentPrice", "priceFin", "finalPrice", "actualPrice") or start_price
        raw_status = self._clean_text(self._pick(
            payload,
            "lotStatus", "status", "statusCode", "lotStatusName", "statusName", "state",
        ))
        published_text = self._pick(payload, "firstVersionPublicationDate", "publishDate", "publicationDate", "publishedAt")
        bidd_end_text = self._pick(payload, "biddEndTime", "biddEndDate", "applicationEndDate", "applicationsEndDate")
        auction_start_text = self._pick(payload, "auctionStartDate", "biddingDate", "tradeDate")
        lot_url = self._clean_text(self._pick(payload, "lotUrl", "url", "href", "link"))
        if not lot_url:
            lot_url = f"{self.BASE_URL}/new/public/lots/lot/{external_base}/(lotInfo:info)?fromRec=false"
        else:
            lot_url = urljoin(self.BASE_URL, lot_url)

        raw_data = {
            "source": "torgi_gov_json",
            "status": raw_status,
            "dates": [
                {"title": "Дата публикации", "text": self._clean_text(published_text)},
                {"title": "Окончание подачи заявок", "text": self._clean_text(bidd_end_text)},
                {"title": "Дата проведения торгов", "text": self._clean_text(auction_start_text)},
            ],
            "notice_number": self._clean_text(notice_number),
            "lot_number": self._clean_text(lot_number),
            "category_code": category_code,
            "category_display": category_name or self.CATEGORY_CODE_LABELS.get(str(category_code), ""),
            "cadastral_numbers": cadastral_numbers,
            "bidd_end_time": self._clean_text(bidd_end_text),
            "auction_start_date": self._clean_text(auction_start_text),
            "raw": payload,
        }

        return NormalizedLot(
            external_id=external_id,
            source="torgi_gov",
            source_system="torgi.gov.ru",
            title=title[:500],
            description=description[:5000],
            category=category,
            region_slug=self._region_slug(region_name),
            region_name=region_name or None,
            address=address or None,
            cadastral_number=cadastral_numbers[0] if cadastral_numbers else None,
            vin=None,
            area=self._pick_area(payload, combined_text),
            start_price=start_price,
            current_price=current_price,
            auction_status=self._normalize_status(raw_status),
            lot_url=lot_url,
            source_url=lot_url,
            detail_level="search",
            raw_data=raw_data,
            published_at=self._parse_date(published_text),
            land_area=extract_land_area(combined_text),
            total_area_gba=extract_building_area(combined_text) or extract_room_area(combined_text),
        )

    def _extract_items(self, payload: Any) -> tuple[list[dict[str, Any]], int | None, str | None]:
        if isinstance(payload, list):
            return [item for item in payload if isinstance(item, dict)], None, None
        if not isinstance(payload, dict):
            return [], None, "JSON API вернул неожиданный тип ответа."

        for key in ("content", "items", "data", "results", "lotCards", "lots", "records"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)], self._extract_total(payload), None
            if isinstance(value, dict):
                nested_items, nested_total, warning = self._extract_items(value)
                if nested_items:
                    return nested_items, nested_total or self._extract_total(payload), warning

        return [], self._extract_total(payload), "JSON API доступен, но список лотов в ответе не найден."

    def _extract_total(self, payload: dict[str, Any]) -> int | None:
        for key in ("totalElements", "total", "totalCount", "total_count", "count"):
            value = payload.get(key)
            if value is None:
                continue
            try:
                return int(value)
            except (TypeError, ValueError):
                continue
        page = payload.get("page")
        if isinstance(page, dict):
            return self._extract_total(page)
        return None

    def _extract_total_pages(self, payload: Any) -> int | None:
        if not isinstance(payload, dict):
            return None
        for key in ("totalPages", "total_pages", "pages", "pageCount"):
            value = payload.get(key)
            if value is None:
                continue
            try:
                return int(value)
            except (TypeError, ValueError):
                continue
        page = payload.get("page")
        if isinstance(page, dict):
            return self._extract_total_pages(page)
        return None

    def _extract_page_info(self, payload: Any) -> dict[str, Any]:
        if not isinstance(payload, dict):
            return {}
        source = payload.get("page") if isinstance(payload.get("page"), dict) else payload
        info: dict[str, Any] = {}
        for key in ("number", "size", "last", "first", "totalPages", "totalElements"):
            if isinstance(source, dict) and key in source:
                info[key] = source.get(key)
            elif key in payload:
                info[key] = payload.get(key)
        return info

    def _external_base_from_payload(self, payload: dict[str, Any]) -> str | None:
        raw_id = self._pick(payload, "id", "lotCardId", "lotId", "lotIdentifier", "lotGuid", "commonId")
        notice_number = self._pick(payload, "noticeNumber", "noticeNo", "noticeId", "noticeRegistryNumber")
        lot_number = self._pick(payload, "lotNumber", "lotNum", "number", "lotNo")
        external_base = raw_id or self._join_id_parts(notice_number, lot_number)
        cleaned = self._clean_text(external_base)
        return cleaned or None

    def _has_more(self, payload: Any, page: int, page_size: int, count: int, total: int | None) -> bool:
        if isinstance(payload, dict):
            if "last" in payload:
                return not bool(payload.get("last"))
            if "hasNext" in payload:
                return bool(payload.get("hasNext"))
        if total is not None:
            return page * page_size < total
        return count >= page_size

    def _pick(self, payload: Any, *keys: str) -> Any:
        for key in keys:
            value = self._find_key(payload, key)
            if value not in (None, "", [], {}):
                if isinstance(value, dict):
                    nested = self._pick(value, "name", "label", "value", "title", "fullName", "address")
                    return nested if nested not in (None, "") else value
                if isinstance(value, list):
                    return ", ".join(self._clean_text(v) for v in value if self._clean_text(v))
                return value
        return None

    def _find_key(self, payload: Any, key: str) -> Any:
        if not isinstance(payload, (dict, list)):
            return None
        if "." in key:
            current: Any = payload
            for part in key.split("."):
                if not isinstance(current, dict):
                    return None
                current = self._dict_get_case_insensitive(current, part)
                if current is None:
                    return None
            return current
        if isinstance(payload, dict):
            direct = self._dict_get_case_insensitive(payload, key)
            if direct not in (None, "", [], {}):
                return direct
            for value in payload.values():
                found = self._find_key(value, key)
                if found not in (None, "", [], {}):
                    return found
        else:
            for value in payload:
                found = self._find_key(value, key)
                if found not in (None, "", [], {}):
                    return found
        return None

    def _dict_get_case_insensitive(self, payload: dict[str, Any], key: str) -> Any:
        if key in payload:
            return payload[key]
        key_lower = key.lower()
        for item_key, value in payload.items():
            if str(item_key).lower() == key_lower:
                return value
        return None

    def _pick_price(self, payload: dict[str, Any], *keys: str) -> float | None:
        for key in keys:
            value = self._find_key(payload, key)
            price = self._coerce_price(value)
            if price is not None:
                return price
        return None

    def _coerce_price(self, value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, dict):
            for key in ("amount", "value", "price", "sum", "rub", "number"):
                price = self._coerce_price(value.get(key))
                if price is not None:
                    return price
            return None
        if isinstance(value, list):
            for item in value:
                price = self._coerce_price(item)
                if price is not None:
                    return price
            return None
        return parse_money(str(value))

    def _pick_area(self, payload: dict[str, Any], text: str) -> float | None:
        for key in ("area", "square", "totalArea", "landArea", "objectArea"):
            value = self._find_key(payload, key)
            if isinstance(value, (int, float)):
                return float(value)
            parsed = extract_area(str(value)) if value not in (None, "") else None
            if parsed:
                return parsed
        return extract_area(text) or extract_land_area(text)

    def _extract_cadastral_numbers(self, payload: dict[str, Any], text: str) -> list[str]:
        values = self._pick(payload, "cadastralNumbers", "cadastralNumber", "cadNumber", "cadNumbers")
        found: list[str] = []
        if values:
            if isinstance(values, str):
                found.extend(extract_cadastral_numbers(values))
            elif isinstance(values, list):
                for item in values:
                    found.extend(extract_cadastral_numbers(str(item)))
        found.extend(extract_cadastral_numbers(text))
        unique: list[str] = []
        for number in found:
            if number not in unique:
                unique.append(number)
        return unique

    def _category_from_payload(self, code: str, name: str, title: str, description: str) -> str:
        code = str(code or "").strip()
        if code in self.CATEGORY_CODE_TO_INTERNAL:
            return self.CATEGORY_CODE_TO_INTERNAL[code]
        name_lower = (name or "").lower()
        if "транспорт" in name_lower:
            return "transport"
        if "земель" in name_lower and "здани" in name_lower:
            return "commercial_building_with_land"
        if "земель" in name_lower:
            return "land"
        if "недвиж" in name_lower:
            return "real_estate"
        return classify_category(title, description)

    def _normalize_status(self, status: str | None) -> str:
        text = self._clean_text(status).lower()
        if not text:
            return "unknown"
        for raw, normalized in self._STATUS_MAP.items():
            if raw.lower() in text:
                return normalized
        if any(word in text for word in ("опублик", "прием", "приём", "заяв")):
            return "active"
        if any(word in text for word in ("проведен", "заверш", "состоял", "отмен", "аннулир")):
            return "closed"
        if any(word in text for word in ("торг", "ожида", "назнач")):
            return "scheduled"
        return "unknown"

    def _parse_date(self, value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value
        text = self._clean_text(value)
        if not text:
            return None
        normalized = text.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(normalized).replace(tzinfo=None)
        except ValueError:
            pass
        for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text[:19], fmt)
            except ValueError:
                continue
        return None

    def _region_slug(self, region_name: str | None) -> str:
        if not region_name:
            return "online"
        translit = {
            "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
            "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
            "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
            "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch",
            "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
        }
        slug = "".join(translit.get(ch, ch) for ch in region_name.lower())
        slug = re.sub(r"[^a-z0-9]+", "-", slug).strip("-")
        return slug[:100] or "online"

    def _extract_region_from_text(self, text: str) -> str | None:
        match = re.search(r"([А-ЯЁ][а-яё-]+(?:\s+[А-ЯЁа-яё-]+){0,2}\s+(?:область|край|республика|округ))", text)
        return match.group(1) if match else None

    def _extract_lot_id_from_url(self, url: str) -> str | None:
        match = re.search(r"/lot/([^/?#]+)", url)
        return match.group(1) if match else None

    def _join_id_parts(self, *parts: Any) -> str | None:
        clean = [self._clean_text(part) for part in parts if self._clean_text(part)]
        return "_".join(clean) if clean else None

    def _clean_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, dict):
            value = self._pick(value, "name", "label", "value", "title", "address", "fullAddress") or ""
        if isinstance(value, list):
            value = " ".join(self._clean_text(item) for item in value)
        text = BeautifulSoup(str(value), "lxml").get_text(" ", strip=True) if "<" in str(value) else str(value)
        return re.sub(r"\s+", " ", text).strip()

    def _format_number(self, value: float) -> str:
        return str(int(value)) if float(value).is_integer() else str(value)

    def _prepare_url(self, url: str, params: dict[str, str]) -> str:
        request = requests.Request("GET", url, params=params).prepare()
        return str(request.url)

    def _respect_rate_limit(self) -> None:
        if self._last_request_at is not None:
            elapsed = time.monotonic() - self._last_request_at
            delay = random.uniform(*self.rate_limit)
            if elapsed < delay:
                time.sleep(delay - elapsed)
        self._last_request_at = time.monotonic()

    def _request_error_message(self, exc: requests.RequestException) -> str:
        if isinstance(exc, (requests.Timeout, requests.ConnectTimeout, requests.ReadTimeout)):
            return "превышено время ожидания ответа"
        response = getattr(exc, "response", None)
        if response is not None:
            return f"HTTP {response.status_code}"
        return str(exc)

class ManualHtmlParser: 
    def __init__(self, config: ParserConfig | None = None): 
        self.config = config or ParserConfig() 
 
    def parse_file(self, file_path: str) -> list[ParsedLotData]: 
        """Надёжное чтение с приоритетом Windows-1251""" 
        encodings = ['cp1251', 'windows-1251', 'utf-8', 'latin1'] 
        html = None 
        
        for enc in encodings: 
            try: 
                with open(file_path, "r", encoding=enc) as f: 
                    html = f.read() 
                logger.info(f"✅ Файл успешно прочитан в кодировке {enc}") 
                break 
            except UnicodeDecodeError: 
                continue 
        
        if html is None: 
            with open(file_path, "r", encoding="utf-8", errors="replace") as f: 
                html = f.read() 
            logger.warning("⚠️ Файл прочитан с заменой символов (utf-8 errors=replace)") 
        
        return self.parse_html(html) 
 
    def parse_html(self, html: str) -> list[ParsedLotData]: 
        soup = BeautifulSoup(html, "lxml") 
 
        containers = soup.select("div.lot_container") 
 
        # fallback для другой верстки, но только если lot_container не найден 
        if not containers: 
            containers = soup.select("div.lot[data-id]") 
 
        logger.info(f"Найдено контейнеров лотов: {len(containers)}") 
 
        lots = [] 
        seen = set() 
 
        for c in containers: 
            lot = self._extract_lot(c) 
            if not lot: 
                continue 
 
            if lot.external_id in seen: 
                continue 
 
            seen.add(lot.external_id) 
            lots.append(lot) 
 
        return lots 
 
    def _extract_lot(self, container: Any) -> ParsedLotData | None: 
        # Главные селекторы TBankrot из сохранённой выдачи 
        lot_div = container.select_one("div.lot[data-id]") 
        title_tag = (
            container.select_one("p.lot_title a")
            or container.select_one(".lot_title a:not(.lot_num):not(.link_new_tab)")
        ) 
        num_tag = container.select_one("a.lot_num") 
        desc_tag = container.select_one(".lot_description .text") 
        price_tag = container.select_one(".lot_prices .current_price span") 
        created_tag = container.select_one(".lot_created") 
        status_icon = container.select_one(".status_icon") 
 
        # Без p.lot_title a это, скорее всего, не карточка лота 
        if not (title_tag or num_tag or lot_div): 
            return None 
 
        # ID лота 
        external_id = None 
 
        if lot_div and lot_div.get("data-id"): 
            external_id = lot_div.get("data-id") 
 
        if not external_id and num_tag: 
            num_text = num_tag.get_text(" ", strip=True) 
            m = re.search(r'\d+', num_text) 
            if m: 
                external_id = m.group(0) 
 
        if title_tag and title_tag.get("href"):
            href = title_tag.get("href", "")
        elif num_tag and num_tag.get("href"):
            href = num_tag.get("href", "")
        else:
            link_tag = container.select_one("a.link_new_tab[href], a[href*='item?id']")
            href = link_tag.get("href", "") if link_tag else ""
 
        if not external_id: 
            m = re.search(r'id=(\d+)', href) 
            if m: 
                external_id = m.group(1) 
 
        if not href and not external_id:
            return None

        if not external_id: 
            external_id = f"tb_{hashlib.md5(href.encode()).hexdigest()[:12]}" 
 
        # URL и название 
        url = urljoin(self.config.base_url, href) 
        title = title_tag.get_text(" ", strip=True) if title_tag else "" 
 
        # Описание — только текст описания, без цены/дат/кнопок 
        description = desc_tag.get_text("\n", strip=True) if desc_tag else "" 
        if not title:
            compact = re.sub(r"\s+", " ", description or container.get_text(" ", strip=True)).strip()
            title = compact[:180] if compact else f"Лот {external_id}"
 
        # Цена — только current_price, не minimal_price 
        price_text = price_tag.get_text(" ", strip=True) if price_tag else "" 
        current_price = parse_money(price_text) 
 
        # Даты 
        dates = [] 
        for d in container.select(".inline_dates .date"): 
            dates.append({ 
                "title": d.get("title") or "", 
                "text": d.get_text(" ", strip=True), 
            }) 
 
        # Блоки типа шаг аукциона / задаток / минимальная цена 
        minimal_blocks = [] 
        for block in container.select(".minimal_price"): 
            label = block.select_one(".small") 
            value = block.select_one(".green-color, .semibold") 
            minimal_blocks.append({ 
                "label": label.get_text(" ", strip=True) if label else "", 
                "text": block.get_text(" ", strip=True), 
                "value": value.get_text(" ", strip=True) if value else "", 
            }) 
 
        # Категории из иконок 
        category_titles = [ 
            img.get("title") 
            for img in container.select(".category_icons img") 
            if img.get("title") 
        ] 
 
        raw_payload = { 
            "source": "manual_html_tbankrot", 
            "price_text": price_text, 
            "dates": dates, 
            "minimal_blocks": minimal_blocks, 
            "created_text": created_tag.get_text(" ", strip=True) if created_tag else "", 
            "status_icon_class": " ".join(status_icon.get("class", [])) if status_icon else "", 
            "category_titles": category_titles, 
            "title": title, 
            "url": url, 
        } 
 
        cadastral_numbers = extract_cadastral_numbers(description)
        address = extract_address(description) or ""
        building_area = extract_building_area(description)
        room_area = extract_room_area(description)
        land_area = extract_land_area(description)

        main_area = building_area or room_area or land_area
        floors = extract_floors(description)
        year_built = extract_year_built(description)
        year_commissioning = extract_commissioning_year(description)
        is_cultural_heritage = extract_cultural_heritage(description)

        return ParsedLotData(
            external_id=str(external_id),
            title=title,
            url=url,
            price_text=price_text,
            current_price=current_price,
            description=description[:5000],
            status=extract_status_from_raw_payload(raw_payload),
            address=address,
            cadastral_number=cadastral_numbers[0] if cadastral_numbers else "",
            cadastral_numbers=cadastral_numbers,
            area=main_area,
            building_area=building_area,
            room_area=room_area,
            land_area=land_area,
            floors=floors,
            raw_payload=raw_payload,
        ) 


def extract_status_from_raw_payload(raw_payload: dict[str, Any]) -> str: 
    dates = raw_payload.get("dates") or [] 
    date_text = " ".join( 
        f"{d.get('title', '')} {d.get('text', '')}" 
        for d in dates 
    ).lower() 
 
    icon_class = str(raw_payload.get("status_icon_class") or "").lower() 
 
    if any(x in date_text for x in [ 
        "прием заявок до", 
        "приём заявок до", 
        "осталось", 
        "до окончания" 
    ]): 
        return "active" 
 
    if "начало торгов" in date_text: 
        return "scheduled" 
 
    if any(x in date_text for x in [ 
        "заверш", 
        "не состоял", 
        "состоял", 
        "прием заявок завершен", 
        "приём заявок завершен" 
    ]): 
        return "closed" 
 
    if "green" in icon_class: 
        return "active" 
 
    return "unknown" 


def import_manual_html(
    session: Session,
    file_path: str,
    city_slug: str = "yaroslavl",
    config: ParserConfig | None = None,
    progress_cb: Optional[Callable[[int, int], None]] = None,
    skip_geo: bool = True,          # По умолчанию ускоренный режим
) -> tuple[int, int, int]:
    from bankrotai.db import ProcessedLot, select
    parser = ManualHtmlParser(config=config)
    parsed_lots = parser.parse_file(file_path)
    
    # Быстрая дедупликация
    seen = set()
    unique_lots = [p for p in parsed_lots if p.external_id not in seen and not seen.add(p.external_id)]
    
    new_cnt = upd_cnt = skip_cnt = 0
    total = len(unique_lots)
    
    # Один запрос — получаем все существующие лоты
    external_ids = [str(p.external_id) for p in unique_lots]
    existing_map = {}
    if external_ids:
        from sqlalchemy import select
        stmt = select(ProcessedLot.external_id, ProcessedLot.review_status).where(
            ProcessedLot.external_id.in_(external_ids)
        )
        for eid, review_status in session.execute(stmt):
            existing_map[eid] = review_status

    for i, p in enumerate(unique_lots, 1):
        ext_id = str(p.external_id)
        existed_before = ext_id in existing_map
        existing_review = existing_map.get(ext_id)
        
        normalized = _normalize_manual_lot(city_slug, p)
        processed = persist_lot(session, normalized)
        
        if not skip_geo:
            upsert_lot_events_from_raw(session, processed, normalized.raw_data)
        else:
            # Быстрый режим — только статус
            processed.auction_status = normalized.auction_status or "active"
        
        if not existed_before:
            new_cnt += 1
        elif existing_review is not None:
            skip_cnt += 1
        else:
            upd_cnt += 1
            
        if progress_cb and i % 5 == 0:
            progress_cb(i, total)
    
    session.flush()
    return new_cnt, upd_cnt, skip_cnt

def _normalize_manual_lot(city_slug: str, p: ParsedLotData) -> NormalizedLot:
    title = p.title.strip()

    return NormalizedLot(
        external_id=p.external_id,
        source="manual_html",
        source_system="manual_html",
        title=title,
        description=p.description[:5000],
        category=classify_category(title, p.description),
        region_slug=city_slug,
        region_name=None,
        address=p.address,
        cadastral_number=p.cadastral_number,
        vin=None,

        # Старое общее поле — для совместимости
        area=p.area,

        start_price=None,
        current_price=p.current_price,
        auction_status=p.status or "unknown",
        lot_url=p.url,
        source_url=p.url,
        detail_level="detail",

        raw_data={
            **p.raw_payload,
            "cadastral_numbers": p.cadastral_numbers,
            "building_area": p.building_area,
            "room_area": p.room_area,
            "land_area": p.land_area,
            "year_built": p.year_built,
            "year_commissioning": p.year_commissioning,
            "is_cultural_heritage": p.is_cultural_heritage,
        },

        # Новые отдельные поля
        total_area_gba=p.building_area or p.room_area,
        land_area=p.land_area,
        floors=p.floors,
        year_built=p.year_built,
        legal_status=extract_legal_status(p.description),
    ) 


# --- Public Sources ---

DEFAULT_CATEGORY_SLUGS = ("realizaciya-imuschestva-dolzhnikov",)

@dataclass
class PublicRealEstateLot:
    source: str
    category: str
    published_at: str
    asset_type: str
    status: str
    price: str
    title: str
    location: str
    url: str
    reference_url: str
    source_label: str
    tbankrot_status: str | None = None
    tbankrot_status_note: str | None = None
    tbankrot_status_checked_at: str | None = None

def _normalize_address_for_search(address: str, region_slug: str) -> str:
    return address

def _parse_public_date(date_text: str) -> datetime | None:
    # Очистка и замена русских месяцев
    # Сначала удаляем запятые, чтобы они не мешали парсингу
    clean_text = date_text.replace(",", "").lower().strip()
    
    months = {
        "января": "january", "февраля": "february", "марта": "march",
        "апреля": "april", "мая": "may", "июня": "june",
        "июля": "july", "августа": "august", "сентября": "september",
        "октября": "october", "ноября": "november", "декабря": "december"
    }
    for ru, en in months.items():
        clean_text = clean_text.replace(ru, en)
    
    # Пытаемся разные форматы
    for fmt in ("%d %B %Y", "%d %b %Y", "%d %m %Y"):
        try:
            return datetime.strptime(clean_text, fmt)
        except ValueError:
            continue
            
    # Если английский формат (напр. "16 April 2026")
    try:
        return datetime.strptime(clean_text, "%d %B %Y")
    except ValueError:
        pass
        
    return None

def _public_lot_from_payload(payload: dict) -> PublicRealEstateLot:
    return PublicRealEstateLot(
        source=payload.get("source", ""),
        category=payload.get("categorySlug", ""),
        published_at=payload.get("published_at_text", ""),
        asset_type=payload.get("asset_type", ""),
        status=payload.get("status", ""),
        price=payload.get("price_text", ""),
        title=payload.get("title", ""),
        location=payload.get("location", ""),
        url=payload.get("url", ""),
        reference_url=payload.get("reference_url", ""),
        source_label=payload.get("sourceLabel", ""),
        tbankrot_status=payload.get("tbankrot_status"),
        tbankrot_status_note=payload.get("tbankrot_status_note"),
        tbankrot_status_checked_at=payload.get("tbankrot_status_checked_at")
    )

class GorodTorgiClient:
    def __init__(self, city_slug: str, **kwargs):
        self.city_slug = city_slug
        self.enrich_details = bool(kwargs.get("enrich_details", True))
        self.resolve_tbankrot = bool(kwargs.get("resolve_tbankrot", True))
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/134.0 Safari/537.36"})

    def _build_tbankrot_query(self, lot: PublicRealEstateLot) -> str:
        return f"{lot.title} {lot.location}"

    def _get_html(self, url: str) -> str:
        resp = self.session.get(url, timeout=20)
        resp.raise_for_status()
        return resp.text

    def _iter_category_pages(self, category_slug: str) -> list[str]:
        base_url = f"https://{self.city_slug}.gorod-torgi.ru/cat/{category_slug}"
        html = self._get_html(base_url)
        soup = BeautifulSoup(html, "lxml")
        
        # Находим ссылки пагинации
        pages = [base_url]
        last_page = 1
        for link in soup.find_all("a", href=re.compile(rf"/cat/{category_slug}/p-\d+")):
            m = re.search(r"/p-(\d+)", link.get("href"))
            if m:
                last_page = max(last_page, int(m.group(1)))
        
        for p in range(2, last_page + 1):
            pages.append(f"{base_url}/p-{p}")
        return pages

    def _parse_page(self, url: str, category_slug: str) -> list[PublicRealEstateLot]:
        html = self._get_html(url)
        soup = BeautifulSoup(html, "lxml")
        containers = soup.select(".card-box, .blog-style-right-card, .lot-item, .item, div[class*='card']")
        lots: list[PublicRealEstateLot] = []
        seen_urls: set[str] = set()

        for item in containers:
            link_tag = item.find("a", href=True)
            if not link_tag:
                continue

            title = link_tag.get_text(" ", strip=True)
            if not title or len(title) < 5:
                continue

            item_url = urljoin(f"https://{self.city_slug}.gorod-torgi.ru", link_tag.get("href", ""))
            if item_url in seen_urls:
                continue
            seen_urls.add(item_url)

            text = item.get_text(" ", strip=True)
            price_match = re.search(r"[\d\s\xa0]+(?:₽|руб|р\.)", text, re.IGNORECASE)
            price_text = price_match.group(0).strip() if price_match else ""
            lot = PublicRealEstateLot(
                source="gorod-torgi.ru",
                category=category_slug,
                published_at="",
                asset_type="",
                status="",
                price=price_text,
                title=title,
                location=extract_address(text) or "",
                url=item_url,
                reference_url=item_url,
                source_label="primary",
            )

            if self.resolve_tbankrot:
                lot = self._resolve_tbankrot_reference_from_text(lot, text, item_url)
            lots.append(lot)

        return lots

    def fetch_listing_lots(self) -> list[PublicRealEstateLot]:
        lots: list[PublicRealEstateLot] = []
        seen_urls: set[str] = set()

        for category_slug in DEFAULT_CATEGORY_SLUGS:
            try:
                pages = self._iter_category_pages(category_slug)
            except requests.RequestException as exc:
                logger.warning("GorodTorgi: failed to load category %s: %s", category_slug, exc)
                continue

            for page_url in pages:
                try:
                    page_lots = self._parse_page(page_url, category_slug)
                except requests.RequestException as exc:
                    logger.warning("GorodTorgi: skipped page %s: %s", page_url, exc)
                    continue

                for lot in page_lots:
                    if lot.url in seen_urls:
                        continue
                    seen_urls.add(lot.url)
                    lots.append(lot)

        return lots

    @staticmethod
    def _extract_tbankrot_status(text: str) -> str | None:
        text = text.lower()
        if "состоял" in text: return "Состоявшийся"
        if "заверш" in text: return "Завершено"
        if "прием" in text: return "Приём заявок"
        return None

    def _resolve_tbankrot_reference_from_text(self, lot: PublicRealEstateLot, container_text: str, item_url: str) -> PublicRealEstateLot:
        status = self._extract_tbankrot_status(container_text)
        if status:
            lot.tbankrot_status = status
            lot.tbankrot_status_note = f"Derived from TBankrot text: {container_text[:50]}"
        
        # Если в тексте есть "прием заявок до [будущая дата]", ставим "Приём заявок"
        if "прием заявок до" in container_text.lower():
            lot.tbankrot_status = "Приём заявок"
            lot.tbankrot_status_note = f"Derived from TBankrot trade date: {container_text}"
            
        return lot

    def fetch_lots(self) -> list[dict]:
        try:
            url = f"https://{self.city_slug}.gorod-torgi.ru/cat/realizaciya-imuschestva-dolzhnikov"
            logger.info(f"GorodTorgi: запрос к {url}")
            html = self._get_html(url)
            soup = BeautifulSoup(html, "lxml")
            # Обновленные селекторы для GorodTorgi
            containers = soup.select(".card-box, .blog-style-right-card, .lot-item, .item, div[class*='card']")
            base_lots = []
            for item in containers:
                title_tag = item.find("a", href=True)
                if not title_tag: continue
                
                title = title_tag.get_text(strip=True)
                if not title or len(title) < 5: continue
                
                link = title_tag.get("href") or ""
                if link and not link.startswith("http"): 
                    link = f"https://{self.city_slug}.gorod-torgi.ru" + link
                
                price_text = ""
                price_tag = item.find(string=re.compile(r"[\d\s\xa0]+(?:₽|руб|р\.)", re.IGNORECASE))
                if price_tag: 
                    price_text = price_tag.strip()
                
                # Избегаем дублей по ссылке
                if any(l["url"] == link for l in base_lots): continue
                
                base_lots.append({"title": title, "price_text": price_text, "url": link, "raw": {"source": "gorod-torgi"}})
            
            logger.info(f"GorodTorgi: найдено {len(base_lots)} уникальных лотов для {self.city_slug}. Загружаем детали...")
            
            detailed = []
            # Ограничиваем количество лотов для детального парсинга, чтобы избежать бесконечного ожидания
            target_lots = base_lots[:60]  # можно увеличить
            
            with ThreadPoolExecutor(max_workers=12) as executor:   # было 5
                futures = {executor.submit(self._fetch_detail, lot): lot for lot in target_lots}
                for future in as_completed(futures):
                    result = future.result()
                    if result:
                        detailed.append(result)
            
            return detailed
        except Exception as e:
            logger.error(f"GorodTorgi error: {e}")
            return []

    def _fetch_detail(self, lot: dict) -> dict | None:
        try:
            if not lot.get('url'):
                return lot
                
            resp = self.session.get(lot['url'], timeout=10)
            resp.raise_for_status()
            
            l_soup = BeautifulSoup(resp.text, 'lxml')
            full_text = l_soup.get_text(" ", strip=True)
            
            # Применяем экстракторы
            lot['description'] = full_text[:2000] # Увеличим лимит описания
            lot['address'] = extract_address(full_text) or ""
            lot['cadastral_number'] = extract_cadastral(full_text) or ""
            lot['area'] = extract_area(full_text)
            lot['land_area'] = extract_land_area(full_text)
            lot['floors'] = extract_floors(full_text)
            lot['legal_status'] = extract_legal_status(full_text)
            
            # Небольшая пауза для вежливости к серверу
            time.sleep(0.08)   # было 0.3 — достаточно для вежливости
            return lot
        except Exception as e:
            logger.warning(f"Ошибка при загрузке деталей лота {lot.get('url')}: {e}")
            return lot # Возвращаем хотя бы базовую информацию



class TBankrotClient:
    BASE_URL = "https://tbankrot.ru"
    SEARCH_ENDPOINT = f"{BASE_URL}/"

    REGION_LABELS = {
        "4": "Амурская область", "10": "Бурятия", "16": "Еврейская АО", "24": "Камчатский край",
        "37": "Магаданская область", "52": "Приморский край", "60": "Сахалинская область",
        "74": "Хабаровский край", "81": "Чукотский АО", "82": "Якутия-Саха",
        "7": "Башкортостан", "28": "Кировская область", "38": "Марий Эл", "39": "Мордовия",
        "44": "Нижегородская область", "48": "Оренбургская область", "50": "Пензенская область",
        "51": "Пермский край", "57": "Самарская область", "59": "Саратовская область",
        "66": "Татарстан", "72": "Удмуртская Республика", "73": "Ульяновская область",
        "80": "Чувашская республика", "5": "Архангельская область", "13": "Вологодская область",
        "21": "Калининградская область", "26": "Карелия", "29": "Коми", "35": "Ленинградская область",
        "42": "Мурманская область", "43": "Ненецкий АО", "45": "Новгородская область",
        "53": "Псковская область", "58": "Санкт-Петербург", "15": "Дагестан", "85": "Ингушетия",
        "20": "Кабардино-Балкарская республика", "25": "Карачаево-Черкесская республика",
        "54": "Северная Осетия-Алания", "64": "Ставропольский край", "78": "Чеченская Республика",
        "2": "Алтай", "3": "Алтайский край", "17": "Забайкальский край", "19": "Иркутская область",
        "27": "Кемеровская область", "32": "Красноярский край", "46": "Новосибирская область",
        "47": "Омская область", "68": "Томская область", "87": "Тыва", "75": "Хакасия",
        "33": "Курганская область", "61": "Свердловская область", "71": "Тюменская область",
        "76": "Ханты-Мансийский АО", "77": "Челябинская область", "83": "Ямало-Ненецкий АО",
        "8": "Белгородская область", "9": "Брянская область", "11": "Владимирская область",
        "14": "Воронежская область", "18": "Ивановская область", "23": "Калужская область",
        "30": "Костромская область", "34": "Курская область", "36": "Липецкая область",
        "40": "Москва", "41": "Московская область", "49": "Орловская область",
        "56": "Рязанская область", "63": "Смоленская область", "65": "Тамбовская область",
        "67": "Тверская область", "70": "Тульская область", "84": "Ярославская область",
        "1": "Адыгея", "6": "Астраханская область", "12": "Волгоградская область",
        "22": "Калмыкия", "31": "Краснодарский край", "88": "Крым",
        "55": "Ростовская область", "62": "Севастополь",
    }

    REGION_SLUGS = {
        "yaroslavl": "84",
        "magadan": "37",
        "sakhalin": "60",
        "orel": "49",
        "murmansk": "42",
        "penza": "50",
        "perm": "51",
        "ryazan": "56",
        "saratov": "59",
        "moscow": "40",
        "moscow_region": "41",
        "spb": "58",
        "saint_petersburg": "58",
    }

    REGION_CODE_TO_PATH = {
        "1": "respublika-adygeya",
        "20": "kabardino-balkarskaya-respublika",
        "25": "karachaevo-cherkesskaya-respublika",
        "31": "krasnodarskiy-kray",
        "55": "rostovskaya-oblast",
        "64": "stavropolskiy-kray",
        "84": "yaroslavskaya-oblast",
        "40": "moskva",
        "41": "moskovskaya-oblast",
        "58": "sankt-peterburg",
    }
    REGION_NAME_TO_PATH = {
        "Алтайский край": "altayskiy-kray",
        "Амурская область": "amurskaya-oblast",
        "Архангельская область": "arhangelskaya-oblast",
        "Астраханская область": "astrahanskaya-oblast",
        "Белгородская область": "belgorodskaya-oblast",
        "Брянская область": "bryanskaya-oblast",
        "Владимирская область": "vladimirskaya-oblast",
        "Волгоградская область": "volgogradskaya-oblast",
        "Вологодская область": "vologodskaya-oblast",
        "Воронежская область": "voronezhskaya-oblast",
        "Еврейская автономная область": "evreyskaya-ao",
        "Забайкальский край": "zabaykalskiy-kray",
        "Ивановская область": "ivanovskaya-oblast",
        "Иркутская область": "irkutskaya-oblast",
        "Кабардино-Балкарская Республика": "kabardino-balkarskaya-respublika",
        "Калининградская область": "kaliningradskaya-oblast",
        "Калужская область": "kaluzhskaya-oblast",
        "Камчатский край": "kamchatskiy-kray",
        "Карачаево-Черкесская Республика": "karachaevo-cherkesskaya-respublika",
        "Кемеровская область": "kemerovskaya-oblast",
        "Кировская область": "kirovskaya-oblast",
        "Костромская область": "kostromskaya-oblast",
        "Краснодарский край": "krasnodarskiy-kray",
        "Красноярский край": "krasnoyarskiy-kray",
        "Курганская область": "kurganskaya-oblast",
        "Курская область": "kurskaya-oblast",
        "Ленинградская область": "leningradskaya-oblast",
        "Липецкая область": "lipeckaya-oblast",
        "Магаданская область": "magadanskaya-oblast",
        "Москва": "moskva",
        "Московская область": "moskovskaya-oblast",
        "Мурманская область": "murmanskaya-oblast",
        "Ненецкий автономный округ": "nenckiy-ao",
        "Нижегородская область": "nizhegorodskaya-oblast",
        "Новгородская область": "novgorodskaya-oblast",
        "Новосибирская область": "novosibirskaya-oblast",
        "Омская область": "omskaya-oblast",
        "Оренбургская область": "orenburgskaya-oblast",
        "Орловская область": "orlovskaya-oblast",
        "Пензенская область": "penzenskaya-oblast",
        "Пермский край": "permskiy-kray",
        "Приморский край": "primorskiy-kray",
        "Псковская область": "pskovskaya-oblast",
        "Республика Адыгея": "respublika-adygeya",
        "Республика Алтай": "respublika-altay",
        "Республика Башкортостан": "respublika-bashkortostan",
        "Республика Бурятия": "respublika-buryatiya",
        "Республика Дагестан": "respublika-dagestan",
        "Республика Ингушетия": "respublika-ingushetiya",
        "Республика Калмыкия": "respublika-kalmykiya",
        "Республика Карелия": "respublika-kareliya",
        "Республика Коми": "respublika-komi",
        "Республика Крым": "respublika-krym",
        "Республика Марий Эл": "respublika-mariy-el",
        "Республика Мордовия": "respublika-mordoviya",
        "Республика Саха (Якутия)": "yakutiya-saha",
        "Республика Северная Осетия — Алания": "severnaya-osetiya-alaniya",
        "Республика Татарстан": "tatarstan",
        "Республика Тыва": "respublika-tyva",
        "Республика Хакасия": "respublika-hakasiya",
        "Ростовская область": "rostovskaya-oblast",
        "Рязанская область": "ryazanskaya-oblast",
        "Самарская область": "samarskaya-oblast",
        "Санкт-Петербург": "sankt-peterburg",
        "Саратовская область": "saratovskaya-oblast",
        "Сахалинская область": "sahalinskaya-oblast",
        "Свердловская область": "sverdlovskaya-oblast",
        "Севастополь": "sevastopol",
        "Смоленская область": "smolenskaya-oblast",
        "Ставропольский край": "stavropolskiy-kray",
        "Тамбовская область": "tambovskaya-oblast",
        "Тверская область": "tverskaya-oblast",
        "Томская область": "tomskaya-oblast",
        "Тульская область": "tulskaya-oblast",
        "Тюменская область": "tyumenskaya-oblast",
        "Удмуртская Республика": "udmurtskaya-respublika",
        "Ульяновская область": "ulyanovskaya-oblast",
        "Хабаровский край": "habarovskiy-kray",
        "Ханты-Мансийский автономный округ": "hanty-mansiyskiy-ao",
        "Челябинская область": "chelyabinskaya-oblast",
        "Чеченская Республика": "chechenskaya-respublika",
        "Чувашская Республика": "chuvashskaya-respublika",
        "Чукотский автономный округ": "chukotskiy-avtonomnyy-okrug",
        "Ямало-Ненецкий автономный округ": "yamalo-neneckiy-ao",
        "Ярославская область": "yaroslavskaya-oblast",
    }
    CADASTRAL_REGION_NAME_TO_CODE = {
        "Республика Адыгея": "01",
        "Адыгея": "01",
        "Республика Башкортостан": "02",
        "Башкортостан": "02",
        "Республика Бурятия": "03",
        "Бурятия": "03",
        "Республика Алтай": "04",
        "Алтай": "04",
        "Республика Дагестан": "05",
        "Дагестан": "05",
        "Республика Ингушетия": "06",
        "Ингушетия": "06",
        "Кабардино-Балкарская Республика": "07",
        "Республика Калмыкия": "08",
        "Калмыкия": "08",
        "Карачаево-Черкесская Республика": "09",
        "Республика Карелия": "10",
        "Карелия": "10",
        "Республика Коми": "11",
        "Коми": "11",
        "Республика Марий Эл": "12",
        "Марий Эл": "12",
        "Республика Мордовия": "13",
        "Мордовия": "13",
        "Республика Саха (Якутия)": "14",
        "Якутия-Саха": "14",
        "Республика Северная Осетия — Алания": "15",
        "Северная Осетия-Алания": "15",
        "Республика Татарстан": "16",
        "Татарстан": "16",
        "Республика Тыва": "17",
        "Тыва": "17",
        "Удмуртская Республика": "18",
        "Республика Хакасия": "19",
        "Хакасия": "19",
        "Чеченская Республика": "20",
        "Чувашская Республика": "21",
        "Чувашская республика": "21",
        "Алтайский край": "22",
        "Краснодарский край": "23",
        "Красноярский край": "24",
        "Приморский край": "25",
        "Ставропольский край": "26",
        "Хабаровский край": "27",
        "Амурская область": "28",
        "Архангельская область": "29",
        "Астраханская область": "30",
        "Белгородская область": "31",
        "Брянская область": "32",
        "Владимирская область": "33",
        "Волгоградская область": "34",
        "Вологодская область": "35",
        "Воронежская область": "36",
        "Ивановская область": "37",
        "Иркутская область": "38",
        "Калининградская область": "39",
        "Калужская область": "40",
        "Камчатский край": "41",
        "Кемеровская область": "42",
        "Кировская область": "43",
        "Костромская область": "44",
        "Курганская область": "45",
        "Курская область": "46",
        "Ленинградская область": "47",
        "Липецкая область": "48",
        "Магаданская область": "49",
        "Московская область": "50",
        "Мурманская область": "51",
        "Нижегородская область": "52",
        "Новгородская область": "53",
        "Новосибирская область": "54",
        "Омская область": "55",
        "Оренбургская область": "56",
        "Орловская область": "57",
        "Пензенская область": "58",
        "Пермский край": "59",
        "Псковская область": "60",
        "Ростовская область": "61",
        "Рязанская область": "62",
        "Самарская область": "63",
        "Саратовская область": "64",
        "Сахалинская область": "65",
        "Свердловская область": "66",
        "Смоленская область": "67",
        "Тамбовская область": "68",
        "Тверская область": "69",
        "Томская область": "70",
        "Тульская область": "71",
        "Тюменская область": "72",
        "Ульяновская область": "73",
        "Челябинская область": "74",
        "Забайкальский край": "75",
        "Ярославская область": "76",
        "Москва": "77",
        "Санкт-Петербург": "78",
        "Еврейская автономная область": "79",
        "Еврейская АО": "79",
        "Ненецкий автономный округ": "83",
        "Ненецкий АО": "83",
        "Ханты-Мансийский автономный округ": "86",
        "Ханты-Мансийский АО": "86",
        "Чукотский автономный округ": "87",
        "Чукотский АО": "87",
        "Ямало-Ненецкий автономный округ": "89",
        "Ямало-Ненецкий АО": "89",
        "Республика Крым": "91",
        "Крым": "91",
        "Севастополь": "92",
    }

    def __init__(self, diagnostics: bool = False):
        self.diagnostics = diagnostics
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        })

    def search_lots(
        self,
        filters_or_city: TBankrotSearchFilters | str,
        search_query: str | None = None,
        max_pages: int = 2,
    ):
        if isinstance(filters_or_city, TBankrotSearchFilters):
            return self.search_filtered_lots(filters_or_city)
        return self._search_lots_legacy(filters_or_city, search_query=search_query, max_pages=max_pages)

    def search_filtered_lots(self, filters: TBankrotSearchFilters) -> tuple[list[NormalizedLot], dict]:
        self._set_page_item_count(filters.page_size)
        params = self._build_query_params(filters)
        endpoint = self._prepare_url(params, filters)
        request_url = self._base_url_for_filters(filters)
        try:
            resp = self.session.get(request_url, params=params, timeout=25)
            resp.raise_for_status()
        except Exception as exc:
            raise RuntimeError(f"TBankrot: не удалось загрузить страницу поиска: {exc}") from exc

        lots = self._parse_listing_html(resp.text, filters=filters, raw_endpoint=resp.url or endpoint)
        pagination = self._extract_pagination_meta(resp.text)
        meta = {
            "source": "tbankrot.ru",
            "mode": "page",
            "page": filters.page,
            "loaded": len(lots),
            "raw_endpoint": resp.url or endpoint,
            "raw_params": params,
            "has_more": bool(pagination.get("total_pages") and filters.page < pagination["total_pages"]) or len(lots) >= filters.page_size,
            "total_pages": pagination.get("total_pages"),
            "page_size": filters.page_size,
            "warnings": [],
        }
        return lots, meta

    def search_all_lots(
        self,
        filters: TBankrotSearchFilters,
        *,
        max_items: int | None = 5000,
        progress_cb: Callable[[int, int | None, int], None] | None = None,
        page_cb: Callable[[list[NormalizedLot], dict], None] | None = None,
        stop_cb: Callable[[], bool] | None = None,
    ) -> tuple[list[NormalizedLot], dict]:
        all_lots: list[NormalizedLot] = []
        seen: set[str] = set()
        page = max(1, filters.page)
        pages_loaded = 0
        duplicates = 0
        stop_reason = None
        page_diagnostics: list[dict[str, Any]] = []

        while True:
            if stop_cb and stop_cb():
                stop_reason = "user_stopped"
                break

            page_filters = replace(filters, page=page)
            lots_on_page, page_meta = self.search_filtered_lots(page_filters)
            pages_loaded += 1
            new_lots: list[NormalizedLot] = []
            page_duplicates = 0

            for lot in lots_on_page:
                if lot.external_id in seen:
                    duplicates += 1
                    page_duplicates += 1
                    continue
                seen.add(lot.external_id)
                new_lots.append(lot)
                all_lots.append(lot)
                if max_items and len(all_lots) >= max_items:
                    stop_reason = "max_items"
                    break

            page_meta = dict(page_meta)
            page_meta.update({
                "mode": "all_pages",
                "loaded": len(all_lots),
                "pages_loaded": pages_loaded,
                "duplicates": duplicates,
            })
            if page_cb and new_lots:
                page_cb(new_lots, page_meta)
            if progress_cb:
                progress_cb(page, None, len(all_lots))
            page_diagnostics.append({
                "page": page,
                "items_on_page": len(lots_on_page),
                "new_unique": len(new_lots),
                "duplicates": page_duplicates,
            })

            if stop_reason:
                break
            total_pages = page_meta.get("total_pages")
            if total_pages:
                if page >= int(total_pages):
                    stop_reason = "last_page"
                    break
            elif len(lots_on_page) < filters.page_size:
                stop_reason = "last_page"
                break
            if lots_on_page and not new_lots:
                stop_reason = "duplicates_only"
                break

            page += 1
            time.sleep(0.2)

        meta = {
            "source": "tbankrot.ru",
            "mode": "all_pages",
            "page": page,
            "loaded": len(all_lots),
            "pages_loaded": pages_loaded,
            "duplicates": duplicates,
            "stop_reason": stop_reason,
            "page_diagnostics": page_diagnostics,
            "has_more": stop_reason not in {"last_page", "duplicates_only"},
            "raw_params": self._build_query_params(filters),
            "raw_endpoint": self._prepare_url(self._build_query_params(filters), filters),
            "warnings": [],
        }
        return all_lots, meta

    def _build_query_params(self, filters: TBankrotSearchFilters) -> list[tuple[str, str]]:
        regional_path = self._regional_path(filters.region)
        params: list[tuple[str, str]] = [] if regional_path else [("p", "search")]

        def add(name: str, value: Any) -> None:
            if value in (None, "", [], {}):
                return
            params.append((name, str(value)))

        add("search", filters.search_text.strip() if filters.search_text else None)
        if not regional_path:
            add("region[]", filters.region)
        add("start_p1", self._format_query_number(filters.price_min))
        add("start_p2", self._format_query_number(filters.price_max))
        add("num", filters.lot_number)
        if filters.trade_type == "auction":
            add("type_2", "on")
        elif filters.trade_type == "public":
            add("type_1", "on")
        if filters.photo_only:
            add("photo", "1")
        add("debtor", filters.debtor)
        add("au", filters.auction_manager)
        add("org", filters.organizer)
        add("stop", filters.stop_words)
        if filters.show_closed:
            add("show_closed", "1")
        if filters.show_paused:
            add("show_paused", "1")
        page = max(1, int(filters.page or 1))
        if regional_path:
            add("sort", "created")
            add("sort_order", "desc")
            add("show_period", "all")
            add("page", page)
        elif page > 1:
            add("page", page)
        return params

    def _prepare_url(self, params: list[tuple[str, str]], filters: TBankrotSearchFilters | None = None) -> str:
        base_url = self._base_url_for_filters(filters) if filters else self.SEARCH_ENDPOINT
        query = urlencode(params)
        return f"{base_url}?{query}" if query else base_url

    def _base_url_for_filters(self, filters: TBankrotSearchFilters | None) -> str:
        regional_path = self._regional_path(filters.region if filters else None)
        if regional_path:
            return f"{self.BASE_URL}/torgi/r/{regional_path}"
        return self.SEARCH_ENDPOINT

    def _regional_path(self, region_code: str | None) -> str | None:
        if not region_code:
            return None
        code = str(region_code)
        if code in self.REGION_CODE_TO_PATH:
            return self.REGION_CODE_TO_PATH[code]

        region_name = self.REGION_LABELS.get(code)
        if not region_name:
            return None

        normalized = self._normalize_region_lookup_name(region_name)
        for known_name, path in self.REGION_NAME_TO_PATH.items():
            if normalized == self._normalize_region_lookup_name(known_name):
                return path

        return self._build_region_path_from_name(region_name)

    def _normalize_region_lookup_name(self, name: str) -> str:
        normalized = name.lower().replace("ё", "е")
        normalized = normalized.replace("—", " ").replace("-", " ")
        normalized = normalized.replace("автономная область", "ао")
        normalized = normalized.replace("автономный округ", "ао")
        normalized = normalized.replace("республика", "")
        normalized = normalized.replace("(", " ").replace(")", " ")
        normalized = re.sub(r"[^а-яa-z0-9]+", " ", normalized)
        return re.sub(r"\s+", " ", normalized).strip()

    def _build_region_path_from_name(self, region_name: str) -> str:
        translit = {
            "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
            "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
            "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
            "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch",
            "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
        }
        source = region_name.lower().replace("ё", "е")
        source = source.replace("автономная область", "ао")
        source = source.replace("автономный округ", "ао")
        slug = "".join(translit.get(ch, ch) for ch in source)
        slug = re.sub(r"[^a-z0-9]+", "-", slug).strip("-")
        return slug[:120]

    def _official_region_code(self, region_name: str | None, cadastral_number: str | None = None) -> str | None:
        cad_code = self._region_code_from_cadastral_number(cadastral_number)
        if cad_code:
            return cad_code
        if not region_name:
            return None

        normalized = self._normalize_region_lookup_name(region_name)
        for known_name, code in self.CADASTRAL_REGION_NAME_TO_CODE.items():
            if normalized == self._normalize_region_lookup_name(known_name):
                return code
        return None

    def _region_code_from_cadastral_number(self, cadastral_number: str | None) -> str | None:
        if not cadastral_number:
            return None
        match = re.search(r"\b(\d{1,2})\s*:", str(cadastral_number))
        if not match:
            return None
        return match.group(1).zfill(2)

    def _set_page_item_count(self, page_size: int | None) -> None:
        count = max(1, int(page_size or 100))
        self.session.cookies.set("pageitemcount", str(count), domain="tbankrot.ru", path="/")

    def _extract_pagination_meta(self, html: str) -> dict[str, int | None]:
        soup = BeautifulSoup(html, "lxml")
        pages = []
        for item in soup.select("#paginator [data-page], #paginator a[href*='page=']"):
            value = item.get("data-page") or item.get_text(" ", strip=True)
            if not value:
                href = item.get("href") or ""
                match = re.search(r"[?&]page=(\d+)", href)
                value = match.group(1) if match else ""
            try:
                pages.append(int(str(value).strip()))
            except ValueError:
                continue
        return {"total_pages": max(pages) if pages else None}

    def _format_query_number(self, value: float | int | None) -> str | None:
        if value is None:
            return None
        try:
            as_float = float(value)
            if as_float.is_integer():
                return str(int(as_float))
            return str(as_float).replace(".", ",")
        except (TypeError, ValueError):
            return str(value)

    def _parse_listing_html(
        self,
        html: str,
        *,
        filters: TBankrotSearchFilters | None = None,
        raw_endpoint: str | None = None,
    ) -> list[NormalizedLot]:
        parser = ManualHtmlParser(ParserConfig(base_url=self.BASE_URL))
        parsed_lots = parser.parse_html(html)
        return [self._normalize_parsed_lot(item, filters=filters, raw_endpoint=raw_endpoint) for item in parsed_lots]

    def _normalize_parsed_lot(
        self,
        item: ParsedLotData,
        *,
        filters: TBankrotSearchFilters | None = None,
        raw_endpoint: str | None = None,
    ) -> NormalizedLot:
        raw_payload = dict(item.raw_payload or {})
        raw_payload.update({
            "source": "tbankrot_search",
            "price_text": item.price_text,
            "raw_endpoint": raw_endpoint,
            "tbankrot_region_code": filters.region if filters else None,
        })
        region_code = filters.region if filters else None
        region_name = self.REGION_LABELS.get(str(region_code)) if region_code else None
        cadastral_hint = item.cadastral_number or (item.cadastral_numbers[0] if item.cadastral_numbers else None)
        official_region_code = self._official_region_code(region_name, cadastral_hint)
        raw_payload["region_code"] = official_region_code
        raw_payload["region_name"] = region_name
        description = item.description or ""
        return NormalizedLot(
            external_id=f"tbankrot:{item.external_id}",
            source="tbankrot",
            source_system="tbankrot.ru",
            title=(item.title or "")[:500],
            description=description[:5000],
            category=classify_category(item.title or "", description),
            region_slug=official_region_code or (str(region_code) if region_code else None),
            region_name=region_name,
            address=item.address or None,
            cadastral_number=item.cadastral_number or None,
            vin=None,
            area=item.area,
            start_price=None,
            current_price=item.current_price,
            auction_status=item.status or "unknown",
            lot_url=item.url,
            source_url=item.url,
            detail_level="search",
            raw_data=raw_payload,
            published_at=None,
            total_area_gba=item.building_area or item.room_area,
            land_area=item.land_area,
            floors=item.floors,
            year_built=item.year_built,
        )

    def _search_lots_legacy(self, city_slug: str, search_query: str | None = None, max_pages: int = 2) -> list[dict]:
        region_id = self.REGION_SLUGS.get(city_slug)
        if not region_id:
            logger.warning(f"Unknown region slug: {city_slug}")
            return []

        all_lots: list[dict] = []
        for page in range(1, max_pages + 1):
            filters = TBankrotSearchFilters(
                search_text=search_query or "",
                region=region_id,
                page=page,
            )
            try:
                logger.info(f"TBankrot: Fetching page {page} for {city_slug} (query: {search_query})")
                lots, _meta = self.search_filtered_lots(filters)
                if not lots:
                    logger.info("TBankrot: No more lots found.")
                    break
                for lot in lots:
                    all_lots.append({
                        "external_id": lot.external_id.replace("tbankrot:", ""),
                        "title": lot.title,
                        "url": lot.lot_url,
                        "price_text": (lot.raw_data or {}).get("price_text") or "",
                        "description": lot.description,
                        "address": lot.address or "",
                        "source": "tbankrot",
                    })
            except Exception as e:
                logger.error(f"TBankrot error on page {page}: {e}")
                break

        return all_lots

def sync_public_real_estate(session: Session, city_slug: str, search: str | None = None) -> list:
    logger.info(f"Запуск синхронизации TBankrot: {city_slug} (search={search})")
    client = TBankrotClient()
    raw_lots = client.search_lots(city_slug, search_query=search)
    
    imported = []
    for raw in raw_lots:
        price_val = extract_price(raw["price_text"])
        
        normalized = NormalizedLot(
            external_id=raw["external_id"],
            source="tbankrot",
            source_system="tbankrot",
            title=raw["title"],
            description=raw["description"],
            category=classify_category(raw["title"], raw["description"]),
            region_slug=city_slug,
            region_name=None,
            address=raw["address"],
            cadastral_number=None,
            vin=None,
            area=None,
            start_price=None,
            current_price=price_val,
            lot_url=raw["url"],
            source_url=raw["url"],
            detail_level="list",
            raw_data=raw,
            auction_status="active"
        )
        processed = persist_lot(session, normalized)
        # We don't do full geo enrichment here to keep it fast, 
        # but persist_lot or subsequent logic might handle it
        imported.append(processed)
    
    # Update RegionSyncState
    from bankrotai.db import upsert_region_sync_state
    upsert_region_sync_state(
        session, 
        city_slug, 
        status="ready", 
        last_success_at=datetime.now(),
        ready_lots=len(imported)
    )
    
    return imported



def ingest_recent_tbankrot(session: Session, city_slug: str = None) -> list:
    logger.info(f"ingest_recent_tbankrot → {city_slug}")
    url = "https://tbankrot.ru/lots"
    try:
        resp = requests.get(url, timeout=15, headers={"User-Agent": "BankrotAI Bot"})
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")
        lot_items = soup.select("div.lot_container, div.lot[data-id]")
        imported = []
        for item in lot_items[:30]:
            title_tag = item.select_one("p.lot_title a")
            if not title_tag:
                continue

            title = title_tag.get_text(" ", strip=True)
            href = title_tag.get("href") or ""
            url = urljoin("https://tbankrot.ru", href)

            # Описание
            desc_tag = item.select_one(".lot_description .text")
            description = desc_tag.get_text("\n", strip=True) if desc_tag else ""

            # Цена
            price_tag = item.select_one(".lot_prices .current_price span")
            price_text = price_tag.get_text(" ", strip=True) if price_tag else ""
            current_price = parse_money(price_text)

            # ID
            lot_div = item if "lot" in item.get("class", []) else item.select_one("div.lot[data-id]")
            external_id = lot_div.get("data-id") if lot_div else None
            if not external_id:
                m = re.search(r'id=(\d+)', href)
                external_id = m.group(1) if m else f"tb_{hashlib.md5(url.encode()).hexdigest()[:12]}"

            raw = {
                "source": "tbankrot_online",
                "title": title,
                "url": url,
                "description": description,
                "price_text": price_text,
            }

            normalized = NormalizedLot(
                external_id=str(external_id),
                source="tbankrot",
                source_system="tbankrot",
                title=title,
                description=description[:5000],
                category=classify_category(title, description),
                region_slug=city_slug,
                region_name=None,
                address=extract_address(description) or "",
                cadastral_number=extract_cadastral(description),
                vin=None,
                area=extract_area(description),
                start_price=None,
                current_price=current_price,
                auction_status="active", # По умолчанию для онлайн-списка, но можно улучшить
                lot_url=url,
                source_url=url,
                detail_level="list",
                raw_data=raw
            )
            processed = persist_lot(session, normalized)
            upsert_lot_events_from_raw(session, processed, raw)
            imported.append(processed)
        return imported
    except Exception as e:
        logger.error(f"tbankrot error: {e}")
        return []
