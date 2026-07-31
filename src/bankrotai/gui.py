from __future__ import annotations

import json
import logging
import os
import re
import socket
import sys
import tempfile
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from typing import Callable
from urllib.parse import parse_qsl, urlencode, urlparse

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

# Desktop HTTPS requests should honor the Windows certificate store. This keeps
# TLS verification enabled while supporting managed root CAs that are absent
# from certifi's public CA bundle.
try:
    import truststore

    truststore.inject_into_ssl()
except Exception:
    logging.getLogger(__name__).warning(
        "Windows system TLS trust store could not be enabled; falling back to the bundled CA certificates",
        exc_info=True,
    )

# Disable hardware acceleration for Qt to avoid Trae Sandbox errors
os.environ["QT_QUICK_BACKEND"] = "software"
os.environ["QT_XCB_GL_INTEGRATION"] = "none"
os.environ["QT_OPENGL"] = "software"
os.environ["QTWEBENGINE_DISABLE_GPU"] = "1"
os.environ["QTWEBENGINE_CHROMIUM_FLAGS"] = "--disable-gpu"
from datetime import datetime

from PySide6.QtCore import (
    Qt, QTimer, QUrl, QThread, Signal, Slot, QObject, QStandardPaths,
    QStringListModel,
)
from PySide6.QtWebEngineWidgets import QWebEngineView
from PySide6.QtWebEngineCore import QWebEngineSettings
from PySide6.QtWebChannel import QWebChannel
from PySide6.QtGui import QColor, QBrush, QDesktopServices
from PySide6.QtWidgets import (
    QAbstractItemView, QApplication, QHBoxLayout, QHeaderView, QLabel,
    QMainWindow, QMessageBox, QPushButton, QSplitter, QTableWidget,
    QTableWidgetItem, QTextEdit, QVBoxLayout, QWidget, QTabWidget, QFileDialog,
    QProgressBar, QStatusBar, QLineEdit, QComboBox, QCheckBox, QGroupBox, QFormLayout,
    QCompleter, QScrollArea, QSizePolicy, QToolButton, QFrame, QDialog, QDoubleSpinBox,
)
from sqlalchemy import and_, desc, or_, select, func

from bankrotai.core import get_logger, get_settings
from bankrotai.db import ProcessedLot, SourceLot, LotGeoSnapshot, session_scope, init_db, RegionSyncState
from bankrotai.scrapers import (
    LotOnlineClient,
    LotOnlineSearchFilters,
    TBankrotClient,
    TBankrotSearchFilters,
    TorgiGovClient,
    TorgiGovSearchFilters,
    import_manual_html,
    sync_public_real_estate,
    ingest_recent_tbankrot,
)
from bankrotai.logic import (
    cleanup_closed_lots,
    delete_lots_batch,
    persist_lot,
    reconcile_cross_source_duplicates,
)
from bankrotai.ai import OpenAIAppraiser, apply_evaluation_to_lot
from openpyxl import Workbook
from bankrotai.domain import NormalizedLot
from bankrotai.geo import nspd_tls_verify
from bankrotai.finance import MaxBidInputs, calculate_max_bid

logger = get_logger("gui")


class WheelSafeComboBox(QComboBox):
    """
    QComboBox that still opens on click, but ignores wheel changes while closed.
    This prevents accidental filter changes when scrolling the left panel.
    """

    def wheelEvent(self, event):
        if self.view() and self.view().isVisible():
            super().wheelEvent(event)
        else:
            event.ignore()


SORT_ROLE = Qt.UserRole + 100
EXTERNAL_ID_ROLE = Qt.UserRole + 101
URL_ROLE = Qt.UserRole + 102
LOT_ID_ROLE = Qt.UserRole + 103
MAP_MARKER_LIMIT = 100_000


def map_assets_directory() -> Path:
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS")) / "bankrotai" / "assets" / "map"
    return Path(__file__).resolve().parent / "assets" / "map"


MAP_PREVIEW_STYLE = """
.lot-preview {
    position: absolute; z-index: 1000; top: 0; left: 0; bottom: 0; width: 380px;
    box-sizing: border-box; overflow-y: auto; background: #fff; color: #273142;
    box-shadow: 4px 0 18px rgba(34, 46, 66, .22); font: 14px Arial, sans-serif;
    transform: translateX(-105%); transition: transform .22s ease;
}
.lot-preview.open { transform: translateX(0); }
.lot-preview__close {
    position: absolute; z-index: 2; top: 10px; right: 10px; width: 34px; height: 34px;
    border: 0; border-radius: 50%; background: rgba(255,255,255,.94); color: #536174;
    font-size: 23px; cursor: pointer; box-shadow: 0 1px 5px rgba(0,0,0,.18);
}
.lot-preview__media { position: relative; height: 245px; background: #edf1f6; }
.lot-preview__photo, .lot-preview__placeholder {
    display: block; width: 100%; height: 245px; object-fit: cover; background: #edf1f6;
}
.lot-preview__placeholder { display: flex; align-items: center; justify-content: center; color: #8995a7; font-size: 15px; }
.lot-preview__arrow {
    position: absolute; z-index: 2; top: 50%; transform: translateY(-50%); width: 38px; height: 48px;
    border: 0; border-radius: 7px; background: rgba(20,30,45,.62); color: white; font-size: 28px;
    cursor: pointer; display: none;
}
.lot-preview__arrow:hover { background: rgba(20,30,45,.82); }
.lot-preview__arrow--prev { left: 10px; }
.lot-preview__arrow--next { right: 10px; }
.lot-preview__counter {
    position: absolute; right: 10px; bottom: 9px; padding: 4px 8px; border-radius: 12px;
    background: rgba(20,30,45,.68); color: white; font-size: 12px; display: none;
}
.lot-preview__body { padding: 18px; }
.lot-preview__source { color: #16866d; font-size: 12px; font-weight: 700; text-transform: uppercase; }
.lot-preview__title { margin: 10px 0; font-size: 17px; line-height: 1.45; }
.lot-preview__description { color: #536174; line-height: 1.45; max-height: 105px; overflow: auto; white-space: pre-wrap; }
.lot-preview__price { margin: 14px 0; font-size: 22px; font-weight: 700; }
.lot-preview__details { margin: 12px 0 16px; line-height: 1.55; color: #536174; }
.lot-preview__details b { color: #273142; }
.lot-preview__links { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 8px; }
.lot-preview__source-button {
    width: 100%; padding: 11px 6px; border: 0; border-radius: 7px; background: #2868e8;
    color: #fff; font-weight: 700; font-size: 15px; cursor: pointer;
}
.lot-preview__source-button[data-kind="gis"] { background: #177d65; }
.lot-preview__source-button[data-kind="etp"] { background: #7654b5; }
.lot-preview__source-button[data-kind="russia"] { background: #596579; }
.lot-preview__source-button:disabled { background: #aab3c1; cursor: default; }
.lot-preview__review-title { margin: 18px 0 9px; font-weight: 700; }
.lot-preview__reviews { display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px; }
.review-button { padding: 10px 4px; border: 2px solid #e1e6ed; border-radius: 8px; background: #fff; cursor: pointer; font-size: 12px; }
.review-button span { display: block; font-size: 23px; margin-bottom: 3px; }
.review-button[data-status="approved"] { color: #188b5b; }
.review-button[data-status="maybe"] { color: #b88700; }
.review-button[data-status="rejected"] { color: #d43f3f; }
.review-button.active[data-status="approved"] { border-color: #22a76f; background: #e8f8f0; }
.review-button.active[data-status="maybe"] { border-color: #e4b72c; background: #fff8d9; }
.review-button.active[data-status="rejected"] { border-color: #df5252; background: #fff0f0; }
"""

MAP_PREVIEW_HTML = """
<aside id="lot-preview" class="lot-preview" aria-hidden="true">
  <button id="lot-preview-close" class="lot-preview__close" title="&#1047;&#1072;&#1082;&#1088;&#1099;&#1090;&#1100;">&times;</button>
  <div class="lot-preview__media">
    <img id="lot-preview-photo" class="lot-preview__photo" alt="&#1060;&#1086;&#1090;&#1086; &#1083;&#1086;&#1090;&#1072;">
    <div id="lot-preview-placeholder" class="lot-preview__placeholder">&#1060;&#1086;&#1090;&#1086; &#1086;&#1090;&#1089;&#1091;&#1090;&#1089;&#1090;&#1074;&#1091;&#1077;&#1090;</div>
    <button id="lot-preview-prev" class="lot-preview__arrow lot-preview__arrow--prev" aria-label="Previous">&#8249;</button>
    <button id="lot-preview-next" class="lot-preview__arrow lot-preview__arrow--next" aria-label="Next">&#8250;</button>
    <div id="lot-preview-counter" class="lot-preview__counter"></div>
  </div>
  <div class="lot-preview__body">
    <div id="lot-preview-source" class="lot-preview__source"></div>
    <h2 id="lot-preview-title" class="lot-preview__title"></h2>
    <div id="lot-preview-description" class="lot-preview__description"></div>
    <div id="lot-preview-price" class="lot-preview__price"></div>
    <div id="lot-preview-details" class="lot-preview__details"></div>
    <div class="lot-preview__links">
      <button class="lot-preview__source-button related-link" data-url-key="source_url">&#1048;&#1089;&#1090;&#1086;&#1095;&#1085;&#1080;&#1082;</button>
      <button class="lot-preview__source-button related-link" data-kind="gis" data-url-key="gis_torgi_url">&#1043;&#1048;&#1057; &#1058;&#1086;&#1088;&#1075;&#1080;</button>
      <button class="lot-preview__source-button related-link" data-kind="etp" data-url-key="etp_url">&#1069;&#1058;&#1055;</button>
      <button class="lot-preview__source-button related-link" data-kind="russia" data-url-key="torgi_russia_url">&#1058;&#1086;&#1088;&#1075;&#1080; &#1056;&#1086;&#1089;&#1089;&#1080;&#1080;</button>
    </div>
    <div class="lot-preview__review-title">&#1054;&#1094;&#1077;&#1085;&#1082;&#1072; &#1083;&#1086;&#1090;&#1072;</div>
    <div class="lot-preview__reviews">
      <button class="review-button" data-status="approved"><span>&#10003;</span>&#1048;&#1085;&#1090;&#1077;&#1088;&#1077;&#1089;&#1077;&#1085;</button>
      <button class="review-button" data-status="maybe"><span>?</span>&#1057;&#1086;&#1084;&#1085;&#1077;&#1074;&#1072;&#1102;&#1089;&#1100;</button>
      <button class="review-button" data-status="rejected"><span>&#10005;</span>&#1055;&#1083;&#1086;&#1093;&#1086;&#1081;</button>
    </div>
  </div>
</aside>
"""

MAP_PREVIEW_SCRIPT = """
let bankrotaiBridge = null;
let selectedPreviewLot = null;
let previewImages = [];
let previewImageIndex = 0;

if (window.qt && window.qt.webChannelTransport) {
    new QWebChannel(qt.webChannelTransport, function(channel) {
        bankrotaiBridge = channel.objects.bankrotaiBridge;
    });
}

function previewText(id, value) {
    document.getElementById(id).textContent = value || '';
}

function setPreviewReviewStatus(status) {
    document.querySelectorAll('.review-button').forEach(function(button) {
        button.classList.toggle('active', button.dataset.status === status);
    });
}

function renderPreviewLinks() {
    document.querySelectorAll('.related-link').forEach(function(button) {
        const url = selectedPreviewLot && selectedPreviewLot[button.dataset.urlKey];
        button.disabled = !url;
    });
}

function renderPreviewImage() {
    const photo = document.getElementById('lot-preview-photo');
    const placeholder = document.getElementById('lot-preview-placeholder');
    const hasImage = previewImages.length > 0;
    photo.style.display = hasImage ? 'block' : 'none';
    placeholder.style.display = hasImage ? 'none' : 'flex';
    if (hasImage) photo.src = previewImages[previewImageIndex]; else photo.removeAttribute('src');
    const multiple = previewImages.length > 1;
    document.getElementById('lot-preview-prev').style.display = multiple ? 'block' : 'none';
    document.getElementById('lot-preview-next').style.display = multiple ? 'block' : 'none';
    const counter = document.getElementById('lot-preview-counter');
    counter.style.display = multiple ? 'block' : 'none';
    counter.textContent = hasImage ? (previewImageIndex + 1) + ' / ' + previewImages.length : '';
}

function movePreviewImage(delta) {
    if (previewImages.length < 2) return;
    previewImageIndex = (previewImageIndex + delta + previewImages.length) % previewImages.length;
    renderPreviewImage();
}

function showLotPreview(lot) {
    selectedPreviewLot = lot;
    const panel = document.getElementById('lot-preview');
    panel.classList.add('open');
    panel.setAttribute('aria-hidden', 'false');
    if (bankrotaiBridge) bankrotaiBridge.previewOpened(mapKind, Number(lot.id));
    previewText('lot-preview-source', lot.source_name || lot.source || 'Источник не указан');
    previewText('lot-preview-title', lot.title || 'Лот без названия');
    previewText('lot-preview-description', lot.description || lot.address || 'Описание отсутствует');
    previewText('lot-preview-price', formatPrice(lot.price));

    const details = [];
    if (lot.address) details.push('<b>Адрес:</b> ' + escapeHtml(lot.address));
    if (lot.cadastral_number) details.push('<b>Кадастр:</b> ' + escapeHtml(lot.cadastral_number));
    if (lot.procedure_number) details.push('<b>Процедура:</b> ' + escapeHtml(lot.procedure_number));
    if (lot.application_deadline) details.push('<b>Приём заявок до:</b> ' + escapeHtml(lot.application_deadline));
    if (lot.auction_at) details.push('<b>Торги:</b> ' + escapeHtml(lot.auction_at));
    document.getElementById('lot-preview-details').innerHTML = details.join('<br>');

    previewImages = Array.from(new Set((lot.image_urls || []).concat(lot.image_url || []).filter(Boolean)));
    previewImageIndex = 0;
    renderPreviewImage();
    renderPreviewLinks();
    setPreviewReviewStatus(lot.review_status || 'new');
}

document.getElementById('lot-preview-photo').addEventListener('error', function() {
    this.style.display = 'none'; document.getElementById('lot-preview-placeholder').style.display = 'flex';
});
document.getElementById('lot-preview-close').addEventListener('click', function() {
    document.getElementById('lot-preview').classList.remove('open');
    document.getElementById('lot-preview').setAttribute('aria-hidden', 'true');
    if (bankrotaiBridge) bankrotaiBridge.previewClosed(mapKind);
});
document.getElementById('lot-preview-prev').addEventListener('click', function() { movePreviewImage(-1); });
document.getElementById('lot-preview-next').addEventListener('click', function() { movePreviewImage(1); });
document.querySelectorAll('.related-link').forEach(function(button) {
    button.addEventListener('click', function() {
        const url = selectedPreviewLot && selectedPreviewLot[button.dataset.urlKey];
        if (!url) return;
        if (bankrotaiBridge) bankrotaiBridge.openSource(url); else window.open(url, '_blank');
    });
});
document.querySelectorAll('.review-button').forEach(function(button) {
    button.addEventListener('click', function() {
        if (!selectedPreviewLot || !bankrotaiBridge) return;
        const status = button.dataset.status;
        bankrotaiBridge.setReviewStatus(Number(selectedPreviewLot.id), status, function(ok) {
            if (ok) {
                selectedPreviewLot.review_status = status;
                setPreviewReviewStatus(status);
                if (window.applyLotReviewStatus) window.applyLotReviewStatus(Number(selectedPreviewLot.id), status);
            }
        });
    });
});

window.showLotPreview = showLotPreview;
window.updateLotPreviewExtras = function(lotId, extras) {
    if (!selectedPreviewLot || Number(selectedPreviewLot.id) !== Number(lotId)) return;
    Object.assign(selectedPreviewLot, extras || {});
    const extraImages = (extras && (extras.torgi_russia_image_urls || extras.image_urls)) || [];
    previewImages = Array.from(new Set(previewImages.concat(extraImages).filter(Boolean)));
    renderPreviewImage();
    renderPreviewLinks();
};
window.setLotReviewStatus = function(lotId, status) {
    if (selectedPreviewLot && Number(selectedPreviewLot.id) === Number(lotId)) {
        selectedPreviewLot.review_status = status;
        setPreviewReviewStatus(status);
    }
    if (window.applyLotReviewStatus) window.applyLotReviewStatus(Number(lotId), status);
};
"""


def extract_preview_image_urls(raw_data: object) -> list[str]:
    """Return unique safe web images found in heterogeneous source payloads."""
    preferred_keys = (
        "image_url", "photo_url", "thumbnail_url", "main_image", "image",
        "photo", "thumbnail", "image_urls", "photo_urls", "images", "photos", "gallery",
    )

    found: list[str] = []

    def collect(value: object, *, depth: int = 0) -> None:
        if depth > 4:
            return
        if isinstance(value, str):
            candidate = value.strip()
            if candidate.startswith("//"):
                candidate = "https:" + candidate
            parsed = urlparse(candidate)
            if parsed.scheme.lower() in {"http", "https"} and parsed.netloc:
                found.append(candidate)
            return
        if isinstance(value, (list, tuple)):
            for item in value:
                collect(item, depth=depth + 1)
            return
        if isinstance(value, dict):
            lowered = {str(key).lower(): item for key, item in value.items()}
            for key in preferred_keys:
                if key in lowered:
                    collect(lowered[key], depth=depth + 1)
            for key, item in lowered.items():
                if any(token in key for token in ("image", "photo", "thumb")):
                    collect(item, depth=depth + 1)

    collect(raw_data)
    return list(dict.fromkeys(found))


def extract_preview_image_url(raw_data: object) -> str | None:
    images = extract_preview_image_urls(raw_data)
    return images[0] if images else None


class MaxBidDialog(QDialog):
    def __init__(self, *, intended_bid: float = 0, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Калькулятор максимальной ставки")
        self.resize(620, 680)
        layout = QVBoxLayout(self)
        intro = QLabel(
            "Введите проверенную консервативную цену продажи и расходы. "
            "AI-оценка автоматически не подставляется."
        )
        intro.setWordWrap(True)
        layout.addWidget(intro)

        form = QFormLayout()
        self.fields: dict[str, QDoubleSpinBox] = {}
        definitions = (
            ("conservative_sale_price", "Консервативная цена продажи, ₽", 0, 1_000_000_000_000, 0),
            ("intended_bid", "Планируемая ставка, ₽", intended_bid, 1_000_000_000_000, 0),
            ("repair_cost", "Ремонт, ₽", 0, 1_000_000_000_000, 0),
            ("legal_cost", "Юридические расходы, ₽", 0, 1_000_000_000_000, 0),
            ("monthly_holding_cost", "Содержание в месяц, ₽", 0, 1_000_000_000_000, 0),
            ("holding_months", "Срок продажи, месяцев", 6, 120, 1),
            ("taxes", "Налоги и сборы, ₽", 0, 1_000_000_000_000, 0),
            ("sale_commission_percent", "Комиссия продажи, %", 0, 100, 2),
            ("target_profit", "Целевая прибыль, ₽", 0, 1_000_000_000_000, 0),
            ("risk_reserve", "Резерв риска, ₽", 0, 1_000_000_000_000, 0),
            ("annual_capital_cost_percent", "Стоимость капитала годовых, %", 0, 100, 2),
        )
        for key, label, value, maximum, decimals in definitions:
            widget = QDoubleSpinBox()
            widget.setRange(0, maximum)
            widget.setDecimals(decimals)
            widget.setValue(float(value or 0))
            widget.setGroupSeparatorShown(True)
            self.fields[key] = widget
            form.addRow(label, widget)
        layout.addLayout(form)

        calculate_button = QPushButton("Рассчитать безопасную ставку")
        calculate_button.clicked.connect(self.calculate)
        layout.addWidget(calculate_button)
        self.result = QTextEdit()
        self.result.setReadOnly(True)
        layout.addWidget(self.result)

        close_button = QPushButton("Закрыть")
        close_button.clicked.connect(self.accept)
        layout.addWidget(close_button)

    def calculate(self):
        values = {key: widget.value() for key, widget in self.fields.items()}
        if values["conservative_sale_price"] <= 0:
            QMessageBox.warning(self, "Исходные данные", "Укажите консервативную цену продажи.")
            return
        if values["holding_months"] <= 0:
            QMessageBox.warning(self, "Исходные данные", "Срок продажи должен быть больше нуля.")
            return
        if values["intended_bid"] <= 0:
            values["intended_bid"] = None
        scenarios = calculate_max_bid(MaxBidInputs(**values))
        labels = {"pessimistic": "Пессимистичный", "base": "Базовый", "optimistic": "Оптимистичный"}
        blocks = []
        for key in ("pessimistic", "base", "optimistic"):
            item = scenarios[key]
            lines = [
                f"{labels[key]} сценарий",
                f"Максимальная ставка: {item.maximum_bid:,.0f} ₽".replace(",", " "),
                f"Цена продажи: {item.sale_price:,.0f} ₽".replace(",", " "),
            ]
            if item.expected_profit is not None:
                lines.extend((
                    f"Ожидаемая прибыль: {item.expected_profit:,.0f} ₽".replace(",", " "),
                    f"ROI: {item.roi_percent:.2f}%",
                    f"Годовая доходность: {item.annualized_return_percent:.2f}%",
                    f"Точка безубыточности: {item.breakeven_sale_price:,.0f} ₽".replace(",", " "),
                ))
            blocks.append("\n".join(lines))
        self.result.setPlainText("\n\n".join(blocks))


class SortableTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        left = self.data(SORT_ROLE)
        right = other.data(SORT_ROLE)

        if left is None:
            left = self.text()
        if right is None:
            right = other.text()

        try:
            return left < right
        except Exception:
            return str(left).lower() < str(right).lower()


def make_text_item(text: str | None) -> QTableWidgetItem:
    value = text or ""
    item = SortableTableWidgetItem(value)
    item.setData(SORT_ROLE, value.lower())
    return item


def make_number_item(value) -> QTableWidgetItem:
    if value is None:
        item = SortableTableWidgetItem("")
        item.setData(SORT_ROLE, -1)
        return item

    try:
        numeric = float(value)
    except Exception:
        numeric = -1

    display = f"{numeric:,.0f}".replace(",", " ")
    item = SortableTableWidgetItem(display)
    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
    item.setData(SORT_ROLE, numeric)
    return item


def parse_date_sort_value(value):
    if not value:
        return datetime.min

    if isinstance(value, datetime):
        return value

    raw = str(value).strip()
    raw = raw.replace("Z", "")
    raw = re.sub(r"([+-]\d{2}:\d{2})$", "", raw)
    raw = raw.split(".000+")[0]
    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(raw, fmt)
        except Exception:
            pass

    return datetime.min


def make_date_item(value) -> QTableWidgetItem:
    sort_value = parse_date_sort_value(value)

    if not value:
        display = ""
    elif isinstance(value, datetime):
        display = value.strftime("%d.%m.%Y")
    else:
        display = str(value)

    item = SortableTableWidgetItem(display)
    item.setData(SORT_ROLE, sort_value)
    return item


class CollapsibleSection(QWidget):
    def __init__(self, title: str, parent=None, expanded: bool = False):
        super().__init__(parent)

        self.toggle_button = QToolButton()
        self.toggle_button.setText(title)
        self.toggle_button.setCheckable(True)
        self.toggle_button.setChecked(expanded)
        self.toggle_button.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.toggle_button.setArrowType(Qt.DownArrow if expanded else Qt.RightArrow)
        self.toggle_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.toggle_button.setStyleSheet("""
            QToolButton {
                border: none;
                background: #eef3fb;
                color: #143370;
                font-weight: 700;
                padding: 7px 6px;
                text-align: left;
                border-radius: 4px;
            }
            QToolButton:hover {
                background: #e2ebfb;
            }
        """)

        self.content = QWidget()
        self.content.setVisible(expanded)
        self.content.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)

        self.content_layout = QVBoxLayout(self.content)
        self.content_layout.setContentsMargins(8, 8, 8, 8)
        self.content_layout.setSpacing(8)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(4)
        root.addWidget(self.toggle_button)
        root.addWidget(self.content)

        self.toggle_button.clicked.connect(self._on_toggled)

    def _on_toggled(self, checked: bool):
        self.content.setVisible(checked)
        self.toggle_button.setArrowType(Qt.DownArrow if checked else Qt.RightArrow)

    def addWidget(self, widget: QWidget):
        self.content_layout.addWidget(widget)

    def addLayout(self, layout):
        self.content_layout.addLayout(layout)


AI_PROVIDER_OPTIONS = [
    ("gemini", "Google Gemini"),
    ("groq", "GroqCloud"),
    ("grok", "Grok / xAI"),
    ("github", "GitHub Models"),
    ("omniroute", "Kiro через OmniRoute"),
    ("nvidia", "NVIDIA"),
]

AI_MODEL_OPTIONS = {
    "omniroute": [
        ("kr/claude-sonnet-4", "Kiro: Claude Sonnet 4"),
    ],
    "nvidia": [
        ("nvidia/llama-3.3-nemotron-super-49b-v1", "NVIDIA: Llama 3.3 Nemotron Super 49B"),
        ("meta/llama-3.3-70b-instruct", "NVIDIA: Llama 3.3 70B Instruct"),
    ],
    "gemini": [
        ("gemini-2.5-flash", "Gemini 2.5 Flash"),
        ("gemini-2.5-pro", "Gemini 2.5 Pro"),
        ("gemini-2.0-flash", "Gemini 2.0 Flash"),
    ],
    "grok": [
        ("grok-4", "Grok 4"),
        ("grok-4-fast-reasoning", "Grok 4 Fast Reasoning"),
        ("grok-3-mini", "Grok 3 Mini"),
    ],
    "groq": [
        ("llama-3.3-70b-versatile", "Groq: Llama 3.3 70B Versatile"),
        ("openai/gpt-oss-120b", "Groq: GPT-OSS 120B"),
        ("llama-3.1-8b-instant", "Groq: Llama 3.1 8B Instant"),
    ],
    "github": [
        ("openai/gpt-4.1-mini", "GitHub: GPT-4.1 Mini"),
        ("openai/gpt-4.1", "GitHub: GPT-4.1"),
        ("xai/grok-3-mini", "GitHub: Grok 3 Mini"),
        ("meta/Llama-4-Scout-17B-16E-Instruct", "GitHub: Llama 4 Scout"),
    ],
}

def translate_status(status: str) -> str: 
    mapping = { 
        "active": "Активен", 
        "closed": "Завершен", 
        "pending": "Ожидается", 
        "scheduled": "Запланирован", 
        "unknown": "Неизвестно", 
    } 
    return mapping.get(status, status) 

def translate_category(category: str) -> str: 
    mapping = { 
        "apartment": "Квартира", 
        "house": "Жилой дом", 
        "commercial": "Коммерческая недв.", 
        "commercial_room": "Нежилое помещение", 
        "commercial_building": "Нежилое здание", 
        "commercial_building_with_land": "Нежилое здание + ЗУ", 
        "complex": "Имущ. комплекс", 
        "office": "Офис", 
        "retail": "Торговое", 
        "land": "Земельный участок", 
        "car": "Автомобиль", 
        "transport": "Транспорт", 
        "vehicle": "Спецтехника", 
        "equipment": "Оборудование", 
        "parking": "Парковка", 
        "unfinished": "Недострой", 
        "receivable": "Права требования", 
        "real_estate": "Недвижимость", 
        "other": "Прочее", 
    } 
    return mapping.get(category, category) 

# --- Background Workers ---

class SyncWorker(QThread):
    finished = Signal(int)
    progress = Signal(str)
    progress_percent = Signal(int)
    error = Signal(str)

    def __init__(self, region: str):
        super().__init__()
        self.region = region

    def run(self):
        try:
            self.progress.emit(f"Запуск синхронизации для региона: {self.region}...")
            with session_scope() as session:
                self.progress.emit("Загрузка данных с GorodTorgi (может занять 1-2 минуты)...")
                lots_gt = sync_public_real_estate(session, self.region)
                
                self.progress.emit("Загрузка данных с TBankrot...")
                lots_tb = ingest_recent_tbankrot(session, self.region)
                
                total = len(lots_gt) + len(lots_tb)
                self.finished.emit(total)
        except Exception as e:
            logger.error(f"SyncWorker error: {e}")
            self.error.emit(str(e))


class TorgiGovSearchWorker(QThread):
    progress = Signal(str)
    progress_percent = Signal(int)
    page_loaded = Signal(list, dict)
    finished = Signal(list, dict)
    error = Signal(str)

    def __init__(
        self,
        filters: TorgiGovSearchFilters,
        *,
        load_all: bool = False,
        max_items: int | None = 5000,
        use_excel: bool = False,
    ):
        super().__init__()
        self.filters = filters
        self.load_all = load_all
        self.max_items = max_items
        self.use_excel = use_excel
        self._stop_requested = False

    def request_stop(self):
        self._stop_requested = True

    def _persist_lots(self, lots: list[NormalizedLot]) -> None:
        if not lots:
            return
        with session_scope() as session:
            for lot in lots:
                persist_lot(session, lot)

    def run(self):
        try:
            self.progress.emit("Подключение к torgi.gov.ru...")
            self.progress_percent.emit(3)

            client = TorgiGovClient(diagnostics=True)

            if self.use_excel:
                self.progress.emit("Скачивание Excel-выгрузки torgi.gov.ru...")
                self.progress_percent.emit(25)
                lots, meta = client.search_lots_excel(self.filters)
                self.progress.emit("Импорт строк из Excel-выгрузки...")
                self.progress_percent.emit(75)
                self._persist_lots(lots)
                self.page_loaded.emit(lots, meta)
            elif self.load_all:
                self.progress.emit("Загрузка всех страниц torgi.gov.ru...")

                def report(page: int, total: int | None, loaded: int):
                    if total:
                        percent = min(99, int((loaded / max(total, 1)) * 100))
                        self.progress_percent.emit(percent)
                        self.progress.emit(f"Загружено {loaded}/{total} лотов, страница {page}...")
                    else:
                        self.progress_percent.emit(min(99, 5 + page))
                        self.progress.emit(f"Загружено {loaded} лотов, страница {page}...")

                def page_loaded(lots_on_page: list[NormalizedLot], page_meta: dict):
                    self.page_loaded.emit(lots_on_page, page_meta)

                lots, meta = client.search_all_lots(
                    self.filters,
                    max_items=self.max_items,
                    progress_cb=report,
                    page_cb=page_loaded,
                    stop_cb=lambda: self._stop_requested,
                )
            else:
                self.progress.emit("Загрузка страницы онлайн-лотов...")
                self.progress_percent.emit(35)
                lots, meta = client.search_lots(self.filters)
                self._persist_lots(lots)
                self.page_loaded.emit(lots, meta)

            self.progress_percent.emit(100)
            self.finished.emit(lots, meta)
        except Exception as e:
            logger.exception("TorgiGovSearchWorker error")
            self.error.emit(str(e))


class TBankrotSearchWorker(QThread):
    progress = Signal(str)
    progress_percent = Signal(int)
    page_loaded = Signal(list, dict)
    finished = Signal(list, dict)
    error = Signal(str)

    def __init__(
        self,
        filters: TBankrotSearchFilters,
        *,
        load_all: bool = False,
        max_items: int | None = 5000,
    ):
        super().__init__()
        self.filters = filters
        self.load_all = load_all
        self.max_items = max_items
        self._stop_requested = False

    def request_stop(self):
        self._stop_requested = True

    def _persist_lots(self, lots: list[NormalizedLot]) -> None:
        if not lots:
            return
        with session_scope() as session:
            for lot in lots:
                persist_lot(session, lot)

    def run(self):
        try:
            self.progress.emit("Подключение к tbankrot.ru...")
            self.progress_percent.emit(3)
            client = TBankrotClient(diagnostics=True)

            if self.load_all:
                self.progress.emit("Загрузка всех страниц TBankrot...")

                def report(page: int, _total: int | None, loaded: int):
                    self.progress_percent.emit(min(99, 5 + page))
                    self.progress.emit(f"Загружено {loaded} лотов, страница {page}...")

                def page_loaded(lots_on_page: list[NormalizedLot], page_meta: dict):
                    self._persist_lots(lots_on_page)
                    self.page_loaded.emit(lots_on_page, page_meta)

                lots, meta = client.search_all_lots(
                    self.filters,
                    max_items=self.max_items,
                    progress_cb=report,
                    page_cb=page_loaded,
                    stop_cb=lambda: self._stop_requested,
                )
            else:
                self.progress.emit("Загрузка страницы TBankrot...")
                self.progress_percent.emit(35)
                lots, meta = client.search_filtered_lots(self.filters)
                self.page_loaded.emit(lots, meta)

            self.progress_percent.emit(100)
            self.finished.emit(lots, meta)
        except Exception as e:
            logger.exception("TBankrotSearchWorker error")
            self.error.emit(str(e))


class LotOnlineSearchWorker(QThread):
    progress = Signal(str)
    progress_percent = Signal(int)
    page_loaded = Signal(list, dict)
    finished = Signal(list, dict)
    error = Signal(str)

    def __init__(
        self,
        filters: LotOnlineSearchFilters,
        *,
        load_all: bool = False,
        max_items: int | None = 5000,
    ) -> None:
        super().__init__()
        self.filters = filters
        self.load_all = load_all
        self.max_items = max_items
        self._stop_requested = False

    def request_stop(self) -> None:
        self._stop_requested = True

    @staticmethod
    def _persist_lots(lots: list[NormalizedLot]) -> None:
        if not lots:
            return
        with session_scope() as session:
            for lot in lots:
                persist_lot(session, lot)

    def run(self) -> None:
        try:
            self.progress.emit("Подключение к РАД / ЛОТ-ОНЛАЙН...")
            self.progress_percent.emit(3)
            client = LotOnlineClient(diagnostics=True)
            if self.load_all:
                def report(page: int, total: int | None, loaded: int) -> None:
                    percent = int(page / max(total or page + 1, 1) * 100)
                    self.progress_percent.emit(min(99, max(5, percent)))
                    self.progress.emit(f"ЛОТ-ОНЛАЙН: загружено {loaded}, страница {page}...")

                def page_loaded(lots_on_page: list[NormalizedLot], page_meta: dict) -> None:
                    self._persist_lots(lots_on_page)
                    self.page_loaded.emit(lots_on_page, page_meta)

                lots, meta = client.search_all_lots(
                    self.filters,
                    max_items=self.max_items,
                    progress_cb=report,
                    page_cb=page_loaded,
                    stop_cb=lambda: self._stop_requested,
                )
            else:
                self.progress_percent.emit(35)
                lots, meta = client.search_lots(self.filters)
                self.page_loaded.emit(lots, meta)
            self.progress_percent.emit(100)
            self.finished.emit(lots, meta)
        except Exception as exc:
            logger.exception("LotOnlineSearchWorker error")
            self.error.emit(str(exc))


class AllRussiaRealEstateWorker(QThread):
    progress = Signal(str)
    source_finished = Signal(str, int, int)
    result_ready = Signal(object, object)
    error = Signal(str)

    def __init__(self) -> None:
        super().__init__()
        self._stop_requested = False
        self._processed_ids: set[int] = set()
        self._source_counts: dict[str, int] = {}
        self._source_errors: dict[str, str] = {}

    def request_stop(self) -> None:
        self._stop_requested = True
        self.requestInterruption()

    def _persist_page(self, source_name: str, lots: list[NormalizedLot]) -> None:
        if not lots:
            return
        with session_scope() as session:
            for lot in lots:
                primary = persist_lot(session, lot)
                self._processed_ids.add(primary.id)
        self._source_counts[source_name] = self._source_counts.get(source_name, 0) + len(lots)
        self.progress.emit(
            f"{source_name}: загружено {self._source_counts[source_name]}, "
            f"уникальных карточек {len(self._processed_ids)}"
        )

    def run(self) -> None:
        sources = (
            (
                "ГИС Торги",
                TorgiGovClient(diagnostics=True),
                TorgiGovSearchFilters(
                    type_transaction="SALE",
                    category_code=TorgiGovClient.REAL_ESTATE_CATEGORY_CODES,
                    lot_status=TorgiGovClient.DEFAULT_LOT_STATUS,
                    page=1,
                    page_size=100,
                ),
            ),
            (
                "TBankrot",
                TBankrotClient(diagnostics=True),
                TBankrotSearchFilters(
                    category_codes=TBankrotClient.REAL_ESTATE_CATEGORY_CODES,
                    page=1,
                    page_size=100,
                ),
            ),
            (
                "РАД / ЛОТ-ОНЛАЙН",
                LotOnlineClient(diagnostics=True),
                LotOnlineSearchFilters(
                    category_id=LotOnlineClient.DEFAULT_CATEGORY_ID,
                    region_feature=None,
                    archive_mode="false",
                    page=1,
                    page_size=96,
                ),
            ),
        )
        for source_name, client, filters in sources:
            if self._stop_requested:
                break
            self.progress.emit(f"{source_name}: поиск всей недвижимости РФ...")
            try:
                page_callback = lambda lots, _meta, name=source_name: self._persist_page(name, lots)
                if isinstance(client, TorgiGovClient):
                    lots, _meta = client.search_all_lots(
                        filters,
                        max_items=None,
                        max_pages=500,
                        page_cb=page_callback,
                        stop_cb=lambda: self._stop_requested,
                    )
                elif isinstance(client, TBankrotClient):
                    lots, _meta = client.search_all_lots(
                        filters,
                        max_items=None,
                        page_cb=page_callback,
                        stop_cb=lambda: self._stop_requested,
                    )
                else:
                    lots, _meta = client.search_all_lots(
                        filters,
                        max_items=None,
                        max_pages=500,
                        page_cb=page_callback,
                        stop_cb=lambda: self._stop_requested,
                    )
                self._source_counts[source_name] = len(lots)
                self.source_finished.emit(source_name, len(lots), len(self._processed_ids))
            except Exception as exc:
                logger.exception("Nationwide search failed for %s", source_name)
                self._source_errors[source_name] = str(exc)
                self.progress.emit(f"{source_name}: ошибка, продолжаю со следующим источником")

        if not self._stop_requested:
            with session_scope() as session:
                reconcile_cross_source_duplicates(session)
                rows = session.execute(
                    select(ProcessedLot.id, ProcessedLot.duplicate_of_id).where(
                        ProcessedLot.id.in_(self._processed_ids)
                    )
                ).all()
                self._processed_ids = {
                    int(duplicate_of_id or processed_id)
                    for processed_id, duplicate_of_id in rows
                }
        if self._source_errors and not self._source_counts:
            self.error.emit("; ".join(f"{name}: {message}" for name, message in self._source_errors.items()))
            return
        self.result_ready.emit(sorted(self._processed_ids), {
            "sources": self._source_counts,
            "errors": self._source_errors,
            "stopped": self._stop_requested,
        })


class ImportWorker(QThread):
    finished = Signal(int, int, int) # new, updated, skipped
    progress = Signal(str)
    progress_percent = Signal(int)
    error = Signal(str)

    def __init__(self, file_path: str, region: str):
        super().__init__()
        self.file_path = file_path
        self.region = region

    def run(self):
        try:
            with session_scope() as session:
                # skip_geo=True — сильно ускоряет импорт
                new, upd, skip = import_manual_html(
                    session, 
                    self.file_path, 
                    self.region, 
                    skip_geo=True,          # ← главное ускорение
                    progress_cb=self._report_progress
                )
                self.finished.emit(new, upd, skip)
        except Exception as e:
            logger.error(f"ImportWorker error: {e}")
            self.error.emit(str(e))

    def _report_progress(self, current, total):
        self.progress.emit(f"Обработано: {current}/{total}")

    def _report_progress_percent(self, current, total):
        if total:
            self.progress_percent.emit(int((current / total) * 100))

class AIWorker(QThread):
    finished = Signal(int)
    lot_finished = Signal(int, int)
    lot_failed = Signal(int, str, int)
    progress = Signal(str)
    progress_percent = Signal(int)
    error = Signal(str)

    def __init__(self, appraiser: OpenAIAppraiser, lot_ids: list[int] | None = None, limit: int = 10):
        super().__init__()
        self.appraiser = appraiser
        self.lot_ids = lot_ids
        self.limit = limit

    def run(self):
        try:
            from bankrotai.db import DB_WRITE_LOCK, find_unappraised_lots
            import time

            processed_count = 0
            failed_count = 0

            with session_scope() as session:
                if self.lot_ids:
                    lot_ids = list(self.lot_ids)
                else:
                    lots = find_unappraised_lots(session, limit=self.limit)
                    lot_ids = [lot.id for lot in lots]

            if not lot_ids:
                self.progress_percent.emit(100)
                self.finished.emit(0, 0)
                return

            total = len(lot_ids)
            for i, lot_id in enumerate(lot_ids):
                self.progress_percent.emit(int((i / total) * 100))

                try:
                    with session_scope() as session:
                        lot = session.get(ProcessedLot, lot_id)
                        if not lot:
                            continue
                        self.progress.emit(f"AI Оценка [{i+1}/{total}]: {lot.title[:40]}...")
                        nl = NormalizedLot.from_processed_lot(lot)

                    evaluation = self.appraiser.evaluate(nl)

                    with DB_WRITE_LOCK:
                        with session_scope() as session:
                            lot = session.get(ProcessedLot, lot_id)
                            if not lot:
                                continue
                            apply_evaluation_to_lot(lot, evaluation)
                            try:
                                cache_key = self.appraiser._get_cache_key(nl)
                                self.appraiser._save_evaluation_to_db(session, nl, evaluation, cache_key)
                            except Exception:
                                logger.exception("Failed to save valuation run for lot %s", lot_id)
                            processed_count += 1

                    self.lot_finished.emit(lot_id, processed_count)
                except Exception as lot_error:
                    failed_count += 1
                    message = str(lot_error)
                    logger.exception("AI evaluation failed for lot %s; continuing batch", lot_id)
                    self.progress.emit(
                        f"Ошибка AI оценки лота ID {lot_id}; продолжаю. Ошибок: {failed_count}"
                    )
                    self.lot_failed.emit(lot_id, message, failed_count)

                self.progress_percent.emit(int(((i + 1) / total) * 100))
                if i < total - 1:
                    time.sleep(2)

            self.finished.emit(processed_count)
        except Exception as e:
            logger.error(f"AIWorker error: {e}")
            provider = getattr(getattr(self.appraiser, "provider", None), "provider", "")
            message = str(e)
            if any(marker in message.lower() for marker in ("connection error", "apiconnectionerror", "timed out", "timeout")):
                message = (
                    f"AI connection error ({provider or 'unknown provider'}): {message}\n\n"
                    "Провайдер недоступен из текущей сети. Для NVIDIA часто требуется зарубежный VPN; "
                    "также можно выбрать OmniRoute/Kiro в настройках AI и проверить API-ключ."
                )
            self.error.emit(message)


class GeoWorker(QThread):
    finished = Signal(int, int, int)
    progress = Signal(str)
    progress_percent = Signal(int)
    lot_processed = Signal(int, bool, int, int)
    error = Signal(str)

    def __init__(
        self,
        limit: int | None = 500,
        lot_ids: list[int] | None = None,
        refresh_existing: bool = False,
    ):
        super().__init__()
        self.limit = limit
        self.lot_ids = lot_ids
        self.refresh_existing = refresh_existing

    def run(self):
        try:
            from bankrotai.db import DB_WRITE_LOCK
            from bankrotai.geo import apply_lot_geo_result, resolve_lot_geo
            from sqlalchemy import exists

            success_count = 0
            failed_count = 0
            skipped_existing = 0

            with session_scope() as session:
                if self.lot_ids:
                    requested_ids = list(dict.fromkeys(int(item) for item in self.lot_ids))
                    if self.refresh_existing:
                        lot_ids = requested_ids
                    else:
                        existing_ids = set(session.scalars(
                            select(LotGeoSnapshot.lot_id).where(LotGeoSnapshot.lot_id.in_(requested_ids))
                        ).all())
                        lot_ids = [item for item in requested_ids if item not in existing_ids]
                        skipped_existing = len(requested_ids) - len(lot_ids)
                else:
                    stmt = select(ProcessedLot.id).where(
                        ~exists().where(LotGeoSnapshot.lot_id == ProcessedLot.id)
                    )
                    if self.limit:
                        stmt = stmt.limit(self.limit)
                    lot_ids = list(session.scalars(stmt).all())

            if not lot_ids:
                self.progress_percent.emit(100)
                self.finished.emit(0, 0, skipped_existing)
                return

            with session_scope() as session:
                rows = session.execute(
                    select(
                        ProcessedLot.id,
                        ProcessedLot.cadastral_number,
                        ProcessedLot.address,
                        ProcessedLot.title,
                        ProcessedLot.region_name,
                    ).where(ProcessedLot.id.in_(lot_ids))
                ).all()
            payloads = [
                {
                    "lot_id": row.id,
                    "cadastral_number": row.cadastral_number,
                    "address": row.address,
                    "title": row.title,
                    "region_name": row.region_name,
                }
                for row in rows
            ]
            total = len(payloads)
            max_workers = min(get_settings().geo_max_workers, max(total, 1))
            self.progress.emit(f"Параллельное геокодирование: {total} лотов, потоков: {max_workers}")

            def resolve(payload: dict):
                return payload["lot_id"], resolve_lot_geo(
                    payload["cadastral_number"],
                    payload["address"],
                    title=payload["title"],
                    region_name=payload["region_name"],
                )

            with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="bankrotai-geo") as executor:
                futures = {executor.submit(resolve, payload): payload for payload in payloads}
                for completed, future in enumerate(as_completed(futures), start=1):
                    payload = futures[future]
                    lot_id = int(payload["lot_id"])
                    try:
                        _resolved_lot_id, geo_result = future.result()
                    except Exception:
                        logger.exception("GEO resolution failed for lot %s", lot_id)
                        geo_result = None

                    success = False
                    with DB_WRITE_LOCK:
                        with session_scope() as session:
                            lot = session.get(ProcessedLot, lot_id)
                            if lot:
                                if self.refresh_existing and geo_result:
                                    session.query(LotGeoSnapshot).filter_by(lot_id=lot.id).delete()
                                    session.flush()
                                success = apply_lot_geo_result(session, lot, geo_result)

                    if success:
                        success_count += 1
                    else:
                        failed_count += 1
                    self.lot_processed.emit(lot_id, success, completed, total)
                    self.progress_percent.emit(int((completed / total) * 100))
                    label = (payload["address"] or payload["cadastral_number"] or payload["title"] or "")[:45]
                    self.progress.emit(
                        f"GEO [{completed}/{total}] ✓ {success_count}, без координат {failed_count}: {label}"
                    )

            self.finished.emit(success_count, failed_count, skipped_existing)
        except Exception as e:
            logger.error(f"GeoWorker error: {e}")
            self.error.emit(str(e))


class CadastreSearchWorker(QThread):
    finished = Signal(object)
    error = Signal(str)

    def __init__(self, query: str):
        super().__init__()
        self.query = query

    def run(self):
        try:
            from bankrotai.geo import CADASTRAL_GEOCODER
            self.finished.emit(CADASTRAL_GEOCODER.search(self.query))
        except Exception as e:
            logger.exception("Cadastre search worker failed")
            self.error.emit(str(e))


class PreviewEnrichmentWorker(QThread):
    result_ready = Signal(int, object)
    failed = Signal(int, str)

    def __init__(self, lot_id: int):
        super().__init__()
        self.lot_id = int(lot_id)

    def run(self):
        try:
            from bankrotai.db import DB_WRITE_LOCK
            from bankrotai.torgi_russia import TorgiRussiaClient

            with session_scope() as session:
                lot = session.get(ProcessedLot, self.lot_id)
                if lot is None:
                    self.result_ready.emit(self.lot_id, {})
                    return
                source_lot = session.scalars(
                    select(SourceLot).where(or_(
                        SourceLot.processed_lot_id == lot.id,
                        and_(
                            SourceLot.source_system == lot.source_system,
                            SourceLot.external_id == lot.external_id,
                        ),
                    ))
                ).first()
                raw = dict(source_lot.raw_data or {}) if source_lot else {}
                related: dict[str, str] = {}
                if source_lot:
                    sibling_sources = session.scalars(
                        select(SourceLot).where(SourceLot.canonical_lot_id == source_lot.canonical_lot_id)
                    ).all()
                    for sibling in sibling_sources:
                        system = (sibling.source_system or "").lower()
                        if sibling.source_url and "torgi.gov" in system:
                            related["gis_torgi_url"] = sibling.source_url
                        if sibling.source_url and "lot-online" in system:
                            related["etp_url"] = sibling.source_url
                cached = {
                    key: raw.get(key)
                    for key in (
                        "torgi_russia_url", "gis_torgi_url", "etp_url",
                        "torgi_russia_image_urls",
                    )
                    if raw.get(key)
                }
                cached.update(related)
                if raw.get("torgi_russia_checked_at"):
                    self.result_ready.emit(self.lot_id, cached)
                    return
                cadastres = list(lot.cadastral_numbers or [])
                if lot.cadastral_number:
                    cadastres.insert(0, lot.cadastral_number)
                source_lot_id = source_lot.id if source_lot else None

            if self.isInterruptionRequested():
                return
            details = TorgiRussiaClient(timeout=8).find_by_cadastral_numbers(cadastres) if cadastres else None
            extras = details.as_dict() if details else {}
            extras.update(related)
            if details and details.procedure_number:
                with session_scope() as session:
                    etp_source = session.scalars(
                        select(SourceLot).where(
                            SourceLot.procedure_number == details.procedure_number,
                            SourceLot.source_system.ilike("%lot-online%"),
                        )
                    ).first()
                    if etp_source and etp_source.source_url:
                        extras["etp_url"] = etp_source.source_url
            if source_lot_id:
                with DB_WRITE_LOCK:
                    with session_scope() as session:
                        source_lot = session.get(SourceLot, source_lot_id)
                        if source_lot:
                            updated = dict(source_lot.raw_data or {})
                            updated.update(extras)
                            updated["torgi_russia_checked_at"] = datetime.now().isoformat(timespec="seconds")
                            source_lot.raw_data = updated
            if not self.isInterruptionRequested():
                self.result_ready.emit(self.lot_id, extras)
        except Exception as exc:
            logger.warning("Torgi Rossii preview enrichment failed for lot %s: %s", self.lot_id, exc)
            if not self.isInterruptionRequested():
                self.failed.emit(self.lot_id, str(exc))


NSPD_REFERER = "https://nspd.gov.ru/map?theme_id=1&is_copy_url=true&active_layers=&baseLayerId="
NSPD_WMS_LAYERS = {
    "land": "36048",
    "buildings": "36328",
}


class CadastralWmsProxyHandler(BaseHTTPRequestHandler):
    timeout = 12

    def log_message(self, format, *args):
        logger.debug("WMS proxy: " + format, *args)

    def do_GET(self):
        parsed = urlparse(self.path)
        layer_key = parsed.path.rsplit("/", 1)[-1]
        layer_id = NSPD_WMS_LAYERS.get(layer_key)

        if not layer_id:
            self.send_error(404, "Unknown cadastral layer")
            return

        params = dict(parse_qsl(parsed.query, keep_blank_values=True))
        params.setdefault("SERVICE", "WMS")
        params.setdefault("VERSION", "1.3.0")
        params.setdefault("REQUEST", "GetMap")
        params.setdefault("FORMAT", "image/png")
        params.setdefault("TRANSPARENT", "true")
        params["LAYERS"] = layer_id

        upstream = f"https://nspd.gov.ru/api/aeggis/v4/{layer_id}/wms?{urlencode(params)}"
        headers = {
            "Referer": NSPD_REFERER,
            "User-Agent": "Mozilla/5.0 BankrotAI/1.0",
            "Accept": "image/png,image/*,*/*;q=0.8",
        }

        try:
            import requests
            response = requests.get(upstream, headers=headers, timeout=(3.05, 8), verify=nspd_tls_verify())
            response.raise_for_status()
        except Exception as e:
            logger.warning("NSPD WMS request failed for %s: %s", layer_key, e)
            self.send_response(204)
            self.end_headers()
            return

        content_type = response.headers.get("Content-Type", "image/png")
        self.send_response(response.status_code)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "public, max-age=300")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(response.content)


def _free_local_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


class MapBridge(QObject):
    review_changed = Signal(int, str)
    preview_opened = Signal(str, int)
    preview_closed = Signal(str)

    @Slot(int, str, result=bool)
    def setReviewStatus(self, lot_id: int, status: str) -> bool:
        if status not in {"approved", "maybe", "rejected"}:
            return False
        with session_scope() as session:
            lot = session.get(ProcessedLot, int(lot_id))
            if lot is None:
                return False
            lot.review_status = status
        self.review_changed.emit(int(lot_id), status)
        return True

    @Slot(str, result=bool)
    def openSource(self, source_url: str) -> bool:
        url = QUrl.fromUserInput(source_url)
        if url.scheme().lower() not in {"http", "https"}:
            return False
        return bool(QDesktopServices.openUrl(url))

    @Slot(str, int)
    def previewOpened(self, map_kind: str, lot_id: int):
        self.preview_opened.emit(map_kind, int(lot_id))

    @Slot(str)
    def previewClosed(self, map_kind: str):
        self.preview_closed.emit(map_kind)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("BankrotAI - Аналитика торгов")
        self.resize(1300, 900)
        self.current_selected_lot_id = None
        self._appraiser = None
        self.cadastre_search_worker = None
        self.cadastral_wms_proxy = None
        self.cadastral_wms_proxy_port = None
        self.torgi_results: list[NormalizedLot] = []
        self.torgi_meta: dict = {}
        self.torgi_current_page = 1
        self.torgi_unsupported_warnings: list[str] = []
        self.tbankrot_results: list[NormalizedLot] = []
        self.tbankrot_meta: dict = {}
        self.tbankrot_current_page = 1
        self._last_tbankrot_sort_column = None
        self._last_tbankrot_sort_order = Qt.AscendingOrder
        self.lot_online_results: list[NormalizedLot] = []
        self.lot_online_meta: dict = {}
        self.lot_online_current_page = 1
        init_db()
        with session_scope() as session:
            reconciled = reconcile_cross_source_duplicates(session)
            if reconciled:
                logger.info("Reconciled %s existing cross-source duplicate lots", reconciled)
        self.map_bridge = MapBridge(self)
        self.map_bridge.review_changed.connect(self.on_map_review_changed)
        self.map_bridge.preview_opened.connect(self.on_map_preview_opened)
        self.map_bridge.preview_closed.connect(self.on_map_preview_closed)
        self.preview_enrichment_workers: dict[int, PreviewEnrichmentWorker] = {}
        self.start_cadastral_wms_proxy()
        self.init_statusbar()
        self.init_ui()
        self.update_dashboard()

    def start_cadastral_wms_proxy(self):
        if self.cadastral_wms_proxy_port:
            return

        port = _free_local_port()
        server = ThreadingHTTPServer(("127.0.0.1", port), CadastralWmsProxyHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()

        self.cadastral_wms_proxy = server
        self.cadastral_wms_proxy_port = port
        logger.info("Local cadastral WMS proxy started on 127.0.0.1:%s", port)

    def init_statusbar(self):
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximumHeight(15)
        self.progress_bar.setMaximumWidth(200)
        self.progress_bar.setVisible(False)
        self.status_bar.addPermanentWidget(self.progress_bar)
        self.status_bar.showMessage("Система готова")

    def start_task_progress(self, key: str, label_text: str):
        if not hasattr(self, "task_progress_widgets"):
            self.task_progress_widgets = {}
        self.finish_task_progress(key)
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(6, 0, 0, 0)
        layout.setSpacing(4)
        label = QLabel(label_text)
        bar = QProgressBar()
        bar.setRange(0, 100)
        bar.setValue(0)
        bar.setMaximumHeight(15)
        bar.setMaximumWidth(150)
        layout.addWidget(label)
        layout.addWidget(bar)
        self.status_bar.addPermanentWidget(widget)
        self.task_progress_widgets[key] = (widget, bar)

    def update_task_progress(self, key: str, value: int):
        item = getattr(self, "task_progress_widgets", {}).get(key)
        if item:
            item[1].setValue(max(0, min(100, int(value))))

    def finish_task_progress(self, key: str):
        item = getattr(self, "task_progress_widgets", {}).pop(key, None)
        if not item:
            return
        widget, _bar = item
        self.status_bar.removeWidget(widget)
        widget.deleteLater()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #dcdcdc; top: -1px; background: white; }
            QTabBar::tab { background: #f0f0f0; border: 1px solid #dcdcdc; padding: 10px 20px; min-width: 120px; }
            QTabBar::tab:selected { background: white; border-bottom-color: white; font-weight: bold; }
        """)
        main_layout.addWidget(self.tabs)

        # 1. Torgi.gov.ru Search Tab
        self.dash_tab = QWidget()
        self.init_dash_tab()
        self.tabs.addTab(self.dash_tab, "Поиск ГИС Торги")

        # 2. TBankrot Search Tab
        self.tbankrot_tab = QWidget()
        self.init_tbankrot_tab()
        self.tabs.addTab(self.tbankrot_tab, "Поиск Т Банкрот")

        # 3. RAD / LOT-ONLINE Search Tab
        self.lot_online_tab = QWidget()
        self.init_lot_online_tab()
        self.tabs.addTab(self.lot_online_tab, "Поиск РАД / ЛОТ-ОНЛАЙН")

        # 4. Registry Tab
        self.registry_tab = QWidget()
        self.init_registry_tab()
        self.tabs.addTab(self.registry_tab, "📋 Реестр лотов")

        # 4. Map Tab
        self.map_tab = QWidget()
        self.init_map_tab()
        self.tabs.addTab(self.map_tab, "🗺️ Карта и Кадастр")

        # 5. Yandex Map Tab
        self.yandex_map_tab = QWidget()
        self.init_yandex_map_tab()
        self.tabs.addTab(self.yandex_map_tab, "Карта и кадастр Яндекс")

        # 6. Tools Tab
        self.tools_tab = QWidget()
        self.init_tools_tab()
        self.tabs.addTab(self.tools_tab, "🛠️ Инструменты")

    @staticmethod
    def _configure_results_table(table: QTableWidget, headers: list[str], stretch_columns: tuple[int, ...]) -> None:
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        header = table.horizontalHeader()
        for column in stretch_columns:
            header.setSectionResizeMode(column, QHeaderView.Stretch)
        header.setSectionsClickable(True)
        header.setSortIndicatorShown(True)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(True)

    def _init_compact_torgi_tab(self) -> None:
        layout = QVBoxLayout(self.dash_tab)
        layout.setContentsMargins(8, 8, 8, 8)
        title = QLabel("Поиск ГИС Торги (torgi.gov.ru)")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #143370;")
        layout.addWidget(title)
        self.stats_label = QLabel()
        self.stats_label.hide()
        group = QGroupBox("Параметры поиска недвижимости (только продажа)")
        form = QFormLayout(group)
        self.torgi_search_input = QLineEdit()
        self.torgi_search_input.setPlaceholderText("Название, адрес, кадастровый номер...")
        form.addRow("Поиск", self.torgi_search_input)
        self.torgi_category_combo = WheelSafeComboBox()
        for label, code in (
            ("Вся недвижимость", TorgiGovClient.REAL_ESTATE_CATEGORY_CODES),
            ("Земельный участок со зданием", "903"),
            ("Недвижимость", "7"),
            ("Земельные участки", "2"),
        ):
            self.torgi_category_combo.addItem(label, code)
        form.addRow("Категория", self.torgi_category_combo)
        self.torgi_subject_combo = WheelSafeComboBox()
        self.torgi_subject_combo.addItem("Все регионы", None)
        for name, code in sorted(TorgiGovClient.SUBJECT_RF_CODES.items()):
            self.torgi_subject_combo.addItem(name, code)
        form.addRow("Регион", self.torgi_subject_combo)
        self.torgi_lot_status_combo = WheelSafeComboBox()
        self.torgi_lot_status_combo.addItem("Активные", TorgiGovClient.DEFAULT_LOT_STATUS)
        self.torgi_lot_status_combo.addItem("Все состояния", None)
        self.torgi_status_combo = self.torgi_lot_status_combo
        form.addRow("Состояние", self.torgi_lot_status_combo)
        self.torgi_price_min_input, self.torgi_price_max_input = QLineEdit(), QLineEdit()
        self.torgi_price_min_input.setPlaceholderText("Цена от")
        self.torgi_price_max_input.setPlaceholderText("Цена до")
        price_row = QHBoxLayout()
        price_row.addWidget(self.torgi_price_min_input)
        price_row.addWidget(self.torgi_price_max_input)
        form.addRow("Цена", price_row)
        self.torgi_load_all_checkbox = QCheckBox("Загрузить все страницы")
        self.torgi_load_all_checkbox.setChecked(True)
        self.torgi_max_items_input = QLineEdit("5000")
        self.torgi_max_items_input.setMaximumWidth(110)
        mode_row = QHBoxLayout()
        mode_row.addWidget(self.torgi_load_all_checkbox)
        mode_row.addWidget(QLabel("Лимит:"))
        mode_row.addWidget(self.torgi_max_items_input)
        mode_row.addStretch()
        form.addRow("Режим", mode_row)
        buttons = QHBoxLayout()
        self.torgi_search_btn = QPushButton("🔎 Найти онлайн")
        self.torgi_search_btn.clicked.connect(self.run_torgi_search)
        self.torgi_excel_search_btn = QPushButton("Поиск через Excel")
        self.torgi_excel_search_btn.clicked.connect(self.run_torgi_excel_search)
        self.torgi_stop_btn = QPushButton("Остановить")
        self.torgi_stop_btn.setEnabled(False)
        self.torgi_stop_btn.clicked.connect(self.stop_torgi_search)
        self.torgi_clear_btn = QPushButton("Очистить")
        self.torgi_clear_btn.clicked.connect(self.clear_torgi_filters)
        self.torgi_open_site_btn = QPushButton("Открыть каталог")
        self.torgi_open_site_btn.clicked.connect(self.open_torgi_site)
        for button in (self.torgi_search_btn, self.torgi_excel_search_btn, self.torgi_stop_btn, self.torgi_clear_btn, self.torgi_open_site_btn):
            buttons.addWidget(button)
        buttons.addStretch()
        form.addRow("", buttons)
        layout.addWidget(group)
        self.torgi_status_label = QLabel("Найдено 0, источник torgi.gov.ru")
        layout.addWidget(self.torgi_status_label)
        self.active_filters_widget = QWidget()
        self.active_filters_layout = QHBoxLayout(self.active_filters_widget)
        self.active_filters_layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.active_filters_widget)
        self.torgi_results_table = QTableWidget()
        self.torgi_table = self.torgi_results_table
        self._configure_results_table(self.torgi_results_table, [
            "В базе", "ID / Извещение", "Название", "Категория", "Регион / адрес",
            "Начальная цена", "Статус", "Дата публикации", "Окончание заявок", "Ссылка",
        ], (2, 4))
        self.torgi_results_table.horizontalHeader().sectionClicked.connect(self.on_torgi_header_clicked)
        self.torgi_results_table.cellClicked.connect(self.open_torgi_link_cell)
        self.torgi_results_table.cellDoubleClicked.connect(self.open_torgi_result_url)
        layout.addWidget(self.torgi_results_table, 1)
        actions = QHBoxLayout()
        self.torgi_import_selected_btn = QPushButton("Импортировать выбранные в базу")
        self.torgi_import_selected_btn.clicked.connect(self.import_selected_torgi_lots)
        self.torgi_import_all_btn = QPushButton("Импортировать все найденные")
        self.torgi_import_all_btn.clicked.connect(self.import_all_torgi_lots)
        self.torgi_prev_btn = QPushButton("Предыдущая страница")
        self.torgi_prev_btn.clicked.connect(self.search_torgi_prev_page)
        self.torgi_next_btn = QPushButton("Следующая страница")
        self.torgi_next_btn.clicked.connect(self.search_torgi_next_page)
        actions.addWidget(self.torgi_import_selected_btn)
        actions.addWidget(self.torgi_import_all_btn)
        actions.addStretch()
        actions.addWidget(self.torgi_prev_btn)
        actions.addWidget(self.torgi_next_btn)
        layout.addLayout(actions)
        self.torgi_unsupported_inputs = {}
        self.restore_torgi_filter_state()
        self.update_active_filter_chips()
        self.render_torgi_results()

    def init_dash_tab(self):
        self._init_compact_torgi_tab()
        return
        layout = QVBoxLayout(self.dash_tab)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        self._last_sort_column = None
        self._last_sort_order = Qt.AscendingOrder

        top_layout = QHBoxLayout()
        title = QLabel("Поиск ГИС Торги (torgi.gov.ru)")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #143370;")
        top_layout.addWidget(title)
        top_layout.addStretch()
        refresh_btn = QPushButton("🔄")
        refresh_btn.setToolTip("Обновить статистику")
        refresh_btn.setFixedSize(42, 34)
        refresh_btn.clicked.connect(self.update_dashboard)
        top_layout.addWidget(refresh_btn)
        layout.addLayout(top_layout)

        self.stats_label = QLabel("Загрузка статистики...")
        self.stats_label.setWordWrap(True)
        self.stats_label.setMaximumHeight(118)
        self.stats_label.setStyleSheet("""
            QLabel {
                font-size: 13px;
                padding: 12px 14px;
                background: #f7f9fc;
                border: 1px solid #dfe7f3;
                border-radius: 8px;
                color: #27364f;
            }
        """)
        layout.addWidget(self.stats_label)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        layout.addWidget(splitter, 1)

        left_panel = QWidget()
        left_panel.setMinimumWidth(380)
        left_panel.setMaximumWidth(430)
        left_panel.setStyleSheet("QWidget { background: white; }")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(8, 8, 8, 8)
        left_layout.setSpacing(8)
        splitter.addWidget(left_panel)

        filter_scroll = QScrollArea()
        filter_scroll.setWidgetResizable(True)
        filter_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        filter_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        filter_scroll.setStyleSheet("QScrollArea { border: 1px solid #dfe7f3; border-radius: 8px; background: white; }")
        filter_widget = QWidget()
        filter_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        filter_layout = QVBoxLayout(filter_widget)
        filter_layout.setContentsMargins(10, 10, 10, 10)
        filter_layout.setSpacing(8)
        filter_scroll.setWidget(filter_widget)
        left_layout.addWidget(filter_scroll, 1)

        def line(placeholder: str = "") -> QLineEdit:
            widget = QLineEdit()
            widget.setPlaceholderText(placeholder)
            widget.setMinimumHeight(32)
            widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            return widget

        def combo(options: list[tuple[str, str | None]], editable: bool = False) -> QComboBox:
            widget = WheelSafeComboBox()
            widget.setEditable(editable)
            widget.setMinimumHeight(32)
            widget.setMaxVisibleItems(14)
            widget.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
            widget.setMinimumContentsLength(14)
            widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            for label, value in options:
                widget.addItem(label, value)
                if value == "__group__":
                    item = widget.model().item(widget.count() - 1)
                    if item:
                        item.setEnabled(False)
            widget.setStyleSheet("""
                QComboBox {
                    background: white;
                    border: 1px solid #cfd8e6;
                    border-radius: 5px;
                    padding: 4px 28px 4px 8px;
                    color: #1f2d3d;
                }
                QComboBox:hover { border-color: #115dee; }
                QComboBox::drop-down {
                    subcontrol-origin: padding;
                    subcontrol-position: top right;
                    width: 24px;
                    border-left: 1px solid #dfe7f3;
                }
                QComboBox QAbstractItemView {
                    background: white;
                    border: 1px solid #cfd8e6;
                    selection-background-color: #eaf1ff;
                    selection-color: #143370;
                    outline: none;
                }
            """)
            return widget

        def section(title_text: str, expanded: bool = False) -> CollapsibleSection:
            item = CollapsibleSection(title_text, expanded=expanded)
            filter_layout.addWidget(item)
            return item

        def add_labeled_field(target_layout: QVBoxLayout, label_text: str, widget: QWidget):
            label = QLabel(label_text)
            label.setWordWrap(True)
            label.setStyleSheet("font-size: 12px; color: #143370; font-weight: 600;")
            widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            target_layout.addWidget(label)
            target_layout.addWidget(widget)

        def two_field_row(left: QWidget, right: QWidget) -> QHBoxLayout:
            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)
            left.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            right.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            row.addWidget(left)
            row.addWidget(right)
            return row

        main_section = section("Основной поиск", expanded=True)
        self.torgi_search_input = line("Введите параметр поиска")
        add_labeled_field(main_section.content_layout, "Введите параметр поиска", self.torgi_search_input)
        self.torgi_load_all_checkbox = QCheckBox("Загрузить все страницы результата")
        self.torgi_load_all_checkbox.setChecked(True)
        self.torgi_load_all_checkbox.setToolTip(
            "Если включено, программа пройдет по всем страницам API и покажет все найденные лоты."
        )
        main_section.addWidget(self.torgi_load_all_checkbox)
        self.torgi_max_items_input = line("например 5000")
        self.torgi_max_items_input.setText("5000")
        add_labeled_field(main_section.content_layout, "Лимит загрузки", self.torgi_max_items_input)

        deal_section = section("Вид сделки")
        self.torgi_type_transaction_combo = combo([
            ("Не выбрано", None),
            ("Продажа", "SALE"),
            ("Аренда", "RENT"),
        ])
        add_labeled_field(deal_section.content_layout, "Вид сделки", self.torgi_type_transaction_combo)

        price_section = section("Начальная цена")
        self.torgi_price_min_input = line("от")
        self.torgi_price_max_input = line("до")
        label = QLabel("Начальная цена")
        label.setWordWrap(True)
        label.setStyleSheet("font-size: 12px; color: #143370; font-weight: 600;")
        price_section.addWidget(label)
        price_section.addLayout(two_field_row(self.torgi_price_min_input, self.torgi_price_max_input))

        location_section = section("Местонахождение")
        region_options = [("Не выбрано", None)]
        for district, region_names in TorgiGovClient.SUBJECT_RF_DISTRICTS.items():
            region_options.append((district, "__group__"))
            region_options.extend(
                (f"  {name}", TorgiGovClient.SUBJECT_RF_CODES[name])
                for name in region_names
                if name in TorgiGovClient.SUBJECT_RF_CODES
            )
        self.torgi_subject_combo = combo(region_options, editable=True)
        add_labeled_field(location_section.content_layout, "Субъект местонахождения имущества", self.torgi_subject_combo)
        self.torgi_fias_input = line("Город, район, адрес")
        add_labeled_field(location_section.content_layout, "Местонахождение имущества", self.torgi_fias_input)
        self.torgi_ownership_combo = combo([
            ("Не выбрано", None),
            ("Федеральная", "FEDERAL"),
            ("Региональная", "REGIONAL"),
            ("Муниципальная", "MUNICIPAL"),
        ])
        add_labeled_field(location_section.content_layout, "Форма собственности", self.torgi_ownership_combo)

        category_section = section("Категория имущества")
        real_estate_codes = ("2", "8", "9", "10", "11", "12", "903")
        self.torgi_category_combo = combo(
            [("Вся недвижимость", TorgiGovClient.REAL_ESTATE_CATEGORY_CODES)]
            + [(TorgiGovClient.CATEGORY_CODE_LABELS[code], code) for code in real_estate_codes]
        )
        add_labeled_field(category_section.content_layout, "Категория имущества", self.torgi_category_combo)

        lot_section = section("Параметры лота")
        self.torgi_lot_status_combo = combo([
            ("Не выбрано", None),
            ("Активные", TorgiGovClient.DEFAULT_LOT_STATUS),
            ("Прием заявок", "APPLICATIONS_SUBMISSION"),
            ("Определение участников", "DETERMINING_PARTICIPANTS"),
            ("Проведение торгов", "BIDDING"),
            ("Подведение итогов", "SUMMING_UP"),
            ("Завершен", "COMPLETED"),
            ("Несостоявшийся", "FAILED"),
        ])
        self._set_combo_data(self.torgi_lot_status_combo, TorgiGovClient.DEFAULT_LOT_STATUS)
        self.torgi_status_combo = self.torgi_lot_status_combo
        add_labeled_field(lot_section.content_layout, "Статус лота", self.torgi_lot_status_combo)
        self.torgi_currency_combo = combo([("Не выбрано", None), ("RUB", "RUB"), ("USD", "USD"), ("EUR", "EUR")])
        add_labeled_field(lot_section.content_layout, "Валюта", self.torgi_currency_combo)
        self.torgi_price_fin_from_input = line("от")
        self.torgi_price_fin_to_input = line("до")
        label = QLabel("Итоговая цена")
        label.setWordWrap(True)
        label.setStyleSheet("font-size: 12px; color: #143370; font-weight: 600;")
        lot_section.addWidget(label)
        lot_section.addLayout(two_field_row(self.torgi_price_fin_from_input, self.torgi_price_fin_to_input))
        self.torgi_is_msp_checkbox = QCheckBox("Торги среди МСП")
        lot_section.addWidget(self.torgi_is_msp_checkbox)

        notice_section = section("Извещение")
        self.torgi_notice_status_combo = combo([
            ("Не выбрано", None),
            ("Опубликованные", "PUBLISHED"),
            ("Завершенные", "COMPLETED"),
            ("Отмененные", "CANCELED"),
        ])
        add_labeled_field(notice_section.content_layout, "Статус извещения", self.torgi_notice_status_combo)
        self.torgi_notice_number_input = line("Номер извещения")
        add_labeled_field(notice_section.content_layout, "Номер извещения", self.torgi_notice_number_input)
        self.torgi_etp_input = line("Код ЭТП")
        add_labeled_field(notice_section.content_layout, "Электронная площадка", self.torgi_etp_input)
        self.torgi_publish_from_input = line("от YYYY-MM-DD")
        self.torgi_publish_to_input = line("до YYYY-MM-DD")
        label = QLabel("Дата публикации")
        label.setWordWrap(True)
        label.setStyleSheet("font-size: 12px; color: #143370; font-weight: 600;")
        notice_section.addWidget(label)
        notice_section.addLayout(two_field_row(self.torgi_publish_from_input, self.torgi_publish_to_input))
        self.torgi_bidd_end_from_input = line("от YYYY-MM-DD")
        self.torgi_bidd_end_to_input = line("до YYYY-MM-DD")
        label = QLabel("Дата окончания подачи заявок")
        label.setWordWrap(True)
        label.setStyleSheet("font-size: 12px; color: #143370; font-weight: 600;")
        notice_section.addWidget(label)
        notice_section.addLayout(two_field_row(self.torgi_bidd_end_from_input, self.torgi_bidd_end_to_input))
        self.torgi_auction_from_input = line("от YYYY-MM-DD")
        self.torgi_auction_to_input = line("до YYYY-MM-DD")
        label = QLabel("Дата проведения торгов")
        label.setWordWrap(True)
        label.setStyleSheet("font-size: 12px; color: #143370; font-weight: 600;")
        notice_section.addWidget(label)
        notice_section.addLayout(two_field_row(self.torgi_auction_from_input, self.torgi_auction_to_input))
        self.torgi_npa_input = line("Нормативный правовой акт")
        add_labeled_field(notice_section.content_layout, "Нормативный правовой акт", self.torgi_npa_input)
        self.torgi_bidd_type_combo = combo([("Не выбрано", None), ("Электронный аукцион", "ELECTRONIC_AUCTION")])
        add_labeled_field(notice_section.content_layout, "Вид торгов", self.torgi_bidd_type_combo)
        self.torgi_bidd_form_combo = combo([("Не выбрано", None), ("Открытая", "OPEN"), ("Закрытая", "CLOSED")])
        add_labeled_field(notice_section.content_layout, "Форма проведения торгов", self.torgi_bidd_form_combo)
        self.torgi_is_stopped_checkbox = QCheckBox("Наличие приостановленных торгов")
        notice_section.addWidget(self.torgi_is_stopped_checkbox)

        organizer_section = section("Организатор торгов")
        self.torgi_organizer_name_input = line("Наименование")
        self.torgi_organizer_inn_input = line("ИНН")
        self.torgi_organizer_kpp_input = line("КПП")
        self.torgi_organizer_ogrn_input = line("ОГРН/ОГРНИП")
        add_labeled_field(organizer_section.content_layout, "Наименование", self.torgi_organizer_name_input)
        add_labeled_field(organizer_section.content_layout, "ИНН", self.torgi_organizer_inn_input)
        add_labeled_field(organizer_section.content_layout, "КПП", self.torgi_organizer_kpp_input)
        add_labeled_field(organizer_section.content_layout, "ОГРН/ОГРНИП", self.torgi_organizer_ogrn_input)

        right_section = section("Правообладатель")
        self.torgi_right_holder_name_input = line("Наименование")
        self.torgi_right_holder_inn_input = line("ИНН")
        self.torgi_right_holder_kpp_input = line("КПП")
        self.torgi_right_holder_ogrn_input = line("ОГРН")
        add_labeled_field(right_section.content_layout, "Наименование", self.torgi_right_holder_name_input)
        add_labeled_field(right_section.content_layout, "ИНН", self.torgi_right_holder_inn_input)
        add_labeled_field(right_section.content_layout, "КПП", self.torgi_right_holder_kpp_input)
        add_labeled_field(right_section.content_layout, "ОГРН", self.torgi_right_holder_ogrn_input)
        self.torgi_rh_gov_prt_checkbox = QCheckBox("Правообладатель включен в перечень компаний с гос. участием")
        right_section.addWidget(self.torgi_rh_gov_prt_checkbox)

        appeals_section = section("Жалобы")
        self.torgi_has_appeals_checkbox = QCheckBox("Наличие жалоб")
        self.torgi_has_solutions_checkbox = QCheckBox("Наличие решений")
        self.torgi_has_prescriptions_checkbox = QCheckBox("Наличие предписаний")
        self.torgi_amo_org_input = line("Орган контроля")
        appeals_section.addWidget(self.torgi_has_appeals_checkbox)
        appeals_section.addWidget(self.torgi_has_solutions_checkbox)
        appeals_section.addWidget(self.torgi_has_prescriptions_checkbox)
        add_labeled_field(appeals_section.content_layout, "Орган контроля", self.torgi_amo_org_input)

        attachment_section = section("Поиск в прикрепленных файлах")
        self.torgi_attachment_input = line("Фраза для поиска")
        add_labeled_field(attachment_section.content_layout, "Фраза для поиска", self.torgi_attachment_input)
        self.torgi_match_phrase_checkbox = QCheckBox("Точное совпадение")
        attachment_section.addWidget(self.torgi_match_phrase_checkbox)
        filter_layout.addStretch()

        bottom_bar = QFrame()
        bottom_bar.setFrameShape(QFrame.NoFrame)
        bottom_layout = QVBoxLayout(bottom_bar)
        bottom_layout.setContentsMargins(0, 8, 0, 0)
        bottom_layout.setSpacing(6)
        self.torgi_search_btn = QPushButton("🔎 Найти онлайн")
        self.torgi_search_btn.setMinimumHeight(38)
        self.torgi_search_btn.setStyleSheet("""
            QPushButton {
                background: #115dee;
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: 700;
                padding: 8px;
            }
            QPushButton:hover { background: #0d4fc9; }
            QPushButton:disabled { background: #9bb6ea; }
        """)
        self.torgi_search_btn.clicked.connect(self.run_torgi_search)
        self.torgi_excel_search_btn = QPushButton("Поиск через эксель")
        self.torgi_excel_search_btn.setMinimumHeight(38)
        self.torgi_excel_search_btn.setStyleSheet("""
            QPushButton {
                background: #1f9d55;
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: 700;
                padding: 8px;
            }
            QPushButton:hover { background: #188447; }
            QPushButton:disabled { background: #9bd3b4; }
        """)
        self.torgi_excel_search_btn.clicked.connect(self.run_torgi_excel_search)
        self.torgi_stop_btn = QPushButton("Остановить поиск")
        self.torgi_stop_btn.setEnabled(False)
        self.torgi_stop_btn.clicked.connect(self.stop_torgi_search)
        self.torgi_clear_btn = QPushButton("Очистить поиск")
        self.torgi_clear_btn.clicked.connect(self.clear_torgi_filters)
        self.torgi_map_btn = QPushButton("Показать на карте")
        self.torgi_map_btn.setEnabled(False)
        bottom_layout.addWidget(self.torgi_search_btn)
        bottom_layout.addWidget(self.torgi_excel_search_btn)
        bottom_layout.addWidget(self.torgi_stop_btn)
        bottom_layout.addWidget(self.torgi_clear_btn)
        bottom_layout.addWidget(self.torgi_map_btn)
        left_layout.addWidget(bottom_bar, 0)

        self.torgi_unsupported_inputs = {
            "Итоговая цена": [self.torgi_price_fin_from_input, self.torgi_price_fin_to_input],
            "Нормативный правовой акт": [self.torgi_npa_input],
            "Приостановленные торги": [self.torgi_is_stopped_checkbox],
            "КПП/ОГРН организатора": [self.torgi_organizer_kpp_input, self.torgi_organizer_ogrn_input],
            "КПП/ОГРН правообладателя": [self.torgi_right_holder_kpp_input, self.torgi_right_holder_ogrn_input],
            "Гос. участие правообладателя": [self.torgi_rh_gov_prt_checkbox],
            "Жалобы/решения/предписания": [
                self.torgi_has_appeals_checkbox,
                self.torgi_has_solutions_checkbox,
                self.torgi_has_prescriptions_checkbox,
                self.torgi_amo_org_input,
            ],
        }

        results_panel = QWidget()
        results_layout = QVBoxLayout(results_panel)
        results_layout.setContentsMargins(14, 0, 0, 0)
        results_layout.setSpacing(10)
        splitter.addWidget(results_panel)

        header_layout = QHBoxLayout()
        self.torgi_status_label = QLabel("Найдено 0, страница 1, источник torgi.gov.ru")
        self.torgi_status_label.setStyleSheet("font-size: 13px; color: #60769f;")
        header_layout.addWidget(self.torgi_status_label)
        header_layout.addStretch()
        self.torgi_debug_btn = QPushButton("Параметры запроса")
        self.torgi_debug_btn.clicked.connect(self.show_torgi_request_diagnostics)
        header_layout.addWidget(self.torgi_debug_btn)
        self.torgi_open_site_btn = QPushButton("Открыть на torgi.gov.ru")
        self.torgi_open_site_btn.clicked.connect(self.open_torgi_site)
        header_layout.addWidget(self.torgi_open_site_btn)
        results_layout.addLayout(header_layout)

        self.active_filters_widget = QWidget()
        self.active_filters_layout = QHBoxLayout(self.active_filters_widget)
        self.active_filters_layout.setContentsMargins(0, 0, 0, 0)
        self.active_filters_layout.setSpacing(6)
        self.active_filters_layout.setAlignment(Qt.AlignLeft)
        results_layout.addWidget(self.active_filters_widget)

        self.torgi_results_table = QTableWidget()
        self.torgi_table = self.torgi_results_table
        self.torgi_results_table.setColumnCount(10)
        self.torgi_results_table.setHorizontalHeaderLabels([
            "В базе",
            "ID / Извещение",
            "Название",
            "Категория",
            "Регион / адрес",
            "Начальная цена",
            "Статус",
            "Дата публикации",
            "Окончание заявок",
            "Ссылка",
        ])
        header = self.torgi_results_table.horizontalHeader()
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.Stretch)
        header.setSectionsClickable(True)
        header.setSortIndicatorShown(True)
        header.sectionClicked.connect(self.on_torgi_header_clicked)
        self.torgi_results_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.torgi_results_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.torgi_results_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.torgi_results_table.setAlternatingRowColors(True)
        self.torgi_results_table.setSortingEnabled(True)
        self.torgi_results_table.cellClicked.connect(self.open_torgi_link_cell)
        self.torgi_results_table.cellDoubleClicked.connect(self.open_torgi_result_url)
        results_layout.addWidget(self.torgi_results_table, 1)

        actions = QHBoxLayout()
        self.torgi_import_selected_btn = QPushButton("Импортировать выбранные в базу")
        self.torgi_import_selected_btn.clicked.connect(self.import_selected_torgi_lots)
        actions.addWidget(self.torgi_import_selected_btn)
        self.torgi_import_all_btn = QPushButton("Импортировать все найденные")
        self.torgi_import_all_btn.clicked.connect(self.import_all_torgi_lots)
        actions.addWidget(self.torgi_import_all_btn)
        actions.addStretch()
        self.torgi_prev_btn = QPushButton("Предыдущая страница")
        self.torgi_prev_btn.clicked.connect(self.search_torgi_prev_page)
        self.torgi_prev_btn.setEnabled(False)
        actions.addWidget(self.torgi_prev_btn)
        self.torgi_next_btn = QPushButton("Следующая страница")
        self.torgi_next_btn.clicked.connect(self.search_torgi_next_page)
        self.torgi_next_btn.setEnabled(False)
        actions.addWidget(self.torgi_next_btn)
        results_layout.addLayout(actions)

        for signal_source in (
            self.torgi_search_input.textChanged,
            self.torgi_price_min_input.textChanged,
            self.torgi_price_max_input.textChanged,
            self.torgi_subject_combo.currentIndexChanged,
            self.torgi_category_combo.currentIndexChanged,
            self.torgi_notice_status_combo.currentIndexChanged,
            self.torgi_lot_status_combo.currentIndexChanged,
        ):
            signal_source.connect(lambda *_: self.update_active_filter_chips())

        splitter.setSizes([400, 900])
        self.restore_torgi_filter_state()
        self.update_active_filter_chips()

    def _init_compact_tbankrot_tab(self) -> None:
        layout = QVBoxLayout(self.tbankrot_tab)
        layout.setContentsMargins(8, 8, 8, 8)
        title = QLabel("Поиск TBankrot (tbankrot.ru)")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #143370;")
        layout.addWidget(title)
        group = QGroupBox("Параметры поиска недвижимости (без аренды)")
        form = QFormLayout(group)
        self.tbankrot_search_input = QLineEdit()
        self.tbankrot_search_input.setPlaceholderText("Название, адрес, номер лота...")
        form.addRow("Поиск", self.tbankrot_search_input)
        self.tbankrot_category_combo = WheelSafeComboBox()
        self.tbankrot_category_combo.addItem("Вся недвижимость", TBankrotClient.REAL_ESTATE_CATEGORY_CODES)
        for code, label in TBankrotClient.REAL_ESTATE_CATEGORY_LABELS.items():
            self.tbankrot_category_combo.addItem(label, code)
        form.addRow("Категория", self.tbankrot_category_combo)
        self.tbankrot_region_combo = WheelSafeComboBox()
        self.tbankrot_region_combo.addItem("Все регионы", None)
        for code, label in sorted(TBankrotClient.REGION_LABELS.items(), key=lambda item: item[1]):
            self.tbankrot_region_combo.addItem(label, code)
        form.addRow("Регион", self.tbankrot_region_combo)
        self.tbankrot_price_min_input, self.tbankrot_price_max_input = QLineEdit(), QLineEdit()
        self.tbankrot_price_min_input.setPlaceholderText("Цена от")
        self.tbankrot_price_max_input.setPlaceholderText("Цена до")
        price_row = QHBoxLayout()
        price_row.addWidget(self.tbankrot_price_min_input)
        price_row.addWidget(self.tbankrot_price_max_input)
        form.addRow("Цена", price_row)
        self.tbankrot_load_all_checkbox = QCheckBox("Загрузить все страницы")
        self.tbankrot_load_all_checkbox.setChecked(True)
        self.tbankrot_max_items_input = QLineEdit("5000")
        self.tbankrot_max_items_input.setMaximumWidth(110)
        mode_row = QHBoxLayout()
        mode_row.addWidget(self.tbankrot_load_all_checkbox)
        mode_row.addWidget(QLabel("Лимит:"))
        mode_row.addWidget(self.tbankrot_max_items_input)
        mode_row.addStretch()
        form.addRow("Режим", mode_row)
        buttons = QHBoxLayout()
        self.tbankrot_search_btn = QPushButton("Найти на TBankrot")
        self.tbankrot_search_btn.clicked.connect(self.run_tbankrot_search)
        self.tbankrot_stop_btn = QPushButton("Остановить")
        self.tbankrot_stop_btn.setEnabled(False)
        self.tbankrot_stop_btn.clicked.connect(self.stop_tbankrot_search)
        self.tbankrot_clear_btn = QPushButton("Очистить")
        self.tbankrot_clear_btn.clicked.connect(self.clear_tbankrot_filters)
        self.tbankrot_open_site_btn = QPushButton("Открыть каталог")
        self.tbankrot_open_site_btn.clicked.connect(self.open_tbankrot_site)
        for button in (self.tbankrot_search_btn, self.tbankrot_stop_btn, self.tbankrot_clear_btn, self.tbankrot_open_site_btn):
            buttons.addWidget(button)
        buttons.addStretch()
        form.addRow("", buttons)
        layout.addWidget(group)
        for name in (
            "tbankrot_lot_number_input", "tbankrot_debtor_input", "tbankrot_auction_manager_input",
            "tbankrot_organizer_input", "tbankrot_stop_words_input",
        ):
            setattr(self, name, QLineEdit())
        self.tbankrot_trade_type_combo = WheelSafeComboBox()
        self.tbankrot_trade_type_combo.addItem("Все типы торгов", None)
        self.tbankrot_photo_only_checkbox = QCheckBox()
        self.tbankrot_show_closed_checkbox = QCheckBox()
        self.tbankrot_show_paused_checkbox = QCheckBox()
        self.tbankrot_status_label = QLabel("Найдено 0, источник tbankrot.ru")
        layout.addWidget(self.tbankrot_status_label)
        self.tbankrot_active_filters_widget = QWidget()
        self.tbankrot_active_filters_layout = QHBoxLayout(self.tbankrot_active_filters_widget)
        self.tbankrot_active_filters_layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.tbankrot_active_filters_widget)
        self.tbankrot_results_table = QTableWidget()
        self._configure_results_table(self.tbankrot_results_table, [
            "В базе", "ID", "Название", "Категория", "Регион / адрес", "Цена",
            "Статус", "Дата публикации", "Окончание заявок", "Ссылка",
        ], (2, 4))
        self.tbankrot_results_table.horizontalHeader().sectionClicked.connect(self.on_tbankrot_header_clicked)
        self.tbankrot_results_table.cellClicked.connect(self.open_tbankrot_link_cell)
        self.tbankrot_results_table.cellDoubleClicked.connect(self.open_tbankrot_result_url)
        layout.addWidget(self.tbankrot_results_table, 1)
        actions = QHBoxLayout()
        self.tbankrot_import_selected_btn = QPushButton("Импортировать выбранные в базу")
        self.tbankrot_import_selected_btn.clicked.connect(self.import_selected_tbankrot_lots)
        self.tbankrot_import_all_btn = QPushButton("Импортировать все найденные")
        self.tbankrot_import_all_btn.clicked.connect(self.import_all_tbankrot_lots)
        self.tbankrot_prev_btn = QPushButton("Предыдущая страница")
        self.tbankrot_prev_btn.clicked.connect(self.search_tbankrot_prev_page)
        self.tbankrot_next_btn = QPushButton("Следующая страница")
        self.tbankrot_next_btn.clicked.connect(self.search_tbankrot_next_page)
        actions.addWidget(self.tbankrot_import_selected_btn)
        actions.addWidget(self.tbankrot_import_all_btn)
        actions.addStretch()
        actions.addWidget(self.tbankrot_prev_btn)
        actions.addWidget(self.tbankrot_next_btn)
        layout.addLayout(actions)
        self.restore_tbankrot_filter_state()
        self.update_tbankrot_filter_chips()
        self.render_tbankrot_results()

    def init_tbankrot_tab(self):
        self._init_compact_tbankrot_tab()
        return
        layout = QVBoxLayout(self.tbankrot_tab)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        top_layout = QHBoxLayout()
        title = QLabel("Поиск Т Банкрот (tbankrot.ru)")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #143370;")
        top_layout.addWidget(title)
        top_layout.addStretch()
        layout.addLayout(top_layout)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        layout.addWidget(splitter, 1)

        left_panel = QWidget()
        left_panel.setMinimumWidth(380)
        left_panel.setMaximumWidth(430)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(8, 8, 8, 8)
        left_layout.setSpacing(8)
        splitter.addWidget(left_panel)

        filter_scroll = QScrollArea()
        filter_scroll.setWidgetResizable(True)
        filter_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        filter_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        filter_scroll.setStyleSheet("QScrollArea { border: 1px solid #dfe7f3; border-radius: 8px; background: white; }")
        filter_widget = QWidget()
        filter_layout = QVBoxLayout(filter_widget)
        filter_layout.setContentsMargins(10, 10, 10, 10)
        filter_layout.setSpacing(8)
        filter_scroll.setWidget(filter_widget)
        left_layout.addWidget(filter_scroll, 1)

        def line(placeholder: str = "") -> QLineEdit:
            widget = QLineEdit()
            widget.setPlaceholderText(placeholder)
            widget.setMinimumHeight(32)
            widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            return widget

        def combo(options: list[tuple[str, str | None]]) -> QComboBox:
            widget = WheelSafeComboBox()
            widget.setMinimumHeight(32)
            widget.setMaxVisibleItems(16)
            widget.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
            widget.setMinimumContentsLength(16)
            widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            for label, value in options:
                widget.addItem(label, value)
            return widget

        def section(title_text: str, expanded: bool = False) -> CollapsibleSection:
            item = CollapsibleSection(title_text, expanded=expanded)
            filter_layout.addWidget(item)
            return item

        def add_labeled_field(target_layout: QVBoxLayout, label_text: str, widget: QWidget):
            label = QLabel(label_text)
            label.setWordWrap(True)
            label.setStyleSheet("font-size: 12px; color: #143370; font-weight: 600;")
            widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            target_layout.addWidget(label)
            target_layout.addWidget(widget)

        def two_field_row(left: QWidget, right: QWidget) -> QHBoxLayout:
            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)
            row.addWidget(left)
            row.addWidget(right)
            return row

        main_section = section("Основной поиск", expanded=True)
        self.tbankrot_search_input = line("Поиск")
        add_labeled_field(main_section.content_layout, "Поиск", self.tbankrot_search_input)
        self.tbankrot_load_all_checkbox = QCheckBox("Загрузить все страницы результата")
        self.tbankrot_load_all_checkbox.setChecked(True)
        main_section.content_layout.addWidget(self.tbankrot_load_all_checkbox)
        self.tbankrot_max_items_input = line("5000")
        self.tbankrot_max_items_input.setText("5000")
        add_labeled_field(main_section.content_layout, "Лимит лотов при загрузке всех страниц", self.tbankrot_max_items_input)

        price_section = section("Цена и регион", expanded=True)
        self.tbankrot_price_min_input = line("Цена от")
        self.tbankrot_price_max_input = line("Цена до")
        price_section.content_layout.addLayout(two_field_row(self.tbankrot_price_min_input, self.tbankrot_price_max_input))
        region_options = [("Не выбрано", None)]
        region_options.extend((label, code) for code, label in sorted(TBankrotClient.REGION_LABELS.items(), key=lambda item: item[1]))
        self.tbankrot_region_combo = combo(region_options)
        add_labeled_field(price_section.content_layout, "Регион", self.tbankrot_region_combo)

        trade_section = section("Торги", expanded=True)
        self.tbankrot_trade_type_combo = combo([
            ("Не выбрано", None),
            ("Аукцион", "auction"),
            ("Публичное предложение", "public"),
        ])
        add_labeled_field(trade_section.content_layout, "Тип торгов", self.tbankrot_trade_type_combo)
        self.tbankrot_lot_number_input = line("Номер лота")
        add_labeled_field(trade_section.content_layout, "Номер лота", self.tbankrot_lot_number_input)
        self.tbankrot_photo_only_checkbox = QCheckBox("Только с фото")
        trade_section.content_layout.addWidget(self.tbankrot_photo_only_checkbox)
        self.tbankrot_show_closed_checkbox = QCheckBox("Показывать завершенные")
        trade_section.content_layout.addWidget(self.tbankrot_show_closed_checkbox)
        self.tbankrot_show_paused_checkbox = QCheckBox("Показывать приостановленные")
        trade_section.content_layout.addWidget(self.tbankrot_show_paused_checkbox)

        parties_section = section("Участники и слова")
        self.tbankrot_debtor_input = line("Должник")
        add_labeled_field(parties_section.content_layout, "Должник", self.tbankrot_debtor_input)
        self.tbankrot_auction_manager_input = line("Арбитражный управляющий")
        add_labeled_field(parties_section.content_layout, "Арбитражный управляющий", self.tbankrot_auction_manager_input)
        self.tbankrot_organizer_input = line("Организатор")
        add_labeled_field(parties_section.content_layout, "Организатор", self.tbankrot_organizer_input)
        self.tbankrot_stop_words_input = line("Стоп-слова")
        add_labeled_field(parties_section.content_layout, "Стоп-слова", self.tbankrot_stop_words_input)
        filter_layout.addStretch()

        bottom_bar = QFrame()
        bottom_layout = QVBoxLayout(bottom_bar)
        bottom_layout.setContentsMargins(0, 8, 0, 0)
        bottom_layout.setSpacing(6)
        self.tbankrot_search_btn = QPushButton("Найти на TBankrot")
        self.tbankrot_search_btn.setMinimumHeight(38)
        self.tbankrot_search_btn.setStyleSheet("""
            QPushButton {
                background: #115dee;
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: 700;
                padding: 8px;
            }
            QPushButton:hover { background: #0d4fc9; }
            QPushButton:disabled { background: #9bb6ea; }
        """)
        self.tbankrot_search_btn.clicked.connect(self.run_tbankrot_search)
        self.tbankrot_stop_btn = QPushButton("Остановить поиск")
        self.tbankrot_stop_btn.setEnabled(False)
        self.tbankrot_stop_btn.clicked.connect(self.stop_tbankrot_search)
        self.tbankrot_clear_btn = QPushButton("Очистить поиск")
        self.tbankrot_clear_btn.clicked.connect(self.clear_tbankrot_filters)
        bottom_layout.addWidget(self.tbankrot_search_btn)
        bottom_layout.addWidget(self.tbankrot_stop_btn)
        bottom_layout.addWidget(self.tbankrot_clear_btn)
        left_layout.addWidget(bottom_bar, 0)

        results_panel = QWidget()
        results_layout = QVBoxLayout(results_panel)
        results_layout.setContentsMargins(14, 0, 0, 0)
        results_layout.setSpacing(10)
        splitter.addWidget(results_panel)

        header_layout = QHBoxLayout()
        self.tbankrot_status_label = QLabel("Найдено 0, страница 1, источник tbankrot.ru")
        self.tbankrot_status_label.setStyleSheet("font-size: 13px; color: #60769f;")
        header_layout.addWidget(self.tbankrot_status_label)
        header_layout.addStretch()
        self.tbankrot_debug_btn = QPushButton("Параметры запроса")
        self.tbankrot_debug_btn.clicked.connect(self.show_tbankrot_request_diagnostics)
        header_layout.addWidget(self.tbankrot_debug_btn)
        self.tbankrot_open_site_btn = QPushButton("Открыть на tbankrot.ru")
        self.tbankrot_open_site_btn.clicked.connect(self.open_tbankrot_site)
        header_layout.addWidget(self.tbankrot_open_site_btn)
        results_layout.addLayout(header_layout)

        self.tbankrot_active_filters_widget = QWidget()
        self.tbankrot_active_filters_layout = QHBoxLayout(self.tbankrot_active_filters_widget)
        self.tbankrot_active_filters_layout.setContentsMargins(0, 0, 0, 0)
        self.tbankrot_active_filters_layout.setSpacing(6)
        self.tbankrot_active_filters_layout.setAlignment(Qt.AlignLeft)
        results_layout.addWidget(self.tbankrot_active_filters_widget)

        self.tbankrot_results_table = QTableWidget()
        self.tbankrot_results_table.setColumnCount(10)
        self.tbankrot_results_table.setHorizontalHeaderLabels([
            "В базе",
            "ID",
            "Название",
            "Категория",
            "Регион / адрес",
            "Цена",
            "Статус",
            "Дата публикации",
            "Окончание заявок",
            "Ссылка",
        ])
        header = self.tbankrot_results_table.horizontalHeader()
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.Stretch)
        header.setSectionsClickable(True)
        header.setSortIndicatorShown(True)
        header.sectionClicked.connect(self.on_tbankrot_header_clicked)
        self.tbankrot_results_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbankrot_results_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.tbankrot_results_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbankrot_results_table.setAlternatingRowColors(True)
        self.tbankrot_results_table.setSortingEnabled(True)
        self.tbankrot_results_table.cellClicked.connect(self.open_tbankrot_link_cell)
        self.tbankrot_results_table.cellDoubleClicked.connect(self.open_tbankrot_result_url)
        results_layout.addWidget(self.tbankrot_results_table, 1)

        actions = QHBoxLayout()
        self.tbankrot_import_selected_btn = QPushButton("Импортировать выбранные в базу")
        self.tbankrot_import_selected_btn.clicked.connect(self.import_selected_tbankrot_lots)
        actions.addWidget(self.tbankrot_import_selected_btn)
        self.tbankrot_import_all_btn = QPushButton("Импортировать все найденные")
        self.tbankrot_import_all_btn.clicked.connect(self.import_all_tbankrot_lots)
        actions.addWidget(self.tbankrot_import_all_btn)
        actions.addStretch()
        self.tbankrot_prev_btn = QPushButton("Предыдущая страница")
        self.tbankrot_prev_btn.clicked.connect(self.search_tbankrot_prev_page)
        self.tbankrot_prev_btn.setEnabled(False)
        actions.addWidget(self.tbankrot_prev_btn)
        self.tbankrot_next_btn = QPushButton("Следующая страница")
        self.tbankrot_next_btn.clicked.connect(self.search_tbankrot_next_page)
        self.tbankrot_next_btn.setEnabled(False)
        actions.addWidget(self.tbankrot_next_btn)
        results_layout.addLayout(actions)

        for signal_source in (
            self.tbankrot_search_input.textChanged,
            self.tbankrot_price_min_input.textChanged,
            self.tbankrot_price_max_input.textChanged,
            self.tbankrot_region_combo.currentIndexChanged,
            self.tbankrot_trade_type_combo.currentIndexChanged,
            self.tbankrot_lot_number_input.textChanged,
        ):
            signal_source.connect(lambda *_: self.update_tbankrot_filter_chips())

        splitter.setSizes([400, 900])
        self.restore_tbankrot_filter_state()
        self.update_tbankrot_filter_chips()
        self.render_tbankrot_results()

    def _combo_value(self, widget: QComboBox) -> str | None:
        data = widget.currentData()
        return str(data) if data not in (None, "") else None

    def _line_text(self, widget: QLineEdit) -> str | None:
        value = widget.text().strip()
        return value or None

    def _line_float(self, widget: QLineEdit) -> float | None:
        value = widget.text().strip().replace(" ", "").replace(",", ".")
        if not value:
            return None
        try:
            return float(value)
        except ValueError:
            raise ValueError(f"Некорректное числовое значение: {widget.text()}")

    def _line_int_or_none(self, widget: QLineEdit) -> int | None:
        value = widget.text().strip().replace(" ", "")
        if not value:
            return None
        try:
            parsed = int(value)
            return parsed if parsed > 0 else None
        except ValueError:
            raise ValueError(f"Некорректное целое значение: {widget.text()}")

    def _set_combo_data(self, widget: QComboBox, value: str | None) -> None:
        if value in (None, ""):
            index = widget.findData(None)
        else:
            index = widget.findData(value)
        if index >= 0:
            widget.setCurrentIndex(index)
        elif widget.isEditable() and value:
            widget.setEditText(str(value))

    def _collect_unsupported_torgi_warnings(self) -> list[str]:
        warnings = []
        for label, widgets in getattr(self, "torgi_unsupported_inputs", {}).items():
            used = False
            for widget in widgets:
                if isinstance(widget, QLineEdit) and widget.text().strip():
                    used = True
                elif isinstance(widget, QCheckBox) and widget.isChecked():
                    used = True
            if used:
                warnings.append(f"Фильтр '{label}' пока отображается в GUI, но не передается в онлайн API.")
        return warnings

    def collect_torgi_filters(self, page: int | None = None) -> TorgiGovSearchFilters:
        self.torgi_unsupported_warnings = []
        return TorgiGovSearchFilters(
            search_text=self.torgi_search_input.text().strip(),
            type_transaction="SALE",
            price_min=self._line_float(self.torgi_price_min_input),
            price_max=self._line_float(self.torgi_price_max_input),
            subject_rf=self._combo_value(self.torgi_subject_combo),
            category_code=(
                self._combo_value(self.torgi_category_combo)
                or TorgiGovClient.REAL_ESTATE_CATEGORY_CODES
            ),
            lot_status=self._combo_value(self.torgi_lot_status_combo),
            page=page or self.torgi_current_page,
            page_size=100,
        )

    def save_torgi_filter_state(self) -> None:
        from bankrotai.core import set_app_setting
        state = {
            "search_text": self.torgi_search_input.text().strip(),
            "subject_rf": self._combo_value(self.torgi_subject_combo),
            "price_min": self.torgi_price_min_input.text().strip(),
            "price_max": self.torgi_price_max_input.text().strip(),
            "category_code": self._combo_value(self.torgi_category_combo),
            "lot_status": self._combo_value(self.torgi_lot_status_combo),
        }
        set_app_setting("torgi_gov_last_filters", json.dumps(state, ensure_ascii=False))

    def restore_torgi_filter_state(self) -> None:
        from bankrotai.core import get_app_setting
        raw = get_app_setting("torgi_gov_last_filters", "")
        if not raw:
            return
        try:
            state = json.loads(raw)
        except json.JSONDecodeError:
            return
        self.torgi_search_input.setText(state.get("search_text", ""))
        self._set_combo_data(self.torgi_subject_combo, state.get("subject_rf"))
        self.torgi_price_min_input.setText(state.get("price_min", ""))
        self.torgi_price_max_input.setText(state.get("price_max", ""))
        self._set_combo_data(self.torgi_category_combo, state.get("category_code"))
        self._set_combo_data(self.torgi_lot_status_combo, state.get("lot_status") or TorgiGovClient.DEFAULT_LOT_STATUS)

    def update_active_filter_chips(self):
        if not hasattr(self, "active_filters_layout"):
            return
        while self.active_filters_layout.count():
            item = self.active_filters_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

        chips = self.get_active_filter_chips()
        if not chips:
            label = QLabel("Фильтры не выбраны")
            label.setStyleSheet("color: #7a8699; font-size: 12px;")
            self.active_filters_layout.addWidget(label)
            return

        for label_text, clear_callback in chips:
            chip = QPushButton(f"{label_text}  ×")
            chip.setCursor(Qt.PointingHandCursor)
            chip.setStyleSheet("""
                QPushButton {
                    background: #eaf1ff;
                    color: #143370;
                    border: 1px solid #c7d8ff;
                    border-radius: 12px;
                    padding: 4px 10px;
                    font-size: 12px;
                }
                QPushButton:hover { background: #dce9ff; }
            """)
            chip.clicked.connect(clear_callback)
            self.active_filters_layout.addWidget(chip)

        self.active_filters_layout.addStretch(1)

    def get_active_filter_chips(self) -> list[tuple[str, Callable[[], None]]]:
        chips: list[tuple[str, Callable[[], None]]] = []

        text = self.torgi_search_input.text().strip()
        if text:
            chips.append((f"Поиск: {text}", lambda: self.clear_filter_widget(self.torgi_search_input)))

        region_label = self.torgi_subject_combo.currentText()
        region_value = self.torgi_subject_combo.currentData()
        if region_value:
            chips.append((region_label, lambda: self.clear_filter_combo(self.torgi_subject_combo)))

        category_label = self.torgi_category_combo.currentText()
        category_value = self.torgi_category_combo.currentData()
        if category_value:
            chips.append((category_label, lambda: self.clear_filter_combo(self.torgi_category_combo)))

        lot_label = self.torgi_lot_status_combo.currentText()
        lot_value = self.torgi_lot_status_combo.currentData()
        if lot_value:
            chips.append((lot_label, lambda: self.clear_filter_combo(self.torgi_lot_status_combo)))

        price_min = self.torgi_price_min_input.text().strip()
        if price_min:
            chips.append((f"Цена от {price_min}", lambda: self.clear_filter_widget(self.torgi_price_min_input)))

        price_max = self.torgi_price_max_input.text().strip()
        if price_max:
            chips.append((f"Цена до {price_max}", lambda: self.clear_filter_widget(self.torgi_price_max_input)))

        return chips

    def clear_filter_widget(self, widget):
        widget.clear()
        self.update_active_filter_chips()
        self.save_torgi_filter_state()

    def clear_filter_combo(self, combo):
        combo.setCurrentIndex(0)
        self.update_active_filter_chips()
        self.save_torgi_filter_state()

    def clear_torgi_filters(self):
        for widget in (self.torgi_search_input, self.torgi_price_min_input, self.torgi_price_max_input):
            widget.clear()
        self._set_combo_data(self.torgi_subject_combo, None)
        self._set_combo_data(self.torgi_category_combo, TorgiGovClient.REAL_ESTATE_CATEGORY_CODES)
        self._set_combo_data(self.torgi_status_combo, TorgiGovClient.DEFAULT_LOT_STATUS)
        self.torgi_load_all_checkbox.setChecked(True)
        self.torgi_max_items_input.setText("5000")
        self.torgi_results = []
        self.torgi_current_page = 1
        self.render_torgi_results()
        self.update_active_filter_chips()
        self.save_torgi_filter_state()

    def run_torgi_search(self, page: int = 1):
        try:
            self.torgi_current_page = max(1, page)
            filters = self.collect_torgi_filters(self.torgi_current_page)
            load_all = self.torgi_load_all_checkbox.isChecked()
            max_items = self._line_int_or_none(self.torgi_max_items_input)
            if load_all:
                filters.page = 1
                filters.page_size = 100
                self.torgi_current_page = 1
            self.save_torgi_filter_state()
        except ValueError as exc:
            QMessageBox.warning(self, "Проверьте фильтры", str(exc))
            return

        self.torgi_search_btn.setEnabled(False)
        self.torgi_search_btn.setText("Идёт поиск...")
        self.torgi_excel_search_btn.setEnabled(False)
        self.torgi_stop_btn.setEnabled(True)
        self.torgi_results = []
        self.torgi_meta = {"mode": "all_pages" if load_all else "page", "loaded": 0}
        self.render_torgi_results()
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.status_bar.showMessage("Онлайн-поиск torgi.gov.ru...")

        self.torgi_worker = TorgiGovSearchWorker(
            filters,
            load_all=load_all,
            max_items=max_items,
        )
        self.torgi_worker.progress.connect(self.status_bar.showMessage)
        self.torgi_worker.progress_percent.connect(self.progress_bar.setValue)
        self.torgi_worker.page_loaded.connect(self.on_torgi_search_page_loaded)
        self.torgi_worker.finished.connect(self.on_torgi_search_finished)
        self.torgi_worker.error.connect(self.on_torgi_search_error)
        self.torgi_worker.start()

    def run_torgi_excel_search(self):
        try:
            self.torgi_current_page = 1
            filters = self.collect_torgi_filters(1)
            filters.page = 1
            filters.page_size = 100
            self.save_torgi_filter_state()
        except ValueError as exc:
            QMessageBox.warning(self, "Проверьте фильтры", str(exc))
            return

        self.torgi_search_btn.setEnabled(False)
        self.torgi_excel_search_btn.setEnabled(False)
        self.torgi_excel_search_btn.setText("Загрузка Excel...")
        self.torgi_stop_btn.setEnabled(False)
        self.torgi_results = []
        self.torgi_meta = {"mode": "excel", "loaded": 0}
        self.render_torgi_results()
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.status_bar.showMessage("Поиск через Excel-выгрузку torgi.gov.ru...")

        self.torgi_worker = TorgiGovSearchWorker(
            filters,
            load_all=False,
            max_items=None,
            use_excel=True,
        )
        self.torgi_worker.progress.connect(self.status_bar.showMessage)
        self.torgi_worker.progress_percent.connect(self.progress_bar.setValue)
        self.torgi_worker.page_loaded.connect(self.on_torgi_search_page_loaded)
        self.torgi_worker.finished.connect(self.on_torgi_search_finished)
        self.torgi_worker.error.connect(self.on_torgi_search_error)
        self.torgi_worker.start()

    def stop_torgi_search(self):
        worker = getattr(self, "torgi_worker", None)
        if worker and worker.isRunning():
            worker.request_stop()
            self.torgi_stop_btn.setEnabled(False)
            self.status_bar.showMessage("Останавливаю поиск после текущего запроса...", 5000)

    def on_torgi_search_page_loaded(self, lots: list, page_meta: dict):
        seen = {lot.external_id for lot in self.torgi_results}
        for lot in lots or []:
            if lot.external_id not in seen:
                self.torgi_results.append(lot)
                seen.add(lot.external_id)
        meta = dict(self.torgi_meta or {})
        meta.update(page_meta or {})
        meta["loaded"] = len(self.torgi_results)
        meta.setdefault("mode", "all_pages" if self.torgi_load_all_checkbox.isChecked() else "page")
        self.torgi_meta = meta
        self.render_torgi_results()

    def on_torgi_search_finished(self, lots: list, meta: dict):
        self.torgi_search_btn.setEnabled(True)
        self.torgi_search_btn.setText("🔎 Найти онлайн")
        self.torgi_excel_search_btn.setEnabled(True)
        self.torgi_excel_search_btn.setText("Поиск через эксель")
        self.torgi_stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        meta = dict(meta or {})
        warnings = list(meta.get("warnings") or [])
        warnings.extend(getattr(self, "torgi_unsupported_warnings", []))
        meta["warnings"] = warnings
        self.torgi_results = list(lots or [])
        self.torgi_meta = meta
        self.render_torgi_results()
        if warnings:
            self.status_bar.showMessage(warnings[0], 10000)
        elif meta.get("stop_reason") == "user_stopped":
            self.status_bar.showMessage(f"Поиск остановлен: сохранено {len(self.torgi_results)} лотов", 5000)
        elif meta.get("mode") == "excel":
            self.status_bar.showMessage(f"Excel-поиск завершен: {len(self.torgi_results)} лотов", 5000)
        else:
            self.status_bar.showMessage(f"Онлайн-поиск завершен: {len(self.torgi_results)} лотов", 5000)

    def on_torgi_search_error(self, error_msg: str):
        self.torgi_search_btn.setEnabled(True)
        self.torgi_search_btn.setText("🔎 Найти онлайн")
        self.torgi_excel_search_btn.setEnabled(True)
        self.torgi_excel_search_btn.setText("Поиск через эксель")
        self.torgi_stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.status_bar.showMessage("Ошибка онлайн-поиска torgi.gov.ru", 10000)
        QMessageBox.warning(self, "Ошибка torgi.gov.ru", error_msg)

    def render_torgi_results(self):
        meta = self.torgi_meta or {}
        loaded = meta.get("loaded", len(self.torgi_results))
        total = meta.get("total")
        page = meta.get("page", self.torgi_current_page)
        source = meta.get("source", "torgi.gov.ru")
        mode = meta.get("mode")
        diagnostics = []
        if meta.get("pages_loaded") is not None:
            diagnostics.append(f"страниц: {meta.get('pages_loaded')}")
        if meta.get("duplicates"):
            diagnostics.append(f"дублей: {meta.get('duplicates')}")
        if meta.get("skipped_without_id"):
            diagnostics.append(f"без ID: {meta.get('skipped_without_id')}")
        if meta.get("stop_reason"):
            diagnostics.append(f"стоп: {meta.get('stop_reason')}")
        if total is not None:
            total_text = f"{loaded} из {total}"
        else:
            total_text = str(loaded)
        if mode == "all_pages":
            mode_text = "все страницы"
        elif mode == "excel":
            mode_text = "Excel-выгрузка"
        else:
            mode_text = f"страница {page}"
        warnings = meta.get("warnings") or []
        suffix_parts = []
        if diagnostics:
            suffix_parts.append("; ".join(diagnostics))
        if warnings:
            suffix_parts.append(f"предупреждений: {len(warnings)}")
        suffix = f" · {' · '.join(suffix_parts)}" if suffix_parts else ""
        self.torgi_status_label.setText(
            f"Найдено {total_text}, режим: {mode_text}, источник {source}{suffix}"
        )
        self.torgi_prev_btn.setEnabled(self.torgi_current_page > 1 and mode != "all_pages")
        self.torgi_next_btn.setEnabled(bool(meta.get("has_more")) and mode != "all_pages")

        existing: set[str] = set()
        external_ids = [lot.external_id for lot in self.torgi_results]
        if external_ids:
            with session_scope() as session:
                existing = set(session.scalars(
                    select(ProcessedLot.external_id).where(
                        ProcessedLot.source_system == "torgi_gov",
                        ProcessedLot.external_id.in_(external_ids),
                    )
                ).all())

        self.torgi_results_table.setSortingEnabled(False)
        self.torgi_results_table.setRowCount(len(self.torgi_results))
        for row, lot in enumerate(self.torgi_results):
            in_db = lot.external_id in existing
            raw = lot.raw_data or {}
            bidd_end = raw.get("bidd_end_time") or raw.get("biddEndTime") or ""
            published = raw.get("published_at") or raw.get("publicationDate") or raw.get("firstVersionPublicationDate") or lot.published_at
            link_url = lot.source_url or lot.lot_url
            items = [
                make_text_item("Да" if in_db else ""),
                make_text_item(lot.external_id.replace("torgi_gov:", "")),
                make_text_item(lot.title),
                make_text_item(translate_category(lot.category)),
                make_text_item(lot.region_name or lot.address or lot.region_slug or ""),
                make_number_item(lot.start_price or lot.current_price),
                make_text_item(translate_status(lot.auction_status)),
                make_date_item(published),
                make_date_item(bidd_end),
                make_text_item("Открыть" if link_url else ""),
            ]
            row_color = QColor("#eaf7ef") if in_db else QColor("white")
            for col, item in enumerate(items):
                item.setData(Qt.UserRole, lot.external_id)
                item.setData(EXTERNAL_ID_ROLE, lot.external_id)
                item.setData(URL_ROLE, link_url)
                item.setBackground(QBrush(row_color))
                self.torgi_results_table.setItem(row, col, item)
        self.torgi_results_table.setSortingEnabled(True)

    def _format_money(self, value: float | None) -> str:
        if value is None:
            return ""
        try:
            return f"{int(float(value)):,} ₽".replace(",", " ")
        except (TypeError, ValueError):
            return str(value)

    def _torgi_lot_by_external_id(self, external_id: str | None) -> NormalizedLot | None:
        if not external_id:
            return None
        for lot in self.torgi_results:
            if lot.external_id == external_id:
                return lot
        return None

    def selected_torgi_lots(self) -> list[NormalizedLot]:
        rows = sorted({item.row() for item in self.torgi_results_table.selectedItems()})
        lots = []
        for row in rows:
            item = self.torgi_results_table.item(row, 1)
            external_id = item.data(EXTERNAL_ID_ROLE) if item else None
            lot = self._torgi_lot_by_external_id(external_id or (item.data(Qt.UserRole) if item else None))
            if lot:
                lots.append(lot)
        return lots

    def import_selected_torgi_lots(self):
        self.import_torgi_lots(self.selected_torgi_lots())

    def import_all_torgi_lots(self):
        total = (self.torgi_meta or {}).get("total")
        loaded = len(self.torgi_results)
        mode = (self.torgi_meta or {}).get("mode")
        if total and loaded < int(total) and mode != "all_pages":
            reply = QMessageBox.question(
                self,
                "Загружена не вся выдача",
                f"Сейчас загружено только {loaded} из {total} лотов.\n\n"
                "Чтобы импортировать все найденные лоты, включите галочку "
                "«Загрузить все страницы результата» и выполните поиск заново.\n\n"
                "Импортировать только загруженные сейчас?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return
        self.import_torgi_lots(list(self.torgi_results))

    def import_torgi_lots(self, lots: list[NormalizedLot]):
        if not lots:
            QMessageBox.information(self, "Импорт", "Нет выбранных онлайн-лотов для импорта.")
            return
        external_ids = [lot.external_id for lot in lots]
        with session_scope() as session:
            existed = set(session.scalars(
                select(ProcessedLot.external_id).where(
                    ProcessedLot.source_system == "torgi_gov",
                    ProcessedLot.external_id.in_(external_ids),
                )
            ).all())
            for lot in lots:
                persist_lot(session, lot)
        added = len([lot for lot in lots if lot.external_id not in existed])
        updated = len(lots) - added
        self.status_bar.showMessage(f"Импорт torgi.gov.ru: +{added}, обновлено {updated}", 5000)
        QMessageBox.information(self, "Импорт завершен", f"Добавлено: {added}\nОбновлено: {updated}")
        self.load_lots()
        self.update_dashboard()
        self.render_torgi_results()

    def search_torgi_next_page(self):
        self.run_torgi_search(self.torgi_current_page + 1)

    def search_torgi_prev_page(self):
        self.run_torgi_search(max(1, self.torgi_current_page - 1))

    def on_torgi_header_clicked(self, column: int):
        numeric_desc_first_cols = {5}

        if self._last_sort_column == column:
            order = Qt.DescendingOrder if self._last_sort_order == Qt.AscendingOrder else Qt.AscendingOrder
        else:
            order = Qt.DescendingOrder if column in numeric_desc_first_cols else Qt.AscendingOrder

        self._last_sort_column = column
        self._last_sort_order = order
        self.torgi_results_table.sortItems(column, order)

    def open_torgi_result_url(self, row: int, column: int) -> None:
        item = self.torgi_results_table.item(row, column) or self.torgi_results_table.item(row, 1)
        external_id = item.data(EXTERNAL_ID_ROLE) if item else None
        lot = self._torgi_lot_by_external_id(external_id or (item.data(Qt.UserRole) if item else None))
        url = (item.data(URL_ROLE) if item else None) or (lot.source_url if lot else None) or (lot.lot_url if lot else None)
        if url:
            QDesktopServices.openUrl(QUrl.fromUserInput(str(url)))
        else:
            QMessageBox.information(self, "Инфо", "Ссылка на лот отсутствует.")

    def open_torgi_link_cell(self, row: int, column: int) -> None:
        if column == 9:
            self.open_torgi_result_url(row, column)

    def show_torgi_request_diagnostics(self):
        meta = self.torgi_meta or {}
        if not meta:
            try:
                filters = self.collect_torgi_filters(self.torgi_current_page)
                params, warnings = TorgiGovClient()._build_query_params(filters)
                endpoint = TorgiGovClient()._prepare_url(TorgiGovClient.SEARCH_ENDPOINT, params)
                meta = {"raw_endpoint": endpoint, "raw_params": params, "warnings": warnings}
            except Exception as exc:
                QMessageBox.warning(self, "Диагностика запроса", str(exc))
                return

        lines = [
            f"Endpoint: {meta.get('raw_endpoint') or TorgiGovClient.SEARCH_ENDPOINT}",
            f"Params: {json.dumps(meta.get('raw_params') or {}, ensure_ascii=False, indent=2)}",
            f"total: {meta.get('total')}",
            f"total_pages: {meta.get('total_pages')}",
            f"loaded: {meta.get('loaded', len(self.torgi_results))}",
            f"raw_items_loaded: {meta.get('raw_items_loaded')}",
            f"pages_loaded: {meta.get('pages_loaded')}",
            f"duplicates: {meta.get('duplicates')}",
            f"skipped_without_id: {meta.get('skipped_without_id')}",
            f"stop_reason: {meta.get('stop_reason')}",
        ]
        page_diagnostics = meta.get("page_diagnostics") or []
        if page_diagnostics:
            lines.append("Pages:")
            for item in page_diagnostics[:20]:
                lines.append(
                    f"page {item.get('page')}: items={item.get('items_on_page')}, "
                    f"new={item.get('new_unique')}, duplicates={item.get('duplicates')}, "
                    f"without_id={item.get('skipped_without_id')}"
                )
            if len(page_diagnostics) > 20:
                lines.append(f"... ещё страниц: {len(page_diagnostics) - 20}")
        warnings = meta.get("warnings") or []
        if warnings:
            lines.append("Warnings:")
            lines.extend(f"- {warning}" for warning in warnings)

        QMessageBox.information(self, "Параметры запроса torgi.gov.ru", "\n".join(lines))

    def open_torgi_site(self):
        try:
            filters = self.collect_torgi_filters(self.torgi_current_page)
            params, _warnings = TorgiGovClient()._build_query_params(filters)
            for key in ("withFacets", "size"):
                params.pop(key, None)
            url = f"{TorgiGovClient.FALLBACK_LIST_URL}?{urlencode(params)}"
        except Exception:
            url = TorgiGovClient.FALLBACK_LIST_URL
        import webbrowser
        webbrowser.open(url)

    def collect_tbankrot_filters(self, page: int | None = None) -> TBankrotSearchFilters:
        return TBankrotSearchFilters(
            search_text=self.tbankrot_search_input.text().strip(),
            region=self._combo_value(self.tbankrot_region_combo),
            price_min=self._line_float(self.tbankrot_price_min_input),
            price_max=self._line_float(self.tbankrot_price_max_input),
            lot_number=self._line_text(self.tbankrot_lot_number_input),
            trade_type=self._combo_value(self.tbankrot_trade_type_combo),
            category_codes=(
                self._combo_value(self.tbankrot_category_combo)
                or TBankrotClient.REAL_ESTATE_CATEGORY_CODES
            ),
            photo_only=self.tbankrot_photo_only_checkbox.isChecked(),
            debtor=self._line_text(self.tbankrot_debtor_input),
            auction_manager=self._line_text(self.tbankrot_auction_manager_input),
            organizer=self._line_text(self.tbankrot_organizer_input),
            stop_words=self._line_text(self.tbankrot_stop_words_input),
            show_closed=self.tbankrot_show_closed_checkbox.isChecked(),
            show_paused=self.tbankrot_show_paused_checkbox.isChecked(),
            page=page or self.tbankrot_current_page,
            page_size=100,
        )

    def save_tbankrot_filter_state(self) -> None:
        from bankrotai.core import set_app_setting
        state = {
            "search_text": self.tbankrot_search_input.text().strip(),
            "region": self._combo_value(self.tbankrot_region_combo),
            "price_min": self.tbankrot_price_min_input.text().strip(),
            "price_max": self.tbankrot_price_max_input.text().strip(),
            "trade_type": self._combo_value(self.tbankrot_trade_type_combo),
            "category_codes": self._combo_value(self.tbankrot_category_combo),
        }
        set_app_setting("tbankrot_last_filters", json.dumps(state, ensure_ascii=False))

    def restore_tbankrot_filter_state(self) -> None:
        from bankrotai.core import get_app_setting
        raw = get_app_setting("tbankrot_last_filters", "")
        if not raw:
            return
        try:
            state = json.loads(raw)
        except json.JSONDecodeError:
            return
        self.tbankrot_search_input.setText(state.get("search_text", ""))
        self._set_combo_data(self.tbankrot_region_combo, state.get("region"))
        self.tbankrot_price_min_input.setText(state.get("price_min", ""))
        self.tbankrot_price_max_input.setText(state.get("price_max", ""))
        self._set_combo_data(self.tbankrot_trade_type_combo, state.get("trade_type"))
        self._set_combo_data(
            self.tbankrot_category_combo,
            state.get("category_codes") or TBankrotClient.REAL_ESTATE_CATEGORY_CODES,
        )

    def update_tbankrot_filter_chips(self):
        if not hasattr(self, "tbankrot_active_filters_layout"):
            return
        while self.tbankrot_active_filters_layout.count():
            item = self.tbankrot_active_filters_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

        chips = self.get_tbankrot_filter_chips()
        if not chips:
            label = QLabel("Фильтры не выбраны")
            label.setStyleSheet("color: #7a8699; font-size: 12px;")
            self.tbankrot_active_filters_layout.addWidget(label)
            return

        for label_text, clear_callback in chips:
            chip = QPushButton(f"{label_text}  ×")
            chip.setCursor(Qt.PointingHandCursor)
            chip.setStyleSheet("""
                QPushButton {
                    background: #eaf1ff;
                    color: #143370;
                    border: 1px solid #c7d8ff;
                    border-radius: 12px;
                    padding: 4px 10px;
                    font-size: 12px;
                }
                QPushButton:hover { background: #dce9ff; }
            """)
            chip.clicked.connect(clear_callback)
            self.tbankrot_active_filters_layout.addWidget(chip)
        self.tbankrot_active_filters_layout.addStretch(1)

    def get_tbankrot_filter_chips(self) -> list[tuple[str, Callable[[], None]]]:
        chips: list[tuple[str, Callable[[], None]]] = []

        text = self.tbankrot_search_input.text().strip()
        if text:
            chips.append((f"Поиск: {text}", lambda: self.clear_tbankrot_filter_widget(self.tbankrot_search_input)))

        region_value = self.tbankrot_region_combo.currentData()
        if region_value:
            chips.append((self.tbankrot_region_combo.currentText(), lambda: self.clear_tbankrot_filter_combo(self.tbankrot_region_combo)))

        category_value = self.tbankrot_category_combo.currentData()
        if category_value:
            chips.append((self.tbankrot_category_combo.currentText(), lambda: self._set_tbankrot_all_categories()))

        trade_value = self.tbankrot_trade_type_combo.currentData()
        if trade_value:
            chips.append((self.tbankrot_trade_type_combo.currentText(), lambda: self.clear_tbankrot_filter_combo(self.tbankrot_trade_type_combo)))

        price_min = self.tbankrot_price_min_input.text().strip()
        if price_min:
            chips.append((f"Цена от {price_min}", lambda: self.clear_tbankrot_filter_widget(self.tbankrot_price_min_input)))

        price_max = self.tbankrot_price_max_input.text().strip()
        if price_max:
            chips.append((f"Цена до {price_max}", lambda: self.clear_tbankrot_filter_widget(self.tbankrot_price_max_input)))

        lot_number = self.tbankrot_lot_number_input.text().strip()
        if lot_number:
            chips.append((f"Лот: {lot_number}", lambda: self.clear_tbankrot_filter_widget(self.tbankrot_lot_number_input)))

        return chips

    def clear_tbankrot_filter_widget(self, widget):
        widget.clear()
        self.update_tbankrot_filter_chips()
        self.save_tbankrot_filter_state()

    def clear_tbankrot_filter_combo(self, combo):
        combo.setCurrentIndex(0)
        self.update_tbankrot_filter_chips()
        self.save_tbankrot_filter_state()

    def _set_tbankrot_all_categories(self):
        self._set_combo_data(self.tbankrot_category_combo, TBankrotClient.REAL_ESTATE_CATEGORY_CODES)
        self.update_tbankrot_filter_chips()
        self.save_tbankrot_filter_state()

    def clear_tbankrot_filters(self):
        for widget in [
            self.tbankrot_search_input, self.tbankrot_price_min_input, self.tbankrot_price_max_input,
            self.tbankrot_lot_number_input, self.tbankrot_debtor_input,
            self.tbankrot_auction_manager_input, self.tbankrot_organizer_input,
            self.tbankrot_stop_words_input,
        ]:
            widget.clear()
        self._set_combo_data(self.tbankrot_region_combo, None)
        self._set_combo_data(self.tbankrot_trade_type_combo, None)
        self._set_combo_data(self.tbankrot_category_combo, TBankrotClient.REAL_ESTATE_CATEGORY_CODES)
        for widget in [
            self.tbankrot_photo_only_checkbox, self.tbankrot_show_closed_checkbox,
            self.tbankrot_show_paused_checkbox,
        ]:
            widget.setChecked(False)
        self.tbankrot_load_all_checkbox.setChecked(True)
        self.tbankrot_max_items_input.setText("5000")
        self.tbankrot_results = []
        self.tbankrot_current_page = 1
        self.render_tbankrot_results()
        self.update_tbankrot_filter_chips()
        self.save_tbankrot_filter_state()

    def run_tbankrot_search(self, page: int = 1):
        try:
            self.tbankrot_current_page = max(1, page)
            filters = self.collect_tbankrot_filters(self.tbankrot_current_page)
            load_all = self.tbankrot_load_all_checkbox.isChecked()
            max_items = self._line_int_or_none(self.tbankrot_max_items_input)
            if load_all:
                filters.page = 1
                self.tbankrot_current_page = 1
            self.save_tbankrot_filter_state()
        except ValueError as exc:
            QMessageBox.warning(self, "Проверьте фильтры", str(exc))
            return

        self.tbankrot_search_btn.setEnabled(False)
        self.tbankrot_search_btn.setText("Идет поиск...")
        self.tbankrot_stop_btn.setEnabled(True)
        self.tbankrot_results = []
        self.tbankrot_meta = {"mode": "all_pages" if load_all else "page", "loaded": 0}
        self.render_tbankrot_results()
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.status_bar.showMessage("Поиск TBankrot...")

        self.tbankrot_worker = TBankrotSearchWorker(
            filters,
            load_all=load_all,
            max_items=max_items,
        )
        self.tbankrot_worker.progress.connect(self.status_bar.showMessage)
        self.tbankrot_worker.progress_percent.connect(self.progress_bar.setValue)
        self.tbankrot_worker.page_loaded.connect(self.on_tbankrot_search_page_loaded)
        self.tbankrot_worker.finished.connect(self.on_tbankrot_search_finished)
        self.tbankrot_worker.error.connect(self.on_tbankrot_search_error)
        self.tbankrot_worker.start()

    def stop_tbankrot_search(self):
        worker = getattr(self, "tbankrot_worker", None)
        if worker and worker.isRunning():
            worker.request_stop()
            self.tbankrot_stop_btn.setEnabled(False)
            self.status_bar.showMessage("Останавливаю поиск TBankrot после текущего запроса...", 5000)

    def on_tbankrot_search_page_loaded(self, lots: list, page_meta: dict):
        seen = {lot.external_id for lot in self.tbankrot_results}
        for lot in lots or []:
            if lot.external_id not in seen:
                self.tbankrot_results.append(lot)
                seen.add(lot.external_id)
        meta = dict(self.tbankrot_meta or {})
        meta.update(page_meta or {})
        meta["loaded"] = len(self.tbankrot_results)
        meta.setdefault("mode", "all_pages" if self.tbankrot_load_all_checkbox.isChecked() else "page")
        self.tbankrot_meta = meta
        self.render_tbankrot_results()

    def on_tbankrot_search_finished(self, lots: list, meta: dict):
        self.tbankrot_search_btn.setEnabled(True)
        self.tbankrot_search_btn.setText("Найти на TBankrot")
        self.tbankrot_stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.tbankrot_results = list(lots or [])
        self.tbankrot_meta = dict(meta or {})
        self.render_tbankrot_results()
        if self.tbankrot_meta.get("stop_reason") == "user_stopped":
            self.status_bar.showMessage(f"Поиск TBankrot остановлен: сохранено {len(self.tbankrot_results)} лотов", 5000)
        else:
            self.status_bar.showMessage(f"Поиск TBankrot завершен: {len(self.tbankrot_results)} лотов", 5000)

    def on_tbankrot_search_error(self, error_msg: str):
        self.tbankrot_search_btn.setEnabled(True)
        self.tbankrot_search_btn.setText("Найти на TBankrot")
        self.tbankrot_stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.status_bar.showMessage("Ошибка поиска TBankrot", 10000)
        QMessageBox.warning(self, "Ошибка TBankrot", error_msg)

    def render_tbankrot_results(self):
        meta = self.tbankrot_meta or {}
        loaded = meta.get("loaded", len(self.tbankrot_results))
        page = meta.get("page", self.tbankrot_current_page)
        source = meta.get("source", "tbankrot.ru")
        mode = meta.get("mode")
        diagnostics = []
        if meta.get("pages_loaded") is not None:
            diagnostics.append(f"страниц: {meta.get('pages_loaded')}")
        if meta.get("duplicates"):
            diagnostics.append(f"дублей: {meta.get('duplicates')}")
        if meta.get("stop_reason"):
            diagnostics.append(f"стоп: {meta.get('stop_reason')}")
        mode_text = "все страницы" if mode == "all_pages" else f"страница {page}"
        suffix = f" · {'; '.join(diagnostics)}" if diagnostics else ""
        self.tbankrot_status_label.setText(
            f"Найдено {loaded}, режим: {mode_text}, источник {source}{suffix}"
        )
        self.tbankrot_prev_btn.setEnabled(self.tbankrot_current_page > 1 and mode != "all_pages")
        self.tbankrot_next_btn.setEnabled(bool(meta.get("has_more")) and mode != "all_pages")

        existing: set[str] = set()
        external_ids = [lot.external_id for lot in self.tbankrot_results]
        if external_ids:
            with session_scope() as session:
                existing = set(session.scalars(
                    select(ProcessedLot.external_id).where(
                        ProcessedLot.source_system == "tbankrot",
                        ProcessedLot.external_id.in_(external_ids),
                    )
                ).all())

        self.tbankrot_results_table.setSortingEnabled(False)
        self.tbankrot_results_table.setRowCount(len(self.tbankrot_results))
        for row, lot in enumerate(self.tbankrot_results):
            in_db = lot.external_id in existing
            raw = lot.raw_data or {}
            dates = raw.get("dates") or []
            bidd_end = ""
            if dates:
                bidd_end = " / ".join(str(item.get("text") or "") for item in dates[:2] if item.get("text"))
            link_url = lot.source_url or lot.lot_url
            items = [
                make_text_item("Да" if in_db else ""),
                make_text_item(lot.external_id.replace("tbankrot:", "")),
                make_text_item(lot.title),
                make_text_item(translate_category(lot.category)),
                make_text_item(lot.region_name or lot.address or lot.region_slug or ""),
                make_number_item(lot.current_price or lot.start_price),
                make_text_item(translate_status(lot.auction_status)),
                make_date_item(lot.published_at),
                make_text_item(bidd_end),
                make_text_item("Открыть" if link_url else ""),
            ]
            row_color = QColor("#eaf7ef") if in_db else QColor("white")
            for col, item in enumerate(items):
                item.setData(Qt.UserRole, lot.external_id)
                item.setData(EXTERNAL_ID_ROLE, lot.external_id)
                item.setData(URL_ROLE, link_url)
                item.setBackground(QBrush(row_color))
                self.tbankrot_results_table.setItem(row, col, item)
        self.tbankrot_results_table.setSortingEnabled(True)

    def _tbankrot_lot_by_external_id(self, external_id: str | None) -> NormalizedLot | None:
        if not external_id:
            return None
        for lot in self.tbankrot_results:
            if lot.external_id == external_id:
                return lot
        return None

    def selected_tbankrot_lots(self) -> list[NormalizedLot]:
        rows = sorted({item.row() for item in self.tbankrot_results_table.selectedItems()})
        lots = []
        for row in rows:
            item = self.tbankrot_results_table.item(row, 1)
            external_id = item.data(EXTERNAL_ID_ROLE) if item else None
            lot = self._tbankrot_lot_by_external_id(external_id or (item.data(Qt.UserRole) if item else None))
            if lot:
                lots.append(lot)
        return lots

    def import_selected_tbankrot_lots(self):
        self.import_tbankrot_lots(self.selected_tbankrot_lots())

    def import_all_tbankrot_lots(self):
        self.import_tbankrot_lots(list(self.tbankrot_results))

    def import_tbankrot_lots(self, lots: list[NormalizedLot]):
        if not lots:
            QMessageBox.information(self, "Импорт", "Нет выбранных лотов TBankrot для импорта.")
            return
        external_ids = [lot.external_id for lot in lots]
        with session_scope() as session:
            existed = set(session.scalars(
                select(ProcessedLot.external_id).where(
                    ProcessedLot.source_system == "tbankrot",
                    ProcessedLot.external_id.in_(external_ids),
                )
            ).all())
            for lot in lots:
                persist_lot(session, lot)
        added = len([lot for lot in lots if lot.external_id not in existed])
        updated = len(lots) - added
        self.status_bar.showMessage(f"Импорт TBankrot: +{added}, обновлено {updated}", 5000)
        QMessageBox.information(self, "Импорт завершен", f"Добавлено: {added}\nОбновлено: {updated}")
        self.load_lots()
        self.update_dashboard()
        self.render_tbankrot_results()

    def search_tbankrot_next_page(self):
        self.run_tbankrot_search(self.tbankrot_current_page + 1)

    def search_tbankrot_prev_page(self):
        self.run_tbankrot_search(max(1, self.tbankrot_current_page - 1))

    def on_tbankrot_header_clicked(self, column: int):
        numeric_desc_first_cols = {5}
        if self._last_tbankrot_sort_column == column:
            order = Qt.DescendingOrder if self._last_tbankrot_sort_order == Qt.AscendingOrder else Qt.AscendingOrder
        else:
            order = Qt.DescendingOrder if column in numeric_desc_first_cols else Qt.AscendingOrder
        self._last_tbankrot_sort_column = column
        self._last_tbankrot_sort_order = order
        self.tbankrot_results_table.sortItems(column, order)

    def open_tbankrot_result_url(self, row: int, column: int) -> None:
        item = self.tbankrot_results_table.item(row, column) or self.tbankrot_results_table.item(row, 1)
        external_id = item.data(EXTERNAL_ID_ROLE) if item else None
        lot = self._tbankrot_lot_by_external_id(external_id or (item.data(Qt.UserRole) if item else None))
        url = (item.data(URL_ROLE) if item else None) or (lot.source_url if lot else None) or (lot.lot_url if lot else None)
        if url:
            QDesktopServices.openUrl(QUrl.fromUserInput(str(url)))
        else:
            QMessageBox.information(self, "Инфо", "Ссылка на лот отсутствует.")

    def open_tbankrot_link_cell(self, row: int, column: int) -> None:
        if column == 9:
            self.open_tbankrot_result_url(row, column)

    def show_tbankrot_request_diagnostics(self):
        meta = self.tbankrot_meta or {}
        if not meta:
            try:
                filters = self.collect_tbankrot_filters(self.tbankrot_current_page)
                params = TBankrotClient()._build_query_params(filters)
                endpoint = TBankrotClient()._prepare_url(params)
                meta = {"raw_endpoint": endpoint, "raw_params": params}
            except Exception as exc:
                QMessageBox.warning(self, "Диагностика запроса", str(exc))
                return

        lines = [
            f"Endpoint: {meta.get('raw_endpoint') or TBankrotClient.SEARCH_ENDPOINT}",
            f"Params: {json.dumps(meta.get('raw_params') or [], ensure_ascii=False, indent=2)}",
            f"loaded: {meta.get('loaded', len(self.tbankrot_results))}",
            f"pages_loaded: {meta.get('pages_loaded')}",
            f"duplicates: {meta.get('duplicates')}",
            f"stop_reason: {meta.get('stop_reason')}",
        ]
        page_diagnostics = meta.get("page_diagnostics") or []
        if page_diagnostics:
            lines.append("Pages:")
            for item in page_diagnostics[:20]:
                lines.append(
                    f"page {item.get('page')}: items={item.get('items_on_page')}, "
                    f"new={item.get('new_unique')}, duplicates={item.get('duplicates')}"
                )
            if len(page_diagnostics) > 20:
                lines.append(f"... еще страниц: {len(page_diagnostics) - 20}")

        QMessageBox.information(self, "Параметры запроса TBankrot", "\n".join(lines))

    def open_tbankrot_site(self):
        try:
            filters = self.collect_tbankrot_filters(self.tbankrot_current_page)
            params = TBankrotClient()._build_query_params(filters)
            url = TBankrotClient()._prepare_url(params)
        except Exception:
            url = TBankrotClient.BASE_URL
        import webbrowser
        webbrowser.open(url)

    def init_lot_online_tab(self):
        layout = QVBoxLayout(self.lot_online_tab)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        title = QLabel("Поиск РАД / ЛОТ-ОНЛАЙН (catalog.lot-online.ru)")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #143370;")
        layout.addWidget(title)

        filters = QGroupBox("Параметры поиска")
        form = QFormLayout(filters)
        self.lot_online_search_input = QLineEdit()
        self.lot_online_search_input.setPlaceholderText("Название, адрес, номер лота...")
        form.addRow("Поиск", self.lot_online_search_input)

        self.lot_online_category_combo = WheelSafeComboBox()
        self.lot_online_category_combo.addItem("Недвижимое имущество", "1")
        form.addRow("Каталог", self.lot_online_category_combo)

        self.lot_online_region_combo = WheelSafeComboBox()
        self.lot_online_region_combo.addItem("Все регионы", None)
        for code, label in sorted(LotOnlineClient.REGION_FEATURES.items(), key=lambda item: item[1]):
            self.lot_online_region_combo.addItem(label, code)
        self._set_combo_data(self.lot_online_region_combo, "24392")
        form.addRow("Регион", self.lot_online_region_combo)

        self.lot_online_archive_combo = WheelSafeComboBox()
        self.lot_online_archive_combo.addItem("Активные", "false")
        self.lot_online_archive_combo.addItem("Все", "all")
        self.lot_online_archive_combo.addItem("Архивные", "true")
        form.addRow("Состояние", self.lot_online_archive_combo)

        mode_row = QHBoxLayout()
        self.lot_online_load_all_checkbox = QCheckBox("Загрузить все страницы")
        self.lot_online_load_all_checkbox.setChecked(True)
        self.lot_online_max_items_input = QLineEdit("5000")
        self.lot_online_max_items_input.setMaximumWidth(110)
        self.lot_online_max_items_input.setPlaceholderText("Лимит")
        mode_row.addWidget(self.lot_online_load_all_checkbox)
        mode_row.addWidget(QLabel("Лимит:"))
        mode_row.addWidget(self.lot_online_max_items_input)
        mode_row.addStretch()
        form.addRow("Режим", mode_row)

        buttons = QHBoxLayout()
        self.lot_online_search_btn = QPushButton("Найти на ЛОТ-ОНЛАЙН")
        self.lot_online_search_btn.setMinimumHeight(38)
        self.lot_online_search_btn.setStyleSheet(
            "QPushButton { background: #115dee; color: white; border: none; border-radius: 6px; "
            "font-weight: 700; padding: 8px; } QPushButton:disabled { background: #9bb6ea; }"
        )
        self.lot_online_search_btn.clicked.connect(lambda: self.run_lot_online_search())
        self.lot_online_stop_btn = QPushButton("Остановить")
        self.lot_online_stop_btn.setEnabled(False)
        self.lot_online_stop_btn.clicked.connect(self.stop_lot_online_search)
        self.lot_online_open_site_btn = QPushButton("Открыть каталог")
        self.lot_online_open_site_btn.clicked.connect(self.open_lot_online_site)
        buttons.addWidget(self.lot_online_search_btn)
        buttons.addWidget(self.lot_online_stop_btn)
        buttons.addWidget(self.lot_online_open_site_btn)
        buttons.addStretch()
        form.addRow("", buttons)
        layout.addWidget(filters)

        self.lot_online_status_label = QLabel("Найдено 0, источник lot-online.ru")
        self.lot_online_status_label.setStyleSheet("font-size: 13px; color: #60769f;")
        layout.addWidget(self.lot_online_status_label)

        self.lot_online_results_table = QTableWidget()
        self.lot_online_results_table.setColumnCount(9)
        self.lot_online_results_table.setHorizontalHeaderLabels([
            "В базе", "ID", "Номер РАД", "Название", "Категория",
            "Регион / адрес", "Цена", "Статус", "Ссылка",
        ])
        header = self.lot_online_results_table.horizontalHeader()
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(5, QHeaderView.Stretch)
        self.lot_online_results_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.lot_online_results_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.lot_online_results_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.lot_online_results_table.setAlternatingRowColors(True)
        self.lot_online_results_table.setSortingEnabled(True)
        self.lot_online_results_table.cellClicked.connect(self.open_lot_online_link_cell)
        self.lot_online_results_table.cellDoubleClicked.connect(self.open_lot_online_result_url)
        layout.addWidget(self.lot_online_results_table, 1)

        actions = QHBoxLayout()
        import_selected = QPushButton("Импортировать выбранные в базу")
        import_selected.clicked.connect(self.import_selected_lot_online_lots)
        import_all = QPushButton("Импортировать все найденные")
        import_all.clicked.connect(self.import_all_lot_online_lots)
        self.lot_online_prev_btn = QPushButton("Предыдущая страница")
        self.lot_online_prev_btn.clicked.connect(self.search_lot_online_prev_page)
        self.lot_online_next_btn = QPushButton("Следующая страница")
        self.lot_online_next_btn.clicked.connect(self.search_lot_online_next_page)
        actions.addWidget(import_selected)
        actions.addWidget(import_all)
        actions.addStretch()
        actions.addWidget(self.lot_online_prev_btn)
        actions.addWidget(self.lot_online_next_btn)
        layout.addLayout(actions)
        self.render_lot_online_results()

    def collect_lot_online_filters(self, page: int | None = None) -> LotOnlineSearchFilters:
        return LotOnlineSearchFilters(
            search_text=self.lot_online_search_input.text().strip(),
            category_id=str(self.lot_online_category_combo.currentData() or "1"),
            region_feature=self.lot_online_region_combo.currentData(),
            archive_mode=str(self.lot_online_archive_combo.currentData() or "false"),
            page=max(1, page or self.lot_online_current_page),
            page_size=96,
        )

    def run_lot_online_search(self, page: int = 1):
        self.lot_online_current_page = max(1, page)
        search_all = self.lot_online_load_all_checkbox.isChecked()
        filters = self.collect_lot_online_filters(1 if search_all else self.lot_online_current_page)
        try:
            max_items = self._line_int_or_none(self.lot_online_max_items_input)
        except ValueError as exc:
            QMessageBox.warning(self, "Проверьте фильтры", str(exc))
            return
        self.lot_online_results = []
        self.lot_online_meta = {"mode": "all_pages" if search_all else "page", "loaded": 0}
        self.render_lot_online_results()
        self.lot_online_search_btn.setEnabled(False)
        self.lot_online_search_btn.setText("Идет поиск...")
        self.lot_online_stop_btn.setEnabled(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)

        self.lot_online_worker = LotOnlineSearchWorker(filters, load_all=search_all, max_items=max_items)
        self.lot_online_worker.progress.connect(self.status_bar.showMessage)
        self.lot_online_worker.progress_percent.connect(self.progress_bar.setValue)
        self.lot_online_worker.page_loaded.connect(self.on_lot_online_page_loaded)
        self.lot_online_worker.finished.connect(self.on_lot_online_search_finished)
        self.lot_online_worker.error.connect(self.on_lot_online_search_error)
        self.lot_online_worker.start()

    def stop_lot_online_search(self):
        worker = getattr(self, "lot_online_worker", None)
        if worker and worker.isRunning():
            worker.request_stop()
            self.lot_online_stop_btn.setEnabled(False)
            self.status_bar.showMessage("Останавливаю поиск ЛОТ-ОНЛАЙН после текущего запроса...", 5000)

    def on_lot_online_page_loaded(self, lots: list, page_meta: dict):
        seen = {lot.external_id for lot in self.lot_online_results}
        for lot in lots or []:
            if lot.external_id not in seen:
                self.lot_online_results.append(lot)
                seen.add(lot.external_id)
        self.lot_online_meta.update(page_meta or {})
        self.lot_online_meta["loaded"] = len(self.lot_online_results)
        self.render_lot_online_results()

    def on_lot_online_search_finished(self, lots: list, meta: dict):
        self.lot_online_search_btn.setEnabled(True)
        self.lot_online_search_btn.setText("Найти на ЛОТ-ОНЛАЙН")
        self.lot_online_stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.lot_online_results = list(lots or [])
        self.lot_online_meta = dict(meta or {})
        self.render_lot_online_results()
        self.status_bar.showMessage(
            f"Поиск ЛОТ-ОНЛАЙН завершен: {len(self.lot_online_results)} лотов", 5000
        )

    def on_lot_online_search_error(self, error_msg: str):
        self.lot_online_search_btn.setEnabled(True)
        self.lot_online_search_btn.setText("Найти на ЛОТ-ОНЛАЙН")
        self.lot_online_stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.status_bar.showMessage("Ошибка поиска ЛОТ-ОНЛАЙН", 10000)
        QMessageBox.warning(self, "Ошибка ЛОТ-ОНЛАЙН", error_msg)

    def render_lot_online_results(self):
        meta = self.lot_online_meta or {}
        mode = meta.get("mode")
        mode_text = "все страницы" if mode == "all_pages" else f"страница {self.lot_online_current_page}"
        self.lot_online_status_label.setText(
            f"Найдено {len(self.lot_online_results)}, режим: {mode_text}, источник lot-online.ru"
        )
        self.lot_online_prev_btn.setEnabled(self.lot_online_current_page > 1 and mode != "all_pages")
        self.lot_online_next_btn.setEnabled(bool(meta.get("has_more")) and mode != "all_pages")
        external_ids = [lot.external_id for lot in self.lot_online_results]
        existing: set[str] = set()
        if external_ids:
            with session_scope() as session:
                existing = set(session.scalars(select(ProcessedLot.external_id).where(
                    ProcessedLot.source_system == "lot-online.ru",
                    ProcessedLot.external_id.in_(external_ids),
                )).all())

        self.lot_online_results_table.setSortingEnabled(False)
        self.lot_online_results_table.setRowCount(len(self.lot_online_results))
        for row, lot in enumerate(self.lot_online_results):
            raw = lot.raw_data or {}
            link_url = lot.source_url or lot.lot_url
            items = [
                make_text_item("Да" if lot.external_id in existing else ""),
                make_text_item(lot.external_id.replace("lot-online:", "")),
                make_text_item(lot.procedure_number or raw.get("procedure_number") or ""),
                make_text_item(lot.title),
                make_text_item(translate_category(lot.category)),
                make_text_item(lot.region_name or lot.address or lot.region_slug or ""),
                make_number_item(lot.current_price or lot.start_price),
                make_text_item(translate_status(lot.auction_status)),
                make_text_item("Открыть" if link_url else ""),
            ]
            for item in items:
                item.setData(EXTERNAL_ID_ROLE, lot.external_id)
                item.setData(URL_ROLE, link_url)
            for column, item in enumerate(items):
                self.lot_online_results_table.setItem(row, column, item)
        self.lot_online_results_table.setSortingEnabled(True)

    def _lot_online_lot_by_external_id(self, external_id: str | None) -> NormalizedLot | None:
        return next((lot for lot in self.lot_online_results if lot.external_id == external_id), None)

    def selected_lot_online_lots(self) -> list[NormalizedLot]:
        rows = sorted({item.row() for item in self.lot_online_results_table.selectedItems()})
        selected = []
        for row in rows:
            item = self.lot_online_results_table.item(row, 1)
            lot = self._lot_online_lot_by_external_id(item.data(EXTERNAL_ID_ROLE) if item else None)
            if lot:
                selected.append(lot)
        return selected

    def import_selected_lot_online_lots(self):
        self.import_lot_online_lots(self.selected_lot_online_lots())

    def import_all_lot_online_lots(self):
        self.import_lot_online_lots(list(self.lot_online_results))

    def import_lot_online_lots(self, lots: list[NormalizedLot]):
        if not lots:
            QMessageBox.information(self, "Импорт", "Нет лотов ЛОТ-ОНЛАЙН для импорта.")
            return
        external_ids = [lot.external_id for lot in lots]
        with session_scope() as session:
            existed = set(session.scalars(select(ProcessedLot.external_id).where(
                ProcessedLot.source_system == "lot-online.ru",
                ProcessedLot.external_id.in_(external_ids),
            )).all())
            for lot in lots:
                persist_lot(session, lot)
        added = sum(lot.external_id not in existed for lot in lots)
        self.status_bar.showMessage(f"Импорт ЛОТ-ОНЛАЙН: +{added}, обновлено {len(lots) - added}", 5000)
        self.load_lots()
        self.update_dashboard()
        self.render_lot_online_results()

    def search_lot_online_next_page(self):
        self.run_lot_online_search(self.lot_online_current_page + 1)

    def search_lot_online_prev_page(self):
        self.run_lot_online_search(max(1, self.lot_online_current_page - 1))

    def open_lot_online_result_url(self, row: int, _column: int):
        item = self.lot_online_results_table.item(row, 1)
        lot = self._lot_online_lot_by_external_id(item.data(EXTERNAL_ID_ROLE) if item else None)
        url = lot.source_url or lot.lot_url if lot else None
        if url:
            QDesktopServices.openUrl(QUrl.fromUserInput(str(url)))

    def open_lot_online_link_cell(self, row: int, column: int):
        if column == 8:
            self.open_lot_online_result_url(row, column)

    def open_lot_online_site(self):
        try:
            params = LotOnlineClient()._build_query_params(self.collect_lot_online_filters())
            url = LotOnlineClient()._prepare_url(params)
        except Exception:
            url = LotOnlineClient.SEARCH_ENDPOINT
        QDesktopServices.openUrl(QUrl.fromUserInput(url))

    def init_registry_tab(self):
        layout = QVBoxLayout(self.registry_tab)
        
        # Toolbar for Registry
        toolbar = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Поиск по названию...")
        self.search_input.setMaximumHeight(35)
        self.search_input.textChanged.connect(lambda: self.load_lots())
        
        self.category_combo = QComboBox()
        self.category_combo.addItems(["Все категории", "Жилая недвижимость", "Коммерческая недвижимость", "Земельные участки", "Транспорт", "Прочее"])
        self.category_combo.setFixedWidth(200)
        self.category_combo.currentIndexChanged.connect(lambda: self.load_lots())
        
        refresh_reg_btn = QPushButton("🔄")
        refresh_reg_btn.setFixedWidth(40)
        refresh_reg_btn.clicked.connect(lambda: self.load_lots())
        
        toolbar.addWidget(self.search_input)
        toolbar.addWidget(self.category_combo)
        toolbar.addWidget(refresh_reg_btn)
        layout.addLayout(toolbar)

        # Sorting Toolbar
        sort_toolbar = QHBoxLayout()
        sort_toolbar.addWidget(QLabel("Сортировка:"))
        
        self.sort_new_btn = QPushButton("Сначала новые")
        self.sort_new_btn.setCheckable(True)
        self.sort_new_btn.setChecked(True)
        self.sort_new_btn.clicked.connect(lambda: self.set_sort("last_update", "desc"))
        
        self.sort_cheap_btn = QPushButton("Дешевле")
        self.sort_cheap_btn.setCheckable(True)
        self.sort_cheap_btn.clicked.connect(lambda: self.set_sort("current_price", "asc"))
        
        self.sort_expensive_btn = QPushButton("Дороже")
        self.sort_expensive_btn.setCheckable(True)
        self.sort_expensive_btn.clicked.connect(lambda: self.set_sort("current_price", "desc"))
        
        self.sort_group = [self.sort_new_btn, self.sort_cheap_btn, self.sort_expensive_btn]
        for btn in self.sort_group: sort_toolbar.addWidget(btn)
        sort_toolbar.addStretch()
        # Sorting is handled by clicking table columns; the old button toolbar is kept out of the UI.

        self.current_sort = ("last_update", "desc")

        self.lots_splitter = QSplitter(Qt.Horizontal)
        layout.addWidget(self.lots_splitter)

        # Table
        self.lots_table = QTableWidget()
        self.lots_table.setColumnCount(12)
        self.lots_table.setHorizontalHeaderLabels([
            "Код", 
            "Название лота", 
            "Регион", 
            "Цена", 
            "Рыночная цена", 
            "Дисконт", 
            "Площадь здания/помещения", 
            "Площадь участка", 
            "Статус", 
            "Категория", 
            "Риск", 
            "Обновление"
        ])
        self.lots_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.lots_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.lots_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.lots_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.lots_table.setSortingEnabled(True) # --- Enable sorting ---
        # Enable double-click to open lot URL
        self.lots_table.cellDoubleClicked.connect(self.open_lot_url)
        self.lots_table.itemSelectionChanged.connect(self.on_lot_selected)
        self.lots_table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        self.lots_splitter.addWidget(self.lots_table)

        # Details
        self.detail_panel = QWidget()
        detail_layout = QVBoxLayout(self.detail_panel)
        
        detail_layout.addWidget(QLabel("<b>📄 Детали лота:</b>"))
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        self.detail_text.setStyleSheet("background-color: #fdfefe; border: 1px solid #d5dbdb;")
        detail_layout.addWidget(self.detail_text)
        
        detail_layout.addWidget(QLabel("<b>🤖 Заключение AI:</b>"))
        self.detail_ai = QTextEdit()
        self.detail_ai.setReadOnly(True)
        self.detail_ai.setMaximumHeight(180)
        self.detail_ai.setStyleSheet("background-color: #f4faff; border: 1px solid #d6eaf8; color: #21618c;")
        detail_layout.addWidget(self.detail_ai)
        
        # AI & Geo Actions
        actions_layout = QHBoxLayout()
        
        self.ai_single_btn = QPushButton("🤖 Оценить AI")
        self.ai_single_btn.setFixedHeight(35)
        self.ai_single_btn.setStyleSheet("background-color: #e8f6f3; color: #16a085; font-weight: bold;")
        self.ai_single_btn.clicked.connect(self.run_ai_single)
        self.ai_single_btn.setEnabled(False)
        actions_layout.addWidget(self.ai_single_btn)
        
        self.geo_fix_btn = QPushButton("🗺️ Гео")
        self.geo_fix_btn.setFixedHeight(35)
        self.geo_fix_btn.setStyleSheet("background-color: #fef9e7; color: #d4ac0d; font-weight: bold;")
        self.geo_fix_btn.clicked.connect(self.refresh_geo)
        self.geo_fix_btn.setEnabled(False)
        actions_layout.addWidget(self.geo_fix_btn)
        
        detail_layout.addLayout(actions_layout)

        # Review Status Buttons
        review_layout = QHBoxLayout()
        self.review_approved_btn = QPushButton("✅")
        self.review_approved_btn.setToolTip("Принять")
        self.review_approved_btn.clicked.connect(lambda: self.change_review_status("approved"))
        
        self.review_maybe_btn = QPushButton("❓")
        self.review_maybe_btn.setToolTip("Под вопросом")
        self.review_maybe_btn.clicked.connect(lambda: self.change_review_status("maybe"))
        
        self.review_rejected_btn = QPushButton("❌")
        self.review_rejected_btn.setToolTip("Отклонить")
        self.review_rejected_btn.clicked.connect(lambda: self.change_review_status("rejected"))
        
        for btn in [self.review_approved_btn, self.review_maybe_btn, self.review_rejected_btn]:
            btn.setFixedHeight(35)
            btn.setEnabled(False)
            review_layout.addWidget(btn)
        
        detail_layout.addLayout(review_layout)

        self.max_bid_btn = QPushButton("🧮 Калькулятор максимальной ставки")
        self.max_bid_btn.setFixedHeight(40)
        self.max_bid_btn.setStyleSheet(
            "background-color: #eaf2f8; color: #2471a3; font-weight: bold; border-radius: 5px;"
        )
        self.max_bid_btn.clicked.connect(self.open_max_bid_calculator)
        self.max_bid_btn.setEnabled(False)
        detail_layout.addWidget(self.max_bid_btn)
        
        self.delete_lot_btn = QPushButton("🗑️ Удалить выбранные")
        self.delete_lot_btn.setFixedHeight(40)
        self.delete_lot_btn.setStyleSheet("background-color: #fadbd8; color: #c0392b; font-weight: bold; border-radius: 5px;")
        self.delete_lot_btn.clicked.connect(self.delete_selected_lots)
        self.delete_lot_btn.setEnabled(False)
        detail_layout.addWidget(self.delete_lot_btn)
        
        self.lots_splitter.addWidget(self.detail_panel)
        self.lots_splitter.setSizes([850, 450])

    def set_sort(self, column, order):
        self.current_sort = (column, order)
        for btn in self.sort_group:
            btn.setChecked(False)
        if column == "last_update": self.sort_new_btn.setChecked(True)
        elif column == "current_price" and order == "asc": self.sort_cheap_btn.setChecked(True)
        elif column == "current_price" and order == "desc": self.sort_expensive_btn.setChecked(True)
        self.load_lots()

    def on_header_clicked(self, index: int):
        # Маппинг колонок на поля ProcessedLot
        col_map = {
            0: "external_id",
            1: "title",
            2: "region_slug",
            3: "current_price",
            4: "market_price",
            5: "discount_percent",
            6: "total_area_gba",
            7: "land_area",
            8: "auction_status",
            9: "category",
            10: "risk_score",
            11: "last_update",
        }
        field = col_map.get(index)
        if field:
            new_order = "asc" if self.current_sort[0] != field or self.current_sort[1] == "desc" else "desc"
            self.set_sort(field, new_order)

    def _init_map_tab_legacy_removed(self):
        main_layout = QHBoxLayout(self.map_tab)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # Левая панель - карта
        map_container = QWidget()
        map_layout = QVBoxLayout(map_container)
        map_layout.setContentsMargins(0, 0, 0, 0)

        self.web_view = QWebEngineView()
        map_layout.addWidget(self.web_view)

        # Контролы карты
        controls = QHBoxLayout()
        controls.setContentsMargins(10, 5, 10, 5)

        map_btn = QPushButton("🔄 Обновить метки")
        map_btn.setFixedSize(150, 35)
        map_btn.setStyleSheet("background-color: #2ecc71; color: white; font-weight: bold; border-radius: 4px;")
        map_btn.clicked.connect(self.update_map)
        controls.addWidget(map_btn)

        self.cad_layer_checkbox = QCheckBox("Кадастровый слой")
        self.cad_layer_checkbox.setStyleSheet("font-weight: bold;")
        self.cad_layer_checkbox.stateChanged.connect(self.toggle_cadastral_layer)
        controls.addWidget(self.cad_layer_checkbox)

        controls.addStretch()
        map_layout.addLayout(controls)

        # Правая панель - поиск по кадастру
        sidebar = QWidget()
        sidebar.setMaximumWidth(350)
        sidebar.setStyleSheet("background: #f8f9fa; border-left: 1px solid #dee2e6;")
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(15, 15, 15, 15)

        # Заголовок
        header = QLabel("🔍 Поиск по кадастру")
        header.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        sidebar_layout.addWidget(header)

        # Поле ввода
        self.cadastral_search_input = QLineEdit()
        self.cadastral_search_input.setPlaceholderText("76:23:010101:15008")
        self.cadastral_search_input.setStyleSheet("padding: 8px; font-size: 13px; border: 1px solid #ced4da; border-radius: 4px;")
        sidebar_layout.addWidget(self.cadastral_search_input)

        # Кнопка поиска
        search_btn = QPushButton("🔎 Найти объект")
        search_btn.setFixedHeight(40)
        search_btn.setStyleSheet("background-color: #007bff; color: white; font-weight: bold; border-radius: 4px;")
        search_btn.clicked.connect(self.search_cadastral_object)
        sidebar_layout.addWidget(search_btn)

        # Результаты поиска
        self.cadastral_result_text = QTextEdit()
        self.cadastral_result_text.setReadOnly(True)
        self.cadastral_result_text.setStyleSheet("background: white; border: 1px solid #ced4da; border-radius: 4px; padding: 10px;")
        sidebar_layout.addWidget(self.cadastral_result_text)

        sidebar_layout.addStretch()

        # Добавляем в главный layout
        main_layout.addWidget(map_container, stretch=3)
        main_layout.addWidget(sidebar, stretch=1)

    def init_map_tab(self):
        main_layout = QHBoxLayout(self.map_tab)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        sidebar = QWidget()
        self.map_cadastre_sidebar = sidebar
        sidebar.setFixedWidth(360)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(12, 12, 12, 12)

        title = QLabel("Кадастровый поиск")
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        sidebar_layout.addWidget(title)

        self.cad_search_input = QLineEdit()
        self.cad_search_input.setPlaceholderText("Кадастровый номер или адрес")
        self.cad_suggestion_model = QStringListModel()
        self.cad_search_completer = QCompleter(self.cad_suggestion_model, self)
        self.cad_search_completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.cad_search_completer.setCompletionMode(QCompleter.PopupCompletion)
        self.cad_search_input.setCompleter(self.cad_search_completer)
        self.cad_search_input.textEdited.connect(self.update_cadastre_address_suggestions)
        sidebar_layout.addWidget(self.cad_search_input)

        self.cad_search_btn = QPushButton("Найти объект")
        self.cad_search_btn.clicked.connect(self.search_cadastre_from_gui)
        sidebar_layout.addWidget(self.cad_search_btn)

        self.cad_layer_checkbox = QCheckBox("Показать кадастровую подложку")
        self.cad_layer_checkbox.stateChanged.connect(self.toggle_cad_layer_from_gui)
        sidebar_layout.addWidget(self.cad_layer_checkbox)

        refresh_btn = QPushButton("Обновить метки лотов")
        refresh_btn.clicked.connect(self.refresh_all_map_markers)
        sidebar_layout.addWidget(refresh_btn)

        self.search_all_russia_btn = QPushButton("Поиск всех лотов РФ")
        self.search_all_russia_btn.setMinimumHeight(42)
        self.search_all_russia_btn.setStyleSheet(
            "QPushButton { background: #1f9d55; color: white; border: none; border-radius: 6px; "
            "font-weight: 700; padding: 8px; } QPushButton:disabled { background: #9bd3b4; }"
        )
        self.search_all_russia_btn.clicked.connect(self.run_all_russia_search)
        sidebar_layout.addWidget(self.search_all_russia_btn)

        self.cad_info_text = QTextEdit()
        self.cad_info_text.setReadOnly(True)
        self.cad_info_text.setPlaceholderText("Введите кадастровый номер или адрес")
        sidebar_layout.addWidget(self.cad_info_text, stretch=1)

        main_layout.addWidget(sidebar)

        self.map_view = QWebEngineView()
        self.map_view.settings().setAttribute(
            QWebEngineSettings.WebAttribute.LocalContentCanAccessRemoteUrls,
            True,
        )
        self.web_view = self.map_view
        self.map_web_channel = QWebChannel(self.map_view.page())
        self.map_web_channel.registerObject("bankrotaiBridge", self.map_bridge)
        self.map_view.page().setWebChannel(self.map_web_channel)
        main_layout.addWidget(self.map_view, stretch=1)

        self.update_map()

    def init_yandex_map_tab(self):
        main_layout = QHBoxLayout(self.yandex_map_tab)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        sidebar = QWidget()
        self.yandex_cadastre_sidebar = sidebar
        sidebar.setFixedWidth(360)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(12, 12, 12, 12)

        title = QLabel("Кадастровый поиск")
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        sidebar_layout.addWidget(title)

        self.yandex_cad_search_input = QLineEdit()
        self.yandex_cad_search_input.setPlaceholderText("Кадастровый номер или адрес")
        self.yandex_cad_suggestion_model = QStringListModel()
        self.yandex_cad_search_completer = QCompleter(self.yandex_cad_suggestion_model, self)
        self.yandex_cad_search_completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.yandex_cad_search_completer.setCompletionMode(QCompleter.PopupCompletion)
        self.yandex_cad_search_input.setCompleter(self.yandex_cad_search_completer)
        self.yandex_cad_search_input.textEdited.connect(self.update_yandex_cadastre_address_suggestions)
        sidebar_layout.addWidget(self.yandex_cad_search_input)

        self.yandex_cad_search_btn = QPushButton("Найти объект")
        self.yandex_cad_search_btn.clicked.connect(self.search_yandex_cadastre_from_gui)
        sidebar_layout.addWidget(self.yandex_cad_search_btn)

        self.yandex_cad_layer_checkbox = QCheckBox("Показать кадастровые границы")
        self.yandex_cad_layer_checkbox.setChecked(True)
        self.yandex_cad_layer_checkbox.stateChanged.connect(self.toggle_yandex_cad_layer_from_gui)
        sidebar_layout.addWidget(self.yandex_cad_layer_checkbox)

        refresh_btn = QPushButton("Обновить метки лотов")
        refresh_btn.clicked.connect(self.refresh_all_map_markers)
        sidebar_layout.addWidget(refresh_btn)

        self.yandex_search_all_russia_btn = QPushButton("Поиск всех лотов РФ")
        self.yandex_search_all_russia_btn.setMinimumHeight(42)
        self.yandex_search_all_russia_btn.setStyleSheet(
            "QPushButton { background: #1f9d55; color: white; border: none; border-radius: 6px; "
            "font-weight: 700; padding: 8px; } QPushButton:disabled { background: #9bd3b4; }"
        )
        self.yandex_search_all_russia_btn.clicked.connect(self.run_all_russia_search)
        sidebar_layout.addWidget(self.yandex_search_all_russia_btn)

        self.yandex_cad_info_text = QTextEdit()
        self.yandex_cad_info_text.setReadOnly(True)
        self.yandex_cad_info_text.setPlaceholderText("Введите кадастровый номер или адрес")
        sidebar_layout.addWidget(self.yandex_cad_info_text, stretch=1)

        main_layout.addWidget(sidebar)

        self.yandex_map_view = QWebEngineView()
        self.yandex_map_view.settings().setAttribute(
            QWebEngineSettings.WebAttribute.LocalContentCanAccessRemoteUrls,
            True,
        )
        self.yandex_map_web_channel = QWebChannel(self.yandex_map_view.page())
        self.yandex_map_web_channel.registerObject("bankrotaiBridge", self.map_bridge)
        self.yandex_map_view.page().setWebChannel(self.yandex_map_web_channel)
        main_layout.addWidget(self.yandex_map_view, stretch=1)

        self.update_yandex_map()

    def init_tools_tab(self):
        scroll_widget = QWidget()
        layout = QVBoxLayout(scroll_widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Section Header Helper
        def add_section(title, color_bg, color_border):
            box = QWidget()
            box.setStyleSheet(f"background: {color_bg}; border: 1px solid {color_border}; border-radius: 8px;")
            box_layout = QVBoxLayout(box)
            lbl = QLabel(f"<b>{title}</b>")
            lbl.setStyleSheet("font-size: 14px; border: none; background: transparent;")
            box_layout.addWidget(lbl)
            return box, box_layout

        # 1. Online Sync (Auto Parser)
        online_box, online_layout = add_section("🌐 Автоматический парсер (Онлайн)", "#e8f8f5", "#a3e4d7")
        desc_online = QLabel("Поиск новых лотов на агрегаторах (GorodTorgi, TBankrot) по заданному региону.")
        desc_online.setStyleSheet("border: none; color: #16a085; background: transparent;")
        online_layout.addWidget(desc_online)
        
        region_layout = QHBoxLayout()
        region_layout.addWidget(QLabel("Регион:"))
        self.region_combo = QComboBox()
        self.region_mapping = {
            "Ярославская обл.": "yaroslavl",
            "Москва": "moskow",
            "Санкт-Петербург": "spb",
            "Ростовская обл.": "rostov",
            "Свердловская обл.": "sverdlovsk",
            "Кировская обл.": "kirov",
            "Псковская обл.": "pskov"
        }
        self.region_combo.addItems(list(self.region_mapping.keys()))
        self.region_combo.setFixedWidth(180)
        region_layout.addWidget(self.region_combo)
        region_layout.addStretch()
        online_layout.addLayout(region_layout)

        self.sync_btn = QPushButton("🚀 Запустить поиск новых лотов")
        self.sync_btn.setFixedHeight(45)
        self.sync_btn.setStyleSheet("background-color: #1abc9c; color: white; font-weight: bold; border-radius: 5px;")
        self.sync_btn.clicked.connect(self.run_online_sync)
        online_layout.addWidget(self.sync_btn)
        # Excel export button
        self.export_btn = QPushButton("📊 Экспорт в Excel")
        self.export_btn.setFixedHeight(45)
        self.export_btn.setStyleSheet("background-color: #2ecc71; color: white; font-weight: bold; border-radius: 5px;")
        self.export_btn.clicked.connect(self.export_to_excel)
        online_layout.addWidget(self.export_btn)
        layout.addWidget(online_box)
        
        # 2. Offline Parser (HTML)
        offline_box, offline_layout = add_section("📂 Оффлайн парсер (HTML файлы)", "#fef9e7", "#f9e79f")
        desc_offline = QLabel("Загрузка лотов из сохраненных HTML-страниц (TBankrot и др.).")
        desc_offline.setStyleSheet("border: none; color: #d4ac0d; background: transparent;")
        offline_layout.addWidget(desc_offline)
        
        self.html_btn = QPushButton("📄 Выбрать HTML файл для импорта")
        self.html_btn.setFixedHeight(45)
        self.html_btn.setStyleSheet("background-color: #f1c40f; color: #7d6608; font-weight: bold; border-radius: 5px;")
        self.html_btn.clicked.connect(self.run_offline_parse)
        offline_layout.addWidget(self.html_btn)
        layout.addWidget(offline_box)
        
        # 3. AI Appraisal
        ai_box, ai_layout = add_section("🤖 Искусственный интеллект (OpenAI)", "#ebf5fb", "#aed6f1")
        desc_ai = QLabel("Массовая оценка инвестиционной привлекательности и рисков для новых лотов.")
        desc_ai.setStyleSheet("border: none; color: #2e86c1; background: transparent;")
        ai_layout.addWidget(desc_ai)
        
        self.ai_batch_btn = QPushButton("🧬 Оценить все необработанные лоты")
        self.ai_batch_btn.setFixedHeight(45)
        self.ai_batch_btn.setStyleSheet("background-color: #3498db; color: white; font-weight: bold; border-radius: 5px;")
        self.ai_batch_btn.clicked.connect(self.run_ai_batch)
        ai_layout.addWidget(self.ai_batch_btn)
        
        self.geo_batch_btn = QPushButton("🗺️ Массовое геокодирование")
        self.geo_batch_btn.setFixedHeight(45)
        self.geo_batch_btn.setStyleSheet("background-color: #f39c12; color: white; font-weight: bold; border-radius: 5px;")
        self.geo_batch_btn.clicked.connect(self.run_geo_batch)
        ai_layout.addWidget(self.geo_batch_btn)
        
        layout.addWidget(ai_box)

        # 2.3 UI: AI status overview (simple panel)
        ai_status_box, ai_status_layout = add_section("🧭 AI Статус", "#eef7ff", "#a5c8e8")
        self.ai_status_label = QLabel("Статус: не запущено")
        ai_status_layout.addWidget(self.ai_status_label)
        self.ai_status_progress = QProgressBar()
        self.ai_status_progress.setRange(0, 100)
        self.ai_status_progress.setValue(0)
        ai_status_layout.addWidget(self.ai_status_progress)
        self.export_btn.setVisible(True)  # ensure export exists on UI
        layout.addWidget(ai_status_box)

        # 4. Maintenance
        clean_box, clean_layout = add_section("🧹 Обслуживание базы данных", "#fdf2f2", "#f5c6cb")
        self.cleanup_btn = QPushButton("🗑️ Удалить завершенные торги (архив)")
        self.cleanup_btn.setFixedHeight(40)
        self.cleanup_btn.setStyleSheet("background-color: #e74c3c; color: white; font-weight: bold; border-radius: 5px;")
        self.cleanup_btn.clicked.connect(self.run_cleanup)
        clean_layout.addWidget(self.cleanup_btn)
        layout.addWidget(clean_box)

        # 5. AI Settings
        from bankrotai.core import get_app_setting
        ai_set_box, ai_set_layout = add_section("⚙️ Настройки AI Провайдера", "#f4f6f7", "#ccd1d1")
        form = QFormLayout()
        
        self.provider_combo = QComboBox()
        for provider_id, label in AI_PROVIDER_OPTIONS:
            self.provider_combo.addItem(label, provider_id)
        current_provider = get_app_setting("ai_provider", "omniroute") or "omniroute"
        provider_index = self.provider_combo.findData(current_provider)
        self.provider_combo.setCurrentIndex(provider_index if provider_index >= 0 else 0)
        self.provider_combo.currentIndexChanged.connect(self.on_ai_provider_changed)
        form.addRow("Провайдер:", self.provider_combo)
        
        self.api_key_input = QLineEdit()
        self.api_key_input.setEchoMode(QLineEdit.Password)
        self.api_key_input.setReadOnly(True)
        self.api_key_input.setPlaceholderText("Задаётся через переменную окружения / secret manager")
        form.addRow("API ключ:", self.api_key_input)
        
        self.model_search_input = QComboBox()
        form.addRow("Модель поиска:", self.model_search_input)
        
        save_settings_btn = QPushButton("💾 Сохранить настройки AI")
        self.on_ai_provider_changed()
        save_settings_btn.clicked.connect(self.save_ai_settings)
        ai_set_layout.addLayout(form)
        ai_set_layout.addWidget(save_settings_btn)
        layout.addWidget(ai_set_box)

        layout.addStretch()
        
        # Set tools_tab content
        main_tools_layout = QVBoxLayout(self.tools_tab)
        main_tools_layout.addWidget(scroll_widget)

    def update_dashboard(self):
        try:
            region_display = self.region_combo.currentText()
            city_slug = self.region_mapping.get(region_display, "yaroslavl")
            
            with session_scope() as session:
                total = session.query(func.count(ProcessedLot.id)).scalar()
                active = session.query(func.count(ProcessedLot.id)).filter(ProcessedLot.auction_status == "active").scalar()
                with_rating = session.query(func.count(ProcessedLot.id)).filter(ProcessedLot.rating.isnot(None)).scalar()
                
                from bankrotai.db import RegionSyncState
                sync_state = session.query(RegionSyncState).filter(RegionSyncState.city_slug == city_slug).first()
                last_sync = sync_state.last_success_at.strftime("%d.%m.%Y %H:%M") if sync_state and sync_state.last_success_at else "никогда"
                
                text = (
                    f"📈 <b>Общая статистика ({region_display}):</b>\n\n"
                    f"• Всего лотов в базе: {total}\n"
                    f"• Активных торгов: {active}\n"
                    f"• Оценено Искусственным Интеллектом: {with_rating}\n\n"
                    f"🔄 Последняя онлайн синхронизация: {last_sync}"
                )
                self.stats_label.setText(text)
        except Exception as e:
            self.stats_label.setText(f"Ошибка статистики: {e}")
        self.load_lots() # Refresh table as well


    def load_lots(self):
        search_text = self.search_input.text().strip()
        category_filter = self.category_combo.currentText()
        sort_col, sort_order = self.current_sort

        # Disable sorting while loading items
        self.lots_table.setSortingEnabled(False)

        with session_scope() as session:
            query = select(ProcessedLot).where(ProcessedLot.duplicate_of_id.is_(None))
            
            if search_text:
                query = query.where(ProcessedLot.title.ilike(f"%{search_text}%"))
            
            if category_filter != "Все категории":
                # Маппинг категорий на внутренние значения
                cat_map = {
                    "Жилая недвижимость": ["apartment", "house", "living"],
                    "Коммерческая недвижимость": [
                        "commercial", 
                        "commercial_room", 
                        "commercial_building", 
                        "commercial_building_with_land", 
                        "complex", 
                        "office", 
                        "retail"
                    ],
                    "Земельные участки": ["land"],
                    "Транспорт": ["car", "transport", "vehicle"],
                    "Прочее": ["equipment", "parking", "unfinished", "other"]
                }
                allowed_cats = cat_map.get(category_filter, [])
                if allowed_cats:
                    query = query.where(ProcessedLot.category.in_(allowed_cats))

            # Сортировка по умолчанию (БД)
            col_attr = getattr(ProcessedLot, sort_col)
            if sort_order == "desc":
                query = query.order_by(desc(col_attr))
            else:
                query = query.order_by(col_attr)

            lots = session.scalars(query).all()
            self.lots_table.setRowCount(len(lots))
            for i, lot in enumerate(lots):
                # Helper for numeric items
                def num_item(val, suffix=""):
                    if val is None: return QTableWidgetItem("—")
                    try:
                        # Convert to float then to int to ensure no decimals and no scientific notation
                        iv = int(float(val))
                        # Use fixed-point formatting with space as thousands separator
                        formatted_val = f"{iv:,}".replace(",", " ")
                        item = QTableWidgetItem(f"{formatted_val}{suffix}")
                        item.setData(Qt.EditRole, iv)
                        return item
                    except (ValueError, TypeError):
                        return QTableWidgetItem(str(val))

                def format_land_area(area_m2: float | None) -> str: 
                    if not area_m2: 
                        return "" 
                    sotki = area_m2 / 100 
                    return f"{area_m2:,.0f} м² / {sotki:.2f} сот.".replace(",", " ")

                items = [
                    QTableWidgetItem(str(lot.external_id)),
                    QTableWidgetItem(lot.title),
                    QTableWidgetItem(lot.region_slug or ""),
                    num_item(lot.current_price, " ₽"),
                    num_item(lot.market_price, " ₽"),
                    num_item(lot.discount_percent, "%"),
                    QTableWidgetItem(f"{lot.total_area_gba:,.1f} м²".replace(",", " ") if lot.total_area_gba else ""),
                    QTableWidgetItem(format_land_area(lot.land_area)),
                    QTableWidgetItem(translate_status(lot.auction_status)),
                    QTableWidgetItem(translate_category(lot.category)),
                    QTableWidgetItem(str(lot.risk_score) if lot.risk_score is not None else ""),
                    QTableWidgetItem(lot.last_update.strftime("%d.%m.%Y %H:%M"))
                ]
                items[0].setData(LOT_ID_ROLE, lot.id)
                
                # Review Status Color
                color = QColor("white")
                if lot.review_status == "approved": color = QColor("#d4edda")
                elif lot.review_status == "rejected": color = QColor("#f8d7da")
                elif lot.review_status == "maybe": color = QColor("#fff3cd")
                
                for j, item in enumerate(items):
                    item.setBackground(QBrush(color))
                    self.lots_table.setItem(i, j, item)

        # Re-enable sorting
        self.lots_table.setSortingEnabled(True)

    def on_lot_selected(self):
        selected_items = self.lots_table.selectedItems()
        if not selected_items:
            self.delete_lot_btn.setEnabled(False)
            self.max_bid_btn.setEnabled(False)
            self.ai_single_btn.setEnabled(False)
            self.geo_fix_btn.setEnabled(False)
            self.geo_fix_btn.setText("🗺️ Гео")
            for btn in [self.review_approved_btn, self.review_maybe_btn, self.review_rejected_btn]:
                btn.setEnabled(False)
            return
        rows = sorted(list(set(item.row() for item in selected_items)))
        count = len(rows)
        self.delete_lot_btn.setEnabled(True)
        self.delete_lot_btn.setText(f"🗑️ Удалить выбранные ({count})")
        self.max_bid_btn.setEnabled(count == 1)
        
        # AI evaluation enabled for 1 or more lots
        self.ai_single_btn.setEnabled(True)
        self.ai_single_btn.setText(f"🤖 Оценить AI ({count})" if count > 1 else "🤖 Оценить AI")
        
        if count == 1:
            self.geo_fix_btn.setEnabled(True)
            self.geo_fix_btn.setText("🗺️ Гео")
            for btn in [self.review_approved_btn, self.review_maybe_btn, self.review_rejected_btn]:
                btn.setEnabled(True)
            
            lot_id = self._lot_id_for_row(rows[0])
            with session_scope() as session:
                lot = session.get(ProcessedLot, lot_id) if lot_id is not None else None
                if lot:
                    self.current_selected_lot_id = lot.id
                    
                    # Helper for integer formatting in text
                    def fmt_int(val):
                        if val is None: return "—"
                        try:
                            return f"{int(float(val)):,}".replace(",", " ")
                        except (ValueError, TypeError):
                            return str(val)

                    info = [
                        f"🏷️ ЛОТ: {lot.title}",
                        f"🆔 КОД: {lot.external_id}",
                        f"📍 АДРЕС: {lot.address or 'не указан'}",
                        f"💰 ЦЕНА: {fmt_int(lot.current_price)} ₽",
                        f"📊 СТАТУС: {translate_status(lot.auction_status)}",
                        f"📈 РЕЙТИНГ: {fmt_int(lot.rating)}",
                    ]
                    
                    # Добавляем новые инвестиционные поля если они заполнены
                    if lot.object_name: info.append(f"🏢 ОБЪЕКТ: {lot.object_name}")
                    if lot.property_type: info.append(f"🏗️ ТИП: {lot.property_type}")
                    if lot.total_area_gba: info.append(f"📐 ПЛОЩАДЬ (GBA): {fmt_int(lot.total_area_gba)} м²")
                    if lot.cadastral_number: info.append(f"🔢 КАДАСТР: {lot.cadastral_number}")
                    if lot.floors: info.append(f"🏢 ЭТАЖЕЙ: {fmt_int(lot.floors)}")
                    if lot.legal_status: info.append(f"⚖️ СТАТУС: {lot.legal_status}")
                    
                    info.append("-" * 20)
                    
                    # Clean description from HTML if any, to avoid accidental italics
                    desc_text = lot.description or ""
                    if "<" in desc_text and ">" in desc_text:
                        desc_text = re.sub('<[^<]+?>', '', desc_text)
                    
                    info.append(f"📑 ОПИСАНИЕ:\n{desc_text}")
                    
                    # Use setPlainText to be 100% sure no HTML formatting is applied
                    self.detail_text.setPlainText("\n".join(info))
                    self.detail_ai.setText(lot.ai_recommendation or "AI анализ не проводился.")
        else:
            self.geo_fix_btn.setEnabled(True)
            self.geo_fix_btn.setText(f"Гео ({count})")
            for btn in [self.review_approved_btn, self.review_maybe_btn, self.review_rejected_btn]:
                btn.setEnabled(False)
            self.detail_text.setPlainText(f"Выбрано объектов: {count}")
            self.detail_ai.clear()

    def open_lot_url(self, row: int, column: int) -> None:
        lot_id = self._lot_id_for_row(row)
        if lot_id is None:
            return
        with session_scope() as session:
            lot = session.get(ProcessedLot, lot_id)
            url = (lot.source_url or lot.lot_url) if lot else None
            if url:
                QDesktopServices.openUrl(QUrl.fromUserInput(str(url)))
            else:
                QMessageBox.information(self, "Инфо", "Ссылка на лот отсутствует.")

    def export_to_excel(self) -> None: 
        """Экспорт текущей таблицы в Excel""" 
        default_name = f"bankrotai_export_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx" 
        desktop = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation) 
        default_path = f"{desktop}/{default_name}" 

        path, _ = QFileDialog.getSaveFileName( 
            self, 
            "Сохранить Excel файл", 
            default_path, 
            "Excel files (*.xlsx)" 
        ) 
        if not path: 
            return 

        try: 
            with session_scope() as session: 
                search_text = self.search_input.text().strip() 
                category_filter = self.category_combo.currentText() 
                sort_col, sort_order = self.current_sort 

                query = select(ProcessedLot) 

                if search_text: 
                    query = query.where(ProcessedLot.title.ilike(f"%{search_text}%")) 

                if category_filter != "Все категории": 
                    cat_map = { 
                        "Жилая недвижимость": ["apartment", "house", "living"], 
                        "Коммерческая недвижимость": [
                            "commercial", 
                            "commercial_room", 
                            "commercial_building", 
                            "commercial_building_with_land", 
                            "complex", 
                            "office", 
                            "retail"
                        ], 
                        "Земельные участки": ["land"], 
                        "Транспорт": ["car", "transport", "vehicle"], 
                        "Прочее": ["equipment", "parking", "unfinished", "other"], 
                    } 
                    allowed = cat_map.get(category_filter, []) 
                    if allowed: 
                        query = query.where(ProcessedLot.category.in_(allowed)) 

                col_attr = getattr(ProcessedLot, sort_col) 
                query = query.order_by(desc(col_attr) if sort_order == "desc" else col_attr) 

                lots = session.scalars(query).all() 

                # Полностью материализуем все нужные данные пока сессия открыта 
                data = [] 
                for lot in lots: 
                    # Helper for safe integer conversion
                    def to_int(val):
                        try:
                            return int(float(val)) if val is not None else None
                        except (ValueError, TypeError):
                            return None

                    row = { 
                        "Код": lot.external_id, 
                        "Название": lot.title, 
                        "Описание": lot.description, 
                        "Категория": translate_category(lot.category), 
                        "Регион": lot.region_slug, 
                        "Адрес": lot.address, 
                        "Кадастровый номер": lot.cadastral_number, 
                        "Все кадастровые номера": ", ".join((lot.cadastral_numbers or [])) if hasattr(lot, "cadastral_numbers") and lot.cadastral_numbers else None,
                        "Текущая цена": to_int(lot.current_price), 
                        "Рыночная цена": to_int(lot.market_price), 
                        "Дисконт %": to_int(lot.discount_percent), 
                        "Риск (1-10)": to_int(lot.risk_score), 
                        "Рейтинг": to_int(lot.rating), 
                        "Статус торгов": translate_status(lot.auction_status), 
                        "Ссылка": lot.lot_url, 
                        "Площадь здания/помещения, м²": lot.total_area_gba, 
                        "Площадь участка, м²": lot.land_area, 
                        "Площадь участка, сот.": round(lot.land_area / 100, 2) if lot.land_area else None, 
                        "Этажность": to_int(lot.floors), 
                        "Юридический статус": lot.legal_status, 
                        "Статус проверки": lot.review_status, 
                        "Обновление": lot.last_update.strftime("%d.%m.%Y %H:%M") if lot.last_update else None, 
                    } 
                    data.append(row) 

            # === Экспорт после закрытия сессии === 
            wb = Workbook() 
            ws = wb.active 
            ws.title = "Реестр лотов" 

            if not data:
                QMessageBox.information(self, "Инфо", "Нет данных для экспорта.")
                return

            headers = list(data[0].keys()) 
            ws.append(headers) 

            # Identify numeric columns for formatting
            numeric_headers = {
                "Текущая цена", "Рыночная цена", "Дисконт %", 
                "Риск (1-10)", "Рейтинг", "Площадь здания/помещения, м²", 
                "Площадь участка, м²", "Площадь участка, сот.", "Этажность"
            }

            for row_dict in data: 
                row_data = [row_dict.get(h) for h in headers]
                ws.append(row_data)
                
                # Format numeric cells in the last added row
                curr_row = ws.max_row
                for idx, header in enumerate(headers):
                    if header in numeric_headers:
                        cell = ws.cell(row=curr_row, column=idx + 1)
                        if cell.value is not None:
                            # Use number format with thousands separator and no scientific notation
                            # Force as integer format
                            cell.number_format = '#,##0'

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 50) # limit width

            wb.save(path) 
            self.status_bar.showMessage(f"Экспорт завершён: {path}", 5000) 
            QMessageBox.information(self, "Успех", f"Файл успешно сохранён:\n{path}") 

        except PermissionError: 
            QMessageBox.critical(self, "Ошибка прав доступа", 
                "Нет прав на запись в выбранную папку.\n\nПопробуйте сохранить на Рабочий стол.") 
        except Exception as e: 
            logger.exception("Export error") 
            QMessageBox.critical(self, "Ошибка экспорта", f"Произошла ошибка:\n{str(e)}") 

    def delete_selected_lots(self):
        selected_items = self.lots_table.selectedItems()
        if not selected_items: return
        rows = sorted(list(set(item.row() for item in selected_items)))
        lot_ids = [lot_id for row in rows if (lot_id := self._lot_id_for_row(row)) is not None]
        if QMessageBox.question(self, "Удаление", f"Удалить {len(lot_ids)} лотов?") == QMessageBox.Yes:
            with session_scope() as session:
                delete_lots_batch(session, lot_ids)
            self.load_lots()
            self.update_dashboard()

    def _lot_id_for_row(self, row: int) -> int | None:
        item = self.lots_table.item(row, 0)
        if item is None:
            return None
        value = item.data(LOT_ID_ROLE)
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def open_max_bid_calculator(self):
        lot_ids = self.get_selected_lot_ids()
        if len(lot_ids) != 1:
            QMessageBox.information(self, "Калькулятор", "Выберите один лот.")
            return
        with session_scope() as session:
            lot = session.get(ProcessedLot, lot_ids[0])
            intended_bid = float(lot.current_price or 0) if lot else 0
        MaxBidDialog(intended_bid=intended_bid, parent=self).exec()

    def get_selected_lot_ids(self) -> list[int]:
        selected_items = self.lots_table.selectedItems()
        if not selected_items:
            return []

        rows = sorted({item.row() for item in selected_items})
        return [lot_id for row in rows if (lot_id := self._lot_id_for_row(row)) is not None]

    def start_geo_worker(
        self,
        *,
        lot_ids: list[int] | None = None,
        refresh_existing: bool = False,
        limit: int | None = None,
    ):
        if getattr(self, "geo_worker", None) and self.geo_worker.isRunning():
            QMessageBox.information(self, "GEO", "Геокодирование уже выполняется. Дождитесь завершения текущей задачи.")
            return
        self.geo_fix_btn.setEnabled(False)
        if hasattr(self, "geo_batch_btn"):
            self.geo_batch_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.start_task_progress("geo", "GEO")

        self.geo_worker = GeoWorker(
            limit=limit,
            lot_ids=lot_ids,
            refresh_existing=refresh_existing,
        )
        self.geo_worker.progress.connect(self.status_bar.showMessage)
        self.geo_worker.progress_percent.connect(self.progress_bar.setValue)
        self.geo_worker.progress_percent.connect(lambda value: self.update_task_progress("geo", value))
        self.geo_worker.lot_processed.connect(self.on_geo_lot_processed)
        self.geo_worker.finished.connect(self.on_geo_finished)
        self.geo_worker.error.connect(self.on_worker_error)
        self.geo_worker.start()

    def run_online_sync(self):
        if getattr(self, "geo_worker", None) and self.geo_worker.isRunning():
            QMessageBox.warning(
                self,
                "Фоновая задача",
                "Сейчас идет массовое геокодирование. Дождитесь завершения GEO перед запуском поиска новых лотов.",
            )
            return
        if getattr(self, "sync_worker", None) and self.sync_worker.isRunning():
            return
        self.sync_btn.setEnabled(False)
        self.sync_btn.setText("⏳ Синхронизация...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0) # Indeterminate
        
        region_display = self.region_combo.currentText()
        city_slug = self.region_mapping.get(region_display, "yaroslavl")
        
        self.sync_worker = SyncWorker(city_slug)
        self.sync_worker.progress.connect(self.status_bar.showMessage)
        self.sync_worker.finished.connect(self.on_sync_finished)
        self.sync_worker.error.connect(self.on_worker_error)
        self.sync_worker.start()

    def on_sync_finished(self, count: int):
        self.sync_btn.setEnabled(True)
        self.sync_btn.setText("🚀 Запустить поиск новых лотов")
        self.progress_bar.setVisible(False)
        self.status_bar.showMessage(f"Синхронизация завершена. Найдено лотов: {count}", 5000)
        QMessageBox.information(self, "Готово", f"Синхронизация завершена! Найдено лотов: {count}")
        self.load_lots()
        self.update_dashboard()

    def run_offline_parse(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Выбрать HTML файл", "", "HTML Files (*.html *.htm)")
        if not file_path: return
        
        self.html_btn.setEnabled(False)
        self.status_bar.showMessage(f"Импорт из файла: {file_path}...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        
        self.import_worker = ImportWorker(file_path, "yaroslavl")
        self.import_worker.progress.connect(self.status_bar.showMessage)
        self.import_worker.finished.connect(self.on_import_finished)
        self.import_worker.error.connect(self.on_worker_error)
        self.import_worker.start()

    def on_import_finished(self, new: int, upd: int, skip: int):
        self.html_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        msg = f"Импорт завершен.\nДобавлено новых: {new}\nОбновлено: {upd}\nПропущено (проверены): {skip}"
        self.status_bar.showMessage(f"Импорт: +{new}, ~{upd}, ={skip}", 5000)
        QMessageBox.information(self, "Результат импорта", msg)
        self.load_lots()
        self.update_dashboard()

    def get_appraiser(self) -> OpenAIAppraiser | None:
        try:
            if self._appraiser is None:
                self._appraiser = OpenAIAppraiser()
            return self._appraiser
        except Exception as e:
            logger.error(f"Failed to initialize AI Appraiser: {e}")
            QMessageBox.critical(self, "Ошибка AI", f"Не удалось инициализировать модуль AI:\n\n{str(e)}")
            return None

    def run_ai_single(self):
        selected_items = self.lots_table.selectedItems()
        if not selected_items: return
        
        rows = sorted(list(set(item.row() for item in selected_items)))
        ext_ids = [self.lots_table.item(r, 0).text() for r in rows]
        
        appraiser = self.get_appraiser()
        if not appraiser: return

        with session_scope() as session:
            lots = session.scalars(select(ProcessedLot.id).where(ProcessedLot.external_id.in_(ext_ids))).all()
            lot_ids = list(lots)

        if not lot_ids: return

        self.ai_single_btn.setEnabled(False)
        self.ai_last_failed_count = 0
        self.ai_status_label.setText("Запуск оценки...") 
        self.ai_status_progress.setValue(0) 
        self.status_bar.showMessage(f"AI Оценка выбранных лотов ({len(lot_ids)})...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        
        self.ai_worker = AIWorker(appraiser=appraiser, lot_ids=lot_ids)
        self.ai_worker.progress.connect(self.status_bar.showMessage)
        self.ai_worker.progress.connect(self.ai_status_label.setText) 
        self.ai_worker.progress_percent.connect(self.ai_status_progress.setValue) 
        self.ai_worker.progress_percent.connect(self.progress_bar.setValue)
        self.ai_worker.progress_percent.connect(lambda value: self.update_task_progress("ai", value))
        self.ai_worker.lot_finished.connect(self.on_ai_lot_finished)
        self.ai_worker.lot_failed.connect(self.on_ai_lot_failed)
        self.ai_worker.finished.connect(self.on_ai_finished)
        self.ai_worker.error.connect(self.on_worker_error)
        self.ai_worker.start()

    def run_ai_batch(self):
        appraiser = self.get_appraiser()
        if not appraiser: return

        self.ai_batch_btn.setEnabled(False)
        self.ai_last_failed_count = 0
        self.ai_status_label.setText("Запуск пакетной оценки...") 
        self.ai_status_progress.setValue(0) 
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.start_task_progress("ai", "AI")
        
        self.ai_worker = AIWorker(appraiser=appraiser, limit=15) # Увеличим лимит до 15
        self.ai_worker.progress.connect(self.status_bar.showMessage)
        self.ai_worker.progress.connect(self.ai_status_label.setText) 
        self.ai_worker.progress_percent.connect(self.ai_status_progress.setValue) 
        self.ai_worker.progress_percent.connect(self.progress_bar.setValue)
        self.ai_worker.progress_percent.connect(lambda value: self.update_task_progress("ai", value))
        self.ai_worker.lot_finished.connect(self.on_ai_lot_finished)
        self.ai_worker.lot_failed.connect(self.on_ai_lot_failed)
        self.ai_worker.finished.connect(self.on_ai_finished)
        self.ai_worker.error.connect(self.on_worker_error)
        self.ai_worker.start()

    def run_geo_batch(self):
        self.start_geo_worker(limit=None, refresh_existing=False)

    def refresh_all_map_markers(self):
        self.status_bar.showMessage("Обновление геометок лотов из базы...")
        self.start_geo_worker(limit=None, refresh_existing=False)

    def _set_all_russia_buttons_enabled(self, enabled: bool) -> None:
        for name in ("search_all_russia_btn", "yandex_search_all_russia_btn"):
            button = getattr(self, name, None)
            if button is not None:
                button.setEnabled(enabled)

    def run_all_russia_search(self) -> None:
        worker = getattr(self, "all_russia_worker", None)
        if worker is not None and worker.isRunning():
            worker.request_stop()
            self.status_bar.showMessage("Останавливаю поиск после текущей страницы...", 5000)
            return
        self._set_all_russia_buttons_enabled(False)
        self.start_task_progress("all_russia", "РФ")
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)
        self.all_russia_worker = AllRussiaRealEstateWorker()
        self.all_russia_worker.progress.connect(self.status_bar.showMessage)
        self.all_russia_worker.source_finished.connect(
            lambda source, count, unique: self.status_bar.showMessage(
                f"{source}: {count}; уникальных карточек РФ: {unique}", 5000
            )
        )
        self.all_russia_worker.result_ready.connect(self.on_all_russia_search_finished)
        self.all_russia_worker.error.connect(self.on_all_russia_search_error)
        self.all_russia_worker.start()

    def on_all_russia_search_finished(self, lot_ids: object, summary: object) -> None:
        self.finish_task_progress("all_russia")
        self._set_all_russia_buttons_enabled(True)
        self.progress_bar.setVisible(False)
        ids = [int(value) for value in lot_ids] if isinstance(lot_ids, list) else []
        details = summary if isinstance(summary, dict) else {}
        self.load_lots()
        self.update_map()
        self.update_yandex_map()
        self.status_bar.showMessage(
            f"Поиск РФ завершён: уникальных карточек {len(ids)}. Запускаю геокодирование...",
            8000,
        )
        if details.get("errors"):
            logger.warning("Nationwide search completed with source errors: %s", details["errors"])
        if ids:
            self.start_geo_worker(lot_ids=ids, refresh_existing=False, limit=None)

    def on_all_russia_search_error(self, message: str) -> None:
        self.finish_task_progress("all_russia")
        self._set_all_russia_buttons_enabled(True)
        self.progress_bar.setVisible(False)
        QMessageBox.warning(self, "Поиск всех лотов РФ", message)

    def on_geo_lot_processed(self, lot_id: int, success: bool, completed: int, total: int):
        if not success:
            return
        payload = self._load_map_lot(lot_id)
        if not payload:
            return
        encoded = json.dumps(payload, ensure_ascii=False)
        js = (
            f"window.__bankrotaiPendingLots = window.__bankrotaiPendingLots || []; "
            f"if (window.upsertLot) {{ window.upsertLot({encoded}); }} "
            f"else {{ window.__bankrotaiPendingLots.push({encoded}); }}"
        )
        self.map_view.page().runJavaScript(js)
        if hasattr(self, "yandex_map_view"):
            self.yandex_map_view.page().runJavaScript(js)
        self.status_bar.showMessage(
            f"GEO: точка добавлена на карту сразу ({completed}/{total})", 1500
        )

    def on_geo_finished(self, count: int, failed_count: int, skipped_existing: int):
        self.finish_task_progress("geo")
        if hasattr(self, "geo_batch_btn"):
            self.geo_batch_btn.setEnabled(True)
        self.geo_fix_btn.setEnabled(bool(self.lots_table.selectedItems()))
        self.progress_bar.setVisible(False)
        if count == 0 and failed_count == 0:
            message = f"Все выбранные лоты уже геокодированы. Пропущено: {skipped_existing}"
            self.status_bar.showMessage(message, 5000)
            QMessageBox.information(self, "Инфо", message)
        else:
            message = (
                f"Геокодирование завершено. Координаты: {count}, "
                f"не найдено: {failed_count}, уже были готовы: {skipped_existing}"
            )
            self.status_bar.showMessage(message, 5000)
            QMessageBox.information(self, "Готово", message)
        self.update_map()
        if hasattr(self, "yandex_map_view"):
            self.update_yandex_map()

    def on_ai_lot_finished(self, lot_id: int, processed_count: int):
        self.load_lots()
        self.update_dashboard()
        self.status_bar.showMessage(f"AI оценка готова для лота ID {lot_id}. Обработано: {processed_count}", 5000)
        self.ai_status_label.setText(f"Готова оценка лота ID {lot_id}. Обработано: {processed_count}")

    def on_ai_lot_failed(self, lot_id: int, error_msg: str, failed_count: int):
        self.ai_last_failed_count = failed_count
        short_error = (error_msg or "").replace("\n", " ")[:160]
        self.status_bar.showMessage(
            f"AI ошибка лота ID {lot_id}; продолжаю. Ошибок: {failed_count}",
            7000,
        )
        self.ai_status_label.setText(
            f"Пропущен лот ID {lot_id}. Ошибок: {failed_count}. {short_error}"
        )

    def on_ai_finished(self, count: int):
        self.finish_task_progress("ai")
        self.ai_batch_btn.setEnabled(True)
        self.ai_single_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.ai_status_label.setText("Готово") 
        self.ai_status_progress.setValue(0) 
        failed_count = getattr(self, "ai_last_failed_count", 0)
        if count == 0 and failed_count:
            message = f"AI оценка завершена с ошибками. Успешно: 0, ошибок: {failed_count}"
            self.status_bar.showMessage(message, 5000)
            QMessageBox.information(self, "Готово", message)
        elif count == 0:
            self.status_bar.showMessage("Нет лотов для оценки", 5000)
            QMessageBox.information(self, "Инфо", "Нет лотов для оценки.")
        elif failed_count:
            message = f"AI оценка завершена. Успешно: {count}, ошибок: {failed_count}"
            self.status_bar.showMessage(message, 5000)
            QMessageBox.information(self, "Готово", message)
        else:
            self.status_bar.showMessage(f"AI Оценка завершена. Обработано: {count}", 5000)
            QMessageBox.information(self, "Готово", f"AI оценка завершена. Обработано лотов: {count}")
        self.load_lots()
        self.update_dashboard()

    def on_worker_error(self, error_msg: str):
        # Сброс состояния всех кнопок
        self.sync_btn.setEnabled(True)
        self.sync_btn.setText("🚀 Запуск поиск новых лотов")
        self.html_btn.setEnabled(True)
        self.ai_batch_btn.setEnabled(True)
        self.ai_single_btn.setEnabled(True)
        if hasattr(self, "geo_batch_btn"):
            self.geo_batch_btn.setEnabled(True)
        if hasattr(self, "geo_fix_btn"):
            self.geo_fix_btn.setEnabled(bool(self.lots_table.selectedItems()))
        self.finish_task_progress("ai")
        self.finish_task_progress("geo")
        self.finish_task_progress("sync")
        self.finish_task_progress("import")
        self.progress_bar.setVisible(False)
        self.ai_status_label.setText("Ошибка") 
        self.ai_status_progress.setValue(0) 
        
        self.status_bar.showMessage(f"Ошибка: {error_msg[:50]}...", 10000)
        
        if "429" in error_msg:
            QMessageBox.warning(self, "Лимит запросов", 
                "Достигнут лимит запросов к AI (Error 429).\n\n"
                "Система автоматически делала повторные попытки, но лимит все еще активен. "
                "Пожалуйста, подождите 1-2 минуты или проверьте настройки API ключа.")
        else:
            QMessageBox.critical(self, "Ошибка воркера", f"Произошла ошибка при выполнении фоновой задачи:\n\n{error_msg}")

    def on_ai_provider_changed(self):
        from bankrotai.core import get_app_setting
        provider = self.provider_combo.currentData() or "omniroute"
        current_model = get_app_setting(f"{provider}_model", "")

        self.api_key_input.clear()
        self.model_search_input.clear()
        for model_id, label in AI_MODEL_OPTIONS.get(provider, []):
            self.model_search_input.addItem(label, model_id)

        model_index = self.model_search_input.findData(current_model)
        self.model_search_input.setCurrentIndex(model_index if model_index >= 0 else 0)

    def save_ai_settings(self):
        provider = self.provider_combo.currentData() or "omniroute"
        model = self.model_search_input.currentData() or ""

        from bankrotai.core import set_app_setting
        set_app_setting("ai_provider", provider)
        if model:
            set_app_setting(f"{provider}_model", model)
        if provider == "omniroute":
            set_app_setting("omniroute_protocol", "openai")
            
        self._appraiser = None # Сброс кеша для создания нового с новыми настройками
        self.status_bar.showMessage("Настройки AI сохранены", 3000)
        QMessageBox.information(
            self,
            "Успех",
            "Провайдер и модель сохранены. API-ключи задаются только через "
            "переменные окружения или secret manager.",
        )

    def change_review_status(self, status: str):
        if not self.current_selected_lot_id: return
        with session_scope() as session:
            lot = session.get(ProcessedLot, self.current_selected_lot_id)
            if lot:
                lot.review_status = status
                session.commit()
        self.status_bar.showMessage(f"Статус изменен: {status}", 3000)
        self.load_lots()

    def on_map_review_changed(self, lot_id: int, status: str):
        labels = {
            "approved": "интересен",
            "maybe": "сомневаюсь",
            "rejected": "не интересен",
        }
        self.status_bar.showMessage(f"Лот №{lot_id}: {labels.get(status, status)}", 3000)
        self.load_lots()
        script = f"window.setLotReviewStatus && window.setLotReviewStatus({int(lot_id)}, {json.dumps(status)});"
        for view_name in ("map_view", "yandex_map_view"):
            view = getattr(self, view_name, None)
            if view is not None:
                view.page().runJavaScript(script)

    def on_map_preview_opened(self, map_kind: str, lot_id: int):
        sidebar = getattr(
            self,
            "yandex_cadastre_sidebar" if map_kind == "yandex" else "map_cadastre_sidebar",
            None,
        )
        if sidebar is not None:
            sidebar.hide()
        if lot_id in self.preview_enrichment_workers:
            return
        worker = PreviewEnrichmentWorker(lot_id)
        self.preview_enrichment_workers[lot_id] = worker
        worker.result_ready.connect(self.on_preview_enrichment_finished)
        worker.failed.connect(self.on_preview_enrichment_error)
        worker.finished.connect(lambda lot_id=lot_id, worker=worker: self.cleanup_preview_enrichment_worker(lot_id, worker))
        worker.start()

    def on_map_preview_closed(self, map_kind: str):
        sidebar = getattr(
            self,
            "yandex_cadastre_sidebar" if map_kind == "yandex" else "map_cadastre_sidebar",
            None,
        )
        if sidebar is not None:
            sidebar.show()

    def on_preview_enrichment_finished(self, lot_id: int, extras: object):
        encoded = json.dumps(extras if isinstance(extras, dict) else {}, ensure_ascii=False)
        script = f"window.updateLotPreviewExtras && window.updateLotPreviewExtras({int(lot_id)}, {encoded});"
        for view_name in ("map_view", "yandex_map_view"):
            view = getattr(self, view_name, None)
            if view is not None:
                view.page().runJavaScript(script)

    def on_preview_enrichment_error(self, lot_id: int, message: str):
        logger.debug("Preview enrichment unavailable for lot %s: %s", lot_id, message)

    def cleanup_preview_enrichment_worker(self, lot_id: int, worker: PreviewEnrichmentWorker):
        if self.preview_enrichment_workers.get(int(lot_id)) is worker:
            self.preview_enrichment_workers.pop(int(lot_id), None)
        worker.deleteLater()

    def closeEvent(self, event):
        nationwide_worker = getattr(self, "all_russia_worker", None)
        if nationwide_worker is not None and nationwide_worker.isRunning():
            nationwide_worker.request_stop()
            if not nationwide_worker.wait(70_000):
                logger.error("Nationwide search worker did not stop before application shutdown")
                event.ignore()
                return
        workers = list(self.preview_enrichment_workers.values())
        for worker in workers:
            worker.requestInterruption()
        for worker in workers:
            if worker.isRunning() and not worker.wait(30_000):
                logger.error("Preview enrichment worker did not stop before application shutdown")
                event.ignore()
                return
        super().closeEvent(event)

    def refresh_geo(self):
        if not self.current_selected_lot_id: return
        self.status_bar.showMessage("Обновление координат...")
        with session_scope() as session:
            lot = session.get(ProcessedLot, self.current_selected_lot_id)
            if lot:
                # Удаляем старые координаты
                from bankrotai.db import LotGeoSnapshot
                session.query(LotGeoSnapshot).filter_by(lot_id=lot.id).delete()
                from bankrotai.geo import enrich_lot_geo
                enrich_lot_geo(session, lot)
                session.commit()
        self.status_bar.showMessage("Геокодирование завершено", 3000)
        self.update_map()

    def refresh_geo(self):
        lot_ids = self.get_selected_lot_ids()
        if not lot_ids:
            return
        self.status_bar.showMessage(f"Проверка геометок выбранных лотов: {len(lot_ids)}")
        self.start_geo_worker(lot_ids=lot_ids, refresh_existing=False, limit=None)

    def run_cleanup(self):
        if QMessageBox.question(self, "Очистка", "Удалить все завершенные торги?") == QMessageBox.Yes:
            with session_scope() as session:
                count = cleanup_closed_lots(session)
            QMessageBox.information(self, "Готово", f"Удалено лотов: {count}")
            self.load_lots()
            self.update_dashboard()

    def _update_map_legacy_removed(self):
        self.status_bar.showMessage("Загрузка карты (Leaflet/OSM)...")
        lots_with_geo = []
        try:
            with session_scope() as session:
                stmt = select(LotGeoSnapshot.centroid_lat, LotGeoSnapshot.centroid_lon, ProcessedLot.title, ProcessedLot.current_price, ProcessedLot.address, ProcessedLot.cadastral_number).join(ProcessedLot, LotGeoSnapshot.lot_id == ProcessedLot.id).where(LotGeoSnapshot.centroid_lat.isnot(None)).limit(700)
                for lat, lon, title, price, addr, cad in session.execute(stmt):
                    lots_with_geo.append({
                        "lat": float(lat),
                        "lng": float(lon),
                        "title": str(title)[:80],
                        "price": f"{float(price or 0):,.0f} ₽",
                        "address": str(addr or "")[:100],
                        "cadastral": str(cad or "")
                    })
        except Exception as e:
            print(f"Map error: {e}")

        if not lots_with_geo:
            lots_with_geo = [{"lat": 57.6261, "lng": 39.8845, "title": "Ярославль (Центр)", "price": "—", "address": "Ярославль", "cadastral": ""}]

        html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
    <link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css" />
    <link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css" />
    <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
    <script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
    <style>
        html, body, #map {{ width: 100%; height: 100%; margin: 0; padding: 0; }}
        .leaflet-popup-content-wrapper {{ border-radius: 8px; font-family: sans-serif; }}
        .popup-title {{ font-weight: bold; color: #2c3e50; font-size: 14px; margin-bottom: 5px; }}
        .popup-price {{ color: #27ae60; font-weight: bold; font-size: 13px; }}
    </style>
</head>
<body>
    <div id="map"></div>
    <script>
        var map = L.map('map').setView([57.62, 39.88], 10);
        L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{
            attribution: '&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> contributors'
        }}).addTo(map);

        var cadastralLayer = L.tileLayer.wms('https://pkk.rosreestr.ru/arcgis/services/Cadastre/Cadastre/MapServer/WmsServer?', {{
            layers: '1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24',
            format: 'image/png',
            transparent: true,
            opacity: 0.7,
            attribution: 'Росреестр'
        }});

        window.setCadLayerVisible = function(visible) {{
            if (visible) {{
                map.addLayer(cadastralLayer);
            }} else {{
                map.removeLayer(cadastralLayer);
            }}
        }};

        window.showCadastreObject = function(data) {{
            if (data.centroid_lat && data.centroid_lon) {{
                map.setView([data.centroid_lat, data.centroid_lon], 16);

                // Если есть геометрия - рисуем полигон
                if (data.geometry_json && data.geometry_json.type === 'Polygon') {{
                    var coords = data.geometry_json.coordinates[0].map(c => [c[1], c[0]]);
                    L.polygon(coords, {{color: 'red', weight: 2}}).addTo(map)
                        .bindPopup('<b>' + (data.name || 'Объект') + '</b><br>' + (data.address || ''));
                }} else {{
                    // Иначе просто маркер
                    L.marker([data.centroid_lat, data.centroid_lon]).addTo(map)
                        .bindPopup('<b>' + (data.name || 'Объект') + '</b><br>' + (data.address || ''))
                        .openPopup();
                }}
            }}
        }};

        var lots = {json.dumps(lots_with_geo, ensure_ascii=False)};
        var markers = L.markerClusterGroup({{ chunkedLoading: true, maxClusterRadius: 50 }});

        lots.forEach(function(lot) {{
            var marker = L.marker([lot.lat, lot.lng])
                .bindPopup('<div class="popup-title">' + lot.title + '</div>' +
                           '<div class="popup-price">Цена: ' + lot.price + '</div>' +
                           '<div style="font-size: 11px; color: #7f8c8d; margin-top: 5px;">' + lot.address + '</div>' +
                           (lot.cadastral ? '<div style="font-size: 10px; color: #95a5a6; margin-top: 3px;">КН: ' + lot.cadastral + '</div>' : ''));
            markers.addLayer(marker);
        }});

        map.addLayer(markers);
        if (lots.length > 0) {{
            map.fitBounds(markers.getBounds(), {{ padding: [30, 30] }});
        }}
    </script>
</body>
</html>'''
        self.web_view.setHtml(html, QUrl("https://unpkg.com"))
        self.status_bar.showMessage("Карта обновлена", 3000)

    def toggle_cadastral_layer(self):
        visible = self.cad_layer_checkbox.isChecked()
        js = f"window.setCadLayerVisible({str(visible).lower()});"
        self.web_view.page().runJavaScript(js)

    def search_cadastral_object(self):
        cadastral_number = self.cadastral_search_input.text().strip()
        if not cadastral_number:
            QMessageBox.warning(self, "Ошибка", "Введите кадастровый номер")
            return

        self.cadastral_result_text.setPlainText("Поиск...")
        self.status_bar.showMessage(f"Поиск объекта {cadastral_number}...")

        from bankrotai.geo import CADASTRAL_GEOCODER
        result = CADASTRAL_GEOCODER.search(cadastral_number)

        if not result:
            self.cadastral_result_text.setPlainText(f"❌ Объект не найден\n\nКадастровый номер: {cadastral_number}\n\nВозможные причины:\n- Неверный формат номера\n- Объект не зарегистрирован в ЕГРН\n- Ошибка связи с сервером Росреестра")
            self.status_bar.showMessage("Объект не найден", 3000)
            return

        # Формируем текст результата
        info_lines = [
            f"✅ Объект найден",
            f"",
            f"Запрос: {cadastral_number}",
            f"Источник: {result.get('source', 'pkk')}",
            f"Уверенность: {result.get('confidence', 'unknown')}",
            f"",
            f"📍 ОСНОВНАЯ ИНФОРМАЦИЯ",
            f"Объект: {result.get('object_type') or '—'}",
            f"Кадастровый номер: {result.get('cadastral_number') or '—'}",
            f"Название: {result.get('name') or '—'}",
            f"Адрес: {result.get('address') or '—'}",
            f"",
        ]

        if result.get('centroid_lat') and result.get('centroid_lon'):
            info_lines.append(f"Координаты: {result['centroid_lat']:.6f}, {result['centroid_lon']:.6f}")
        else:
            info_lines.append("Координаты: не найдены")

        if result.get('geometry_json'):
            info_lines.append("Границы: есть координаты границ")
        else:
            info_lines.append("Границы: без координат границ")

        info_lines.append("")
        info_lines.append("📊 ДЕТАЛЬНАЯ ИНФОРМАЦИЯ")

        # Дополнительные поля
        fields = {
            "Назначение": result.get('purpose'),
            "Статус": result.get('status'),
            "Площадь": f"{result.get('area_value')} {result.get('area_unit', '')}" if result.get('area_value') else None,
            "Кадастровая стоимость": f"{result.get('cad_cost')} руб." if result.get('cad_cost') else None,
            "Этажность": result.get('floors'),
            "Год постройки": result.get('year_built'),
            "Год ввода в эксплуатацию": result.get('year_commissioning'),
            "Кадастровый квартал": result.get('cadastral_quarter'),
            "Дата создания записи": result.get('date_create'),
            "Дата оценки стоимости": result.get('date_cost'),
            "ОКН": "Да" if result.get('cultural_heritage') else None,
        }

        for label, value in fields.items():
            if value:
                info_lines.append(f"{label}: {value}")
            else:
                info_lines.append(f"{label}: —")

        self.cadastral_result_text.setPlainText("\n".join(info_lines))
        self.status_bar.showMessage("Объект найден", 3000)

        # Показываем на карте
        js = f"window.showCadastreObject({json.dumps(result, ensure_ascii=False)});"
        self.web_view.page().runJavaScript(js)

    @staticmethod
    def _map_payload(
        lot: ProcessedLot,
        geo: LotGeoSnapshot,
        source_lots: list[SourceLot] | None = None,
    ) -> dict:
        def display_datetime(value: datetime | None) -> str | None:
            return value.strftime("%d.%m.%Y %H:%M") if value else None

        source_lots = source_lots or []
        primary_source = next(
            (
                source
                for source in source_lots
                if source.source_system == lot.source_system and source.external_id == lot.external_id
            ),
            source_lots[0] if source_lots else None,
        )
        source_url = (
            primary_source.source_url if primary_source and primary_source.source_url
            else lot.source_url or lot.lot_url
        )
        source_system = (lot.source_system or lot.source or "").lower()
        image_urls: list[str] = []
        gis_torgi_url = None
        etp_url = None
        torgi_russia_url = None
        for source in source_lots:
            raw_data = source.raw_data if isinstance(source.raw_data, dict) else {}
            for image_url in extract_preview_image_urls(raw_data):
                if image_url not in image_urls:
                    image_urls.append(image_url)
            system = (source.source_system or "").casefold()
            candidate_url = source.source_url
            gis_torgi_url = gis_torgi_url or raw_data.get("gis_torgi_url") or (
                candidate_url if "torgi.gov" in system else None
            )
            etp_url = etp_url or raw_data.get("etp_url") or (
                candidate_url if "lot-online" in system else None
            )
            torgi_russia_url = torgi_russia_url or raw_data.get("torgi_russia_url")
        return {
            "id": lot.id,
            "title": lot.title,
            "description": lot.description,
            "price": float(lot.current_price) if lot.current_price else None,
            "market_price": float(lot.market_price) if lot.market_price else None,
            "discount": lot.discount_percent,
            "risk": lot.risk_score,
            "rating": lot.rating,
            "address": lot.address,
            "cadastral_number": lot.cadastral_number,
            "category": lot.category,
            "status": lot.auction_status,
            "review_status": lot.review_status,
            "lat": geo.centroid_lat,
            "lon": geo.centroid_lon,
            "geo_source": geo.geo_source,
            "geo_confidence": geo.geo_confidence,
            "geometry": geo.geometry_json,
            "metadata": geo.metadata_json,
            "url": lot.lot_url,
            "source": lot.source_system or lot.source,
            "source_name": (
                " / ".join(dict.fromkeys(
                    source.platform_name or source.source_system for source in source_lots
                ))
                if source_lots else lot.source_system or lot.source
            ),
            "source_url": source_url,
            "gis_torgi_url": gis_torgi_url or (source_url if "torgi.gov" in source_system else None),
            "etp_url": etp_url or (source_url if "lot-online" in source_system else None),
            "torgi_russia_url": torgi_russia_url,
            "image_url": image_urls[0] if image_urls else None,
            "image_urls": image_urls,
            "procedure_number": next((source.procedure_number for source in source_lots if source.procedure_number), None),
            "application_deadline": display_datetime(next((source.application_deadline for source in source_lots if source.application_deadline), None)),
            "auction_at": display_datetime(next((source.auction_at for source in source_lots if source.auction_at), None)),
        }

    def _load_map_lots(self, *, lot_id: int | None = None, limit: int = MAP_MARKER_LIMIT) -> list[dict]:
        with session_scope() as session:
            latest_geo = (
                select(
                    LotGeoSnapshot.lot_id,
                    func.max(LotGeoSnapshot.id).label("geo_id"),
                )
                .where(
                    LotGeoSnapshot.centroid_lat.isnot(None),
                    LotGeoSnapshot.centroid_lon.isnot(None),
                )
                .group_by(LotGeoSnapshot.lot_id)
                .subquery()
            )
            stmt = (
                select(ProcessedLot, LotGeoSnapshot)
                .join(latest_geo, latest_geo.c.lot_id == ProcessedLot.id)
                .join(LotGeoSnapshot, LotGeoSnapshot.id == latest_geo.c.geo_id)
                .where(ProcessedLot.duplicate_of_id.is_(None))
                .order_by(LotGeoSnapshot.observed_at.desc(), LotGeoSnapshot.id.desc())
            )
            if lot_id is not None:
                stmt = stmt.where(ProcessedLot.id == lot_id)
            else:
                stmt = stmt.limit(limit)

            rows = list(session.execute(stmt))
            primary_ids = [lot.id for lot, _geo in rows]
            source_map: dict[int, list[SourceLot]] = {lot_id: [] for lot_id in primary_ids}
            if primary_ids:
                source_rows = session.execute(
                    select(ProcessedLot.id, ProcessedLot.duplicate_of_id, SourceLot)
                    .join(SourceLot, SourceLot.processed_lot_id == ProcessedLot.id)
                    .where(or_(
                        ProcessedLot.id.in_(primary_ids),
                        ProcessedLot.duplicate_of_id.in_(primary_ids),
                    ))
                )
                for processed_id, duplicate_of_id, source_lot in source_rows:
                    source_map.setdefault(duplicate_of_id or processed_id, []).append(source_lot)
            return [
                self._map_payload(lot, geo, source_map.get(lot.id, []))
                for lot, geo in rows
            ]

    def _load_map_lot(self, lot_id: int) -> dict | None:
        lots = self._load_map_lots(lot_id=lot_id)
        return lots[0] if lots else None

    def update_map(self):
        lots = self._load_map_lots()
        html = self.build_map_html([])
        base_url = QUrl.fromLocalFile(str(map_assets_directory()) + os.sep)
        self._set_map_document(self.map_view, html, lots, "_leaflet_load_handler", base_url)
        self.status_bar.showMessage("Карта обновлена", 3000)

    def update_yandex_map(self):
        lots = self._load_map_lots()
        html = self.build_yandex_map_html([])
        base_url = QUrl.fromLocalFile(str(map_assets_directory()) + os.sep)
        self._set_map_document(
            self.yandex_map_view,
            html,
            lots,
            "_yandex_load_handler",
            base_url,
        )
        self.status_bar.showMessage("Яндекс-карта обновлена", 3000)

    def _set_map_document(
        self,
        view: QWebEngineView,
        html: str,
        lots: list[dict],
        handler_attribute: str,
        base_url: QUrl,
    ) -> None:
        previous = getattr(self, handler_attribute, None)
        if previous is not None:
            try:
                view.loadFinished.disconnect(previous)
            except (RuntimeError, TypeError):
                pass

        def loaded(ok: bool) -> None:
            try:
                view.loadFinished.disconnect(loaded)
            except (RuntimeError, TypeError):
                pass
            setattr(self, handler_attribute, None)
            if not ok:
                logger.error("Map HTML failed to load for %s", handler_attribute)
                return
            for offset in range(0, len(lots), 50):
                payload = json.dumps(lots[offset:offset + 50], ensure_ascii=True)
                view.page().runJavaScript(
                    "(function(batch) {"
                    " window.__bankrotaiPendingLots = window.__bankrotaiPendingLots || [];"
                    " batch.forEach(function(lot) {"
                    "   if (window.upsertLot) { window.upsertLot(lot); }"
                    "   else { window.__bankrotaiPendingLots.push(lot); }"
                    " });"
                    f"}})({payload});"
                )
            view.page().runJavaScript(
                "if (window.fitAllLots) { window.fitAllLots(); }"
            )

        setattr(self, handler_attribute, loaded)
        view.loadFinished.connect(loaded)
        view.setHtml(html, base_url)

    def build_yandex_map_html(self, lots: list[dict]) -> str:
        lots_json = json.dumps(lots, ensure_ascii=False)
        return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>BankrotAI Yandex Map</title>
<link rel="stylesheet" href="leaflet.css">
<script src="https://api-maps.yandex.ru/2.1/?lang=ru_RU" type="text/javascript"></script>
<script src="leaflet.js"></script>
<script src="qrc:///qtwebchannel/qwebchannel.js"></script>
<style>
html, body, #map {{
    width: 100%;
    height: 100%;
    margin: 0;
    padding: 0;
}}
.hint {{
    position: absolute;
    z-index: 10;
    left: 12px;
    top: 12px;
    background: rgba(255,255,255,.94);
    border: 1px solid #d8dde6;
    border-radius: 6px;
    padding: 8px 10px;
    font: 13px Arial, sans-serif;
    color: #2e3c54;
}}
{MAP_PREVIEW_STYLE}
</style>
</head>
<body>
<div id="map"></div>
<div id="hint" class="hint">Загрузка Яндекс.Карт...</div>
{MAP_PREVIEW_HTML}
<script>
const lots = {lots_json};
const mapKind = 'yandex';
let map;
let lotCollection;
let boundaryCollection;
let selectedObjectCollection;
let boundaryVisible = true;
const lotPlacemarks = new Map();

function escapeHtml(text) {{
    if (text === null || text === undefined) return '';
    return String(text)
        .replaceAll('&', '&amp;')
        .replaceAll('<', '&lt;')
        .replaceAll('>', '&gt;')
        .replaceAll('"', '&quot;')
        .replaceAll("'", '&#039;');
}}

function formatPrice(value) {{
    if (!value) return '—';
    return new Intl.NumberFormat('ru-RU').format(value) + ' ₽';
}}

function markerColor(status) {{
    return status === 'approved' ? '#24a269' : status === 'maybe' ? '#e0aa16' : status === 'rejected' ? '#d94b4b' : '#7d8795';
}}

function markerSvg(lot) {{
    const color = markerColor(lot.review_status);
    return `<svg xmlns="http://www.w3.org/2000/svg" width="38" height="48" viewBox="0 0 38 48"><path d="M19 1C9.1 1 1 9.1 1 19c0 13.2 18 28 18 28s18-14.8 18-28C37 9.1 28.9 1 19 1z" fill="${{color}}" stroke="white" stroke-width="2"/><path d="M12 23V14h14v9M10 23h18M15 18h2m4 0h2" fill="none" stroke="white" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"/></svg>`;
}}

function yandexIconOptions(lot) {{
    return {{
        iconLayout: 'default#image',
        iconImageHref: 'data:image/svg+xml;charset=UTF-8,' + encodeURIComponent(markerSvg(lot)),
        iconImageSize: [38, 48],
        iconImageOffset: [-19, -48]
    }};
}}

function geoJsonCoordinatesToYandex(coords) {{
    if (!Array.isArray(coords)) return coords;
    if (coords.length === 2 && typeof coords[0] === 'number' && typeof coords[1] === 'number') {{
        return [coords[1], coords[0]];
    }}
    return coords.map(geoJsonCoordinatesToYandex);
}}

function addGeometry(geometry, options) {{
    if (!geometry || !geometry.type || !geometry.coordinates) return;
    if (geometry.type === 'Polygon') {{
        const polygon = new ymaps.Polygon(
            geoJsonCoordinatesToYandex(geometry.coordinates),
            {{}},
            options
        );
        boundaryCollection.add(polygon);
    }} else if (geometry.type === 'MultiPolygon') {{
        geometry.coordinates.forEach(poly => {{
            const polygon = new ymaps.Polygon(
                geoJsonCoordinatesToYandex(poly),
                {{}},
                options
            );
            boundaryCollection.add(polygon);
        }});
    }}
}}

function upsertLot(lot) {{
    if (!lotCollection) {{
        window.__bankrotaiPendingLots = window.__bankrotaiPendingLots || [];
        window.__bankrotaiPendingLots.push(lot);
        return;
    }}
    if (!lot.lat || !lot.lon) return;
    const key = String(lot.id);
    const previous = lotPlacemarks.get(key);
    if (previous) lotCollection.remove(previous);
    const placemark = new ymaps.Placemark(
            [lot.lat, lot.lon],
            {{
                hintContent: escapeHtml(lot.title)
            }},
            yandexIconOptions(lot)
    );
    placemark.events.add('click', function(event) {{
        event.preventDefault();
        showLotPreview(lot);
    }});
    lotCollection.add(placemark);
    lotPlacemarks.set(key, placemark);
    if (lot.geometry) {{
        addGeometry(lot.geometry, {{
            strokeColor: '#2468d8',
            strokeWidth: 2,
            fillColor: 'rgba(36,104,216,0.08)'
        }});
    }}
}}

function addLots() {{
    lotCollection.removeAll();
    boundaryCollection.removeAll();
    lotPlacemarks.clear();

    lots.forEach(upsertLot);

    map.geoObjects.add(lotCollection);
    map.geoObjects.add(boundaryCollection);
    if (lotCollection.getLength() > 0) {{
        map.setBounds(lotCollection.getBounds(), {{ checkZoomRange: true, zoomMargin: 35 }});
    }}
}}

function showCadastreObject(data) {{
    selectedObjectCollection.removeAll();
    const body =
        '<b>Кадастр:</b> ' + escapeHtml(data.cadastral_number || '—') + '<br>' +
        '<b>Название:</b> ' + escapeHtml(data.title || '—') + '<br>' +
        '<b>Адрес:</b> ' + escapeHtml(data.address || '—') + '<br>' +
        '<b>Границы:</b> ' + (data.has_boundary ? 'получены' : 'без координат границ') + '<br>' +
        '<b>Источник:</b> ' + escapeHtml(data.source || '—');

    if (data.geometry) {{
        const beforeCount = selectedObjectCollection.getLength();
        const options = {{
            strokeColor: '#d92323',
            strokeWidth: 4,
            fillColor: 'rgba(217,35,35,0.16)'
        }};
        const oldBoundary = boundaryCollection;
        boundaryCollection = selectedObjectCollection;
        addGeometry(data.geometry, options);
        boundaryCollection = oldBoundary;
        if (selectedObjectCollection.getLength() > beforeCount) {{
            map.geoObjects.add(selectedObjectCollection);
            map.setBounds(selectedObjectCollection.getBounds(), {{ checkZoomRange: true, zoomMargin: 45 }});
            return;
        }}
    }}

    if (data.lat && data.lon) {{
        const marker = new ymaps.Placemark(
            [data.lat, data.lon],
            {{
                balloonContentHeader: escapeHtml(data.object_type || 'Объект'),
                balloonContentBody: body
            }},
            yandexIconOptions({{ category: 'real_estate' }})
        );
        selectedObjectCollection.add(marker);
        map.geoObjects.add(selectedObjectCollection);
        map.setCenter([data.lat, data.lon], 17);
        marker.balloon.open();
    }}
}}

function setCadLayerVisible(visible) {{
    boundaryVisible = visible;
    if (visible) {{
        map.geoObjects.add(boundaryCollection);
    }} else {{
        map.geoObjects.remove(boundaryCollection);
    }}
}}

window.showCadastreObject = showCadastreObject;
window.setCadLayerVisible = setCadLayerVisible;
window.upsertLot = upsertLot;
window.applyLotReviewStatus = function(lotId, status) {{
    const lot = lots.find(item => Number(item.id) === Number(lotId));
    if (lot) lot.review_status = status;
    const placemark = lotPlacemarks.get(String(lotId));
    if (placemark) placemark.options.set(yandexIconOptions(lot || {{ review_status: status }}));
}};
window.fitAllLots = function() {{
    if (lotCollection && lotCollection.getLength() > 0) {{
        map.setBounds(lotCollection.getBounds(), {{ checkZoomRange: true, zoomMargin: 35 }});
    }}
}};

{MAP_PREVIEW_SCRIPT}

function initYandexMap() {{
    map = new ymaps.Map('map', {{
        center: [57.6261, 39.8845],
        zoom: 8,
        controls: ['zoomControl', 'typeSelector', 'fullscreenControl']
    }});
    lotCollection = new ymaps.GeoObjectCollection();
    boundaryCollection = new ymaps.GeoObjectCollection();
    selectedObjectCollection = new ymaps.GeoObjectCollection();
    document.getElementById('hint').style.display = 'none';
    addLots();
    const pending = window.__bankrotaiPendingLots || [];
    window.__bankrotaiPendingLots = [];
    pending.forEach(upsertLot);
}}

function initLeafletFallback() {{
    const hint = document.getElementById('hint');
    hint.textContent = 'Яндекс.Карты недоступны — включена резервная карта';
    map = L.map('map').setView([57.6261, 39.8845], 8);
    let fallbackMapSized = false;
    new ResizeObserver(function() {{
        map.invalidateSize(false);
        const element = document.getElementById('map');
        if (!fallbackMapSized && element.clientWidth > 0 && element.clientHeight > 0) {{
            fallbackMapSized = true;
            setTimeout(function() {{ if (window.fitAllLots) window.fitAllLots(); }}, 50);
        }}
    }}).observe(document.getElementById('map'));
    L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{
        maxZoom: 19,
        attribution: '&copy; OpenStreetMap contributors'
    }}).addTo(map);
    const fallbackLayer = L.layerGroup().addTo(map);
    const fallbackMarkers = new Map();
    const fallbackBounds = L.latLngBounds([]);

    function fallbackIcon(lot) {{
        return L.divIcon({{
            className: '',
            html: markerSvg(lot),
            iconSize: [38, 48],
            iconAnchor: [19, 48]
        }});
    }}
    function upsertFallbackLot(lot) {{
        if (!lot.lat || !lot.lon) return;
        const key = String(lot.id);
        const previous = fallbackMarkers.get(key);
        if (previous) fallbackLayer.removeLayer(previous);
        const marker = L.marker([lot.lat, lot.lon], {{ icon: fallbackIcon(lot) }});
        marker.on('click', function() {{ showLotPreview(lot); }});
        marker.addTo(fallbackLayer);
        fallbackMarkers.set(key, marker);
        fallbackBounds.extend([lot.lat, lot.lon]);
    }}
    window.upsertLot = upsertFallbackLot;
    window.applyLotReviewStatus = function(lotId, status) {{
        const lot = lots.find(item => Number(item.id) === Number(lotId));
        if (lot) lot.review_status = status;
        const marker = fallbackMarkers.get(String(lotId));
        if (marker) marker.setIcon(fallbackIcon(lot || {{ review_status: status }}));
    }};
    window.showCadastreObject = function(data) {{
        if (data.geometry) {{
            const layer = L.geoJSON(data.geometry, {{ style: {{ color: '#d92323', weight: 4 }} }}).addTo(map);
            map.fitBounds(layer.getBounds(), {{ padding: [40, 40] }});
        }} else if (data.lat && data.lon) {{
            L.marker([data.lat, data.lon], {{ icon: fallbackIcon({{}}) }}).addTo(map);
            map.setView([data.lat, data.lon], 17);
        }}
    }};
    window.setCadLayerVisible = function() {{}};
    window.fitAllLots = function() {{
        if (fallbackBounds.isValid()) map.fitBounds(fallbackBounds, {{ padding: [30, 30] }});
    }};
    lots.forEach(upsertFallbackLot);
    if (fallbackBounds.isValid()) map.fitBounds(fallbackBounds, {{ padding: [30, 30] }});
    const pending = window.__bankrotaiPendingLots || [];
    window.__bankrotaiPendingLots = [];
    pending.forEach(upsertFallbackLot);
}}

if (typeof ymaps !== 'undefined') {{
    ymaps.ready(initYandexMap);
}} else {{
    initLeafletFallback();
}}
</script>
</body>
</html>
"""


    def build_map_html(self, lots: list[dict]) -> str:
        lots_json = json.dumps(lots, ensure_ascii=False)
        wms_base_url = f"http://127.0.0.1:{self.cadastral_wms_proxy_port or 0}/nspd"

        return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>BankrotAI Map</title>

<link rel="stylesheet" href="leaflet.css">
<link rel="stylesheet" href="MarkerCluster.css">
<link rel="stylesheet" href="MarkerCluster.Default.css">
<script src="qrc:///qtwebchannel/qwebchannel.js"></script>

<style>
html, body, #map {{
    height: 100%;
    width: 100%;
    margin: 0;
    padding: 0;
}}

.popup-title {{
    font-weight: bold;
    margin-bottom: 6px;
}}
{MAP_PREVIEW_STYLE}
</style>
</head>

<body>
<div id="map"></div>
{MAP_PREVIEW_HTML}

<script src="leaflet.js"></script>
<script src="leaflet.markercluster.js"></script>

<script>
const lots = {lots_json};
const mapKind = 'leaflet';

const map = L.map('map').setView([57.6261, 39.8845], 8);
let mapSized = false;
new ResizeObserver(function() {{
    map.invalidateSize(false);
    const element = document.getElementById('map');
    if (!mapSized && element.clientWidth > 0 && element.clientHeight > 0) {{
        mapSized = true;
        setTimeout(function() {{ if (window.fitAllLots) window.fitAllLots(); }}, 50);
    }}
}}).observe(document.getElementById('map'));

const osm = L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{
    maxZoom: 19,
    attribution: '&copy; OpenStreetMap contributors'
}}).addTo(map);

const cadastralLandLayer = L.tileLayer.wms(
    '{wms_base_url}/land',
    {{
        layers: '36048',
        version: '1.3.0',
        format: 'image/png',
        transparent: true,
        opacity: 0.65,
        attribution: 'НСПД / Росреестр'
    }}
);

const cadastralBuildingsLayer = L.tileLayer.wms(
    '{wms_base_url}/buildings',
    {{
        layers: '36328',
        version: '1.3.0',
        format: 'image/png',
        transparent: true,
        opacity: 0.65,
        attribution: 'НСПД / Росреестр'
    }}
);

const markers = L.markerClusterGroup();
const boundaries = L.layerGroup().addTo(map);
const lotMarkers = new Map();

let selectedObjectLayer = null;
let selectedObjectMarker = null;

function escapeHtml(text) {{
    if (text === null || text === undefined) return '';
    return String(text)
        .replaceAll('&', '&amp;')
        .replaceAll('<', '&lt;')
        .replaceAll('>', '&gt;')
        .replaceAll('"', '&quot;')
        .replaceAll("'", '&#039;');
}}

function formatPrice(value) {{
    if (!value) return '—';
    return new Intl.NumberFormat('ru-RU').format(value) + ' ₽';
}}

function markerColor(status) {{
    return status === 'approved' ? '#24a269' : status === 'maybe' ? '#e0aa16' : status === 'rejected' ? '#d94b4b' : '#7d8795';
}}

function markerSvg(lot) {{
    const color = markerColor(lot.review_status);
    return `<svg xmlns="http://www.w3.org/2000/svg" width="38" height="48" viewBox="0 0 38 48"><path d="M19 1C9.1 1 1 9.1 1 19c0 13.2 18 28 18 28s18-14.8 18-28C37 9.1 28.9 1 19 1z" fill="${{color}}" stroke="white" stroke-width="2"/><path d="M12 23V14h14v9M10 23h18M15 18h2m4 0h2" fill="none" stroke="white" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"/></svg>`;
}}

function makeIcon(lot) {{
    return L.divIcon({{
        className: '', html: markerSvg(lot), iconSize: [38, 48], iconAnchor: [19, 48], popupAnchor: [0, -46]
    }});
}}

function upsertLot(lot) {{
    if (!lot.lat || !lot.lon) return;
    const key = String(lot.id);
    const previous = lotMarkers.get(key);
    if (previous) markers.removeLayer(previous);
    const marker = L.marker([lot.lat, lot.lon], {{
            icon: makeIcon(lot)
    }});

    marker.on('click', function() {{ showLotPreview(lot); }});
    markers.addLayer(marker);
    lotMarkers.set(key, marker);

    if (lot.geometry) {{
        L.geoJSON(lot.geometry, {{
            style: {{
                weight: 1,
                opacity: 0.5,
                fillOpacity: 0.05
            }}
        }}).addTo(boundaries);
    }}
    if (!map.hasLayer(markers)) map.addLayer(markers);
}}

function addLots() {{
    markers.clearLayers();
    boundaries.clearLayers();
    lotMarkers.clear();

    lots.forEach(upsertLot);

    map.addLayer(markers);

    if (markers.getLayers().length > 0) {{
        map.fitBounds(markers.getBounds(), {{ padding: [30, 30] }});
    }}
}}

function showCadastreObject(data) {{
    if (selectedObjectLayer) {{
        map.removeLayer(selectedObjectLayer);
        selectedObjectLayer = null;
    }}

    if (selectedObjectMarker) {{
        map.removeLayer(selectedObjectMarker);
        selectedObjectMarker = null;
    }}

    const popup = `
        <div class="popup-title">${{escapeHtml(data.object_type || 'Объект')}}</div>
        <div><b>Кадастр:</b> ${{escapeHtml(data.cadastral_number || '—')}}</div>
        <div><b>Название:</b> ${{escapeHtml(data.title || '—')}}</div>
        <div><b>Адрес:</b> ${{escapeHtml(data.address || '—')}}</div>
        <div><b>Границы:</b> ${{data.has_boundary ? 'получены' : 'без координат границ'}}</div>
        <div><b>Источник:</b> ${{escapeHtml(data.source || '—')}}</div>
    `;

    if (data.geometry) {{
        selectedObjectLayer = L.geoJSON(data.geometry, {{
            style: {{
                weight: 4,
                opacity: 1,
                fillOpacity: 0.18
            }}
        }}).addTo(map);

        selectedObjectLayer.bindPopup(popup).openPopup();
        map.fitBounds(selectedObjectLayer.getBounds(), {{ padding: [40, 40] }});
        return;
    }}

    if (data.lat && data.lon) {{
        selectedObjectMarker = L.marker([data.lat, data.lon], {{
            icon: makeIcon({{ category: 'real_estate' }})
        }}).addTo(map);
        selectedObjectMarker.bindPopup(popup).openPopup();
        map.setView([data.lat, data.lon], 17);
    }}
}}

function setCadLayerVisible(visible) {{
    if (visible) {{
        if (!map.hasLayer(cadastralLandLayer)) map.addLayer(cadastralLandLayer);
        if (!map.hasLayer(cadastralBuildingsLayer)) map.addLayer(cadastralBuildingsLayer);
    }} else {{
        if (map.hasLayer(cadastralLandLayer)) map.removeLayer(cadastralLandLayer);
        if (map.hasLayer(cadastralBuildingsLayer)) map.removeLayer(cadastralBuildingsLayer);
    }}
}}

window.showCadastreObject = showCadastreObject;
window.setCadLayerVisible = setCadLayerVisible;
window.upsertLot = upsertLot;
window.applyLotReviewStatus = function(lotId, status) {{
    const lot = lots.find(item => Number(item.id) === Number(lotId));
    if (lot) lot.review_status = status;
    const marker = lotMarkers.get(String(lotId));
    if (marker) marker.setIcon(makeIcon(lot || {{ review_status: status }}));
}};
window.fitAllLots = function() {{
    if (markers.getLayers().length > 0) {{
        map.fitBounds(markers.getBounds(), {{ padding: [30, 30] }});
    }}
}};

{MAP_PREVIEW_SCRIPT}

addLots();
const pending = window.__bankrotaiPendingLots || [];
window.__bankrotaiPendingLots = [];
pending.forEach(upsertLot);
</script>
</body>
</html>
"""

    def search_cadastre_from_gui(self):
        query = self.normalize_cadastre_query(self.cad_search_input.text().strip())
        if not query:
            QMessageBox.warning(self, "Поиск", "Введите кадастровый номер или адрес")
            return

        self.cad_search_btn.setEnabled(False)
        self.cad_search_btn.setText("Ищу...")
        self.cad_info_text.setPlainText("Ищу объект...")
        self.status_bar.showMessage(f"Кадастровый поиск: {query}")

        self.cadastre_search_worker = CadastreSearchWorker(query)
        self.cadastre_search_worker.finished.connect(self.on_cadastre_search_finished)
        self.cadastre_search_worker.error.connect(self.on_cadastre_search_error)
        self.cadastre_search_worker.start()

    def normalize_cadastre_query(self, query: str) -> str:
        if not query:
            return ""
        if re.match(r"^\d{2}:\d{2}:\d{6,7}:\d+$", query):
            return query
        lower = query.lower()
        known_place = any(
            place in lower
            for place in ("ярослав", "рыбинск", "архангельск", "москва", "санкт-петербург", "ленинградская")
        )
        if known_place:
            return query
        return f"Ярославль, {query}"

    def update_cadastre_address_suggestions(self, text: str):
        q = text.strip()
        if not q or re.match(r"^\d{2}:\d{2}:\d{0,7}:?\d*$", q):
            self.cad_suggestion_model.setStringList([])
            return

        lower = q.lower()
        if any(place in lower for place in ("ярослав", "рыбинск", "архангельск", "москва", "санкт-петербург")):
            suggestions = [q]
        else:
            suggestions = [
                f"Ярославль, {q}",
                f"Ярославская область, Ярославль, {q}",
                f"Рыбинск, {q}",
            ]
        self.cad_suggestion_model.setStringList(suggestions)

    def on_cadastre_search_finished(self, result):
        self.cad_search_btn.setEnabled(True)
        self.cad_search_btn.setText("Найти объект")
        self.status_bar.showMessage("Кадастровый поиск завершен", 3000)
        self.show_cadastre_info(result)
        self.show_cadastre_object_on_map(result)

    def on_cadastre_search_error(self, error_msg: str):
        self.cad_search_btn.setEnabled(True)
        self.cad_search_btn.setText("Найти объект")
        self.status_bar.showMessage("Ошибка кадастрового поиска", 5000)
        self.cad_info_text.setPlainText(f"Ошибка поиска: {error_msg}")

    def show_cadastre_info(self, result):
        if result.error:
            self.cad_info_text.setPlainText(
                f"Ошибка: {result.error}\n\n"
                f"Запрос: {result.query}"
            )
            return

        lines = []

        lines.append(f"Запрос: {result.query}")
        lines.append(f"Источник: {result.source}")
        lines.append(f"Уверенность: {result.confidence}")
        lines.append("")

        lines.append(f"Объект: {result.object_type or '—'}")
        lines.append(f"Кадастровый номер: {result.cadastral_number or '—'}")
        lines.append(f"Название: {result.title or '—'}")
        lines.append(f"Адрес: {result.address or '—'}")
        lines.append("")

        if result.lat and result.lon:
            lines.append(f"Координаты: {result.lat:.6f}, {result.lon:.6f}")
        else:
            lines.append("Координаты: —")

        lines.append("Границы: получены" if result.has_boundary else "Границы: без координат границ")
        lines.append("")
        lines.append("Информация:")

        for key, value in (result.info or {}).items():
            lines.append(f"{key}: {value if value not in (None, '') else '—'}")

        self.cad_info_text.setPlainText("\n".join(lines))

    def show_cadastre_object_on_map(self, result):
        payload = {
            "query": result.query,
            "cadastral_number": result.cadastral_number,
            "object_type": result.object_type,
            "title": result.title,
            "address": result.address,
            "lat": result.lat,
            "lon": result.lon,
            "geometry": result.geometry_json,
            "has_boundary": result.has_boundary,
            "source": result.source,
            "confidence": result.confidence,
            "info": result.info,
        }

        js = f"window.showCadastreObject({json.dumps(payload, ensure_ascii=False)});"
        self.map_view.page().runJavaScript(js)

    def toggle_cad_layer_from_gui(self):
        visible = self.cad_layer_checkbox.isChecked()
        js = f"window.setCadLayerVisible({str(visible).lower()});"
        self.map_view.page().runJavaScript(js)

    def search_yandex_cadastre_from_gui(self):
        query = self.normalize_cadastre_query(self.yandex_cad_search_input.text().strip())
        if not query:
            QMessageBox.warning(self, "Поиск", "Введите кадастровый номер или адрес")
            return

        self.yandex_cad_search_btn.setEnabled(False)
        self.yandex_cad_search_btn.setText("Ищу...")
        self.yandex_cad_info_text.setPlainText("Ищу объект...")
        self.status_bar.showMessage(f"Кадастровый поиск: {query}")

        self.yandex_cadastre_search_worker = CadastreSearchWorker(query)
        self.yandex_cadastre_search_worker.finished.connect(self.on_yandex_cadastre_search_finished)
        self.yandex_cadastre_search_worker.error.connect(self.on_yandex_cadastre_search_error)
        self.yandex_cadastre_search_worker.start()

    def update_yandex_cadastre_address_suggestions(self, text: str):
        q = text.strip()
        if not q or re.match(r"^\d{2}:\d{2}:\d{0,7}:?\d*$", q):
            self.yandex_cad_suggestion_model.setStringList([])
            return

        lower = q.lower()
        if any(place in lower for place in ("ярослав", "рыбинск", "архангельск", "москва", "санкт-петербург")):
            suggestions = [q]
        else:
            suggestions = [
                f"Ярославль, {q}",
                f"Ярославская область, Ярославль, {q}",
                f"Рыбинск, {q}",
            ]
        self.yandex_cad_suggestion_model.setStringList(suggestions)

    def on_yandex_cadastre_search_finished(self, result):
        self.yandex_cad_search_btn.setEnabled(True)
        self.yandex_cad_search_btn.setText("Найти объект")
        self.status_bar.showMessage("Кадастровый поиск завершен", 3000)
        self.show_yandex_cadastre_info(result)
        self.show_yandex_cadastre_object_on_map(result)

    def on_yandex_cadastre_search_error(self, error_msg: str):
        self.yandex_cad_search_btn.setEnabled(True)
        self.yandex_cad_search_btn.setText("Найти объект")
        self.status_bar.showMessage("Ошибка кадастрового поиска", 5000)
        self.yandex_cad_info_text.setPlainText(f"Ошибка поиска: {error_msg}")

    def show_yandex_cadastre_info(self, result):
        if result.error:
            self.yandex_cad_info_text.setPlainText(
                f"Ошибка: {result.error}\n\n"
                f"Запрос: {result.query}"
            )
            return

        lines = [
            f"Запрос: {result.query}",
            f"Источник: {result.source}",
            f"Уверенность: {result.confidence}",
            "",
            f"Объект: {result.object_type or '—'}",
            f"Кадастровый номер: {result.cadastral_number or '—'}",
            f"Название: {result.title or '—'}",
            f"Адрес: {result.address or '—'}",
            "",
        ]

        if result.lat and result.lon:
            lines.append(f"Координаты: {result.lat:.6f}, {result.lon:.6f}")
        else:
            lines.append("Координаты: —")

        lines.append("Границы: получены" if result.has_boundary else "Границы: без координат границ")
        lines.append("")
        lines.append("Информация:")
        for key, value in (result.info or {}).items():
            lines.append(f"{key}: {value if value not in (None, '') else '—'}")

        self.yandex_cad_info_text.setPlainText("\n".join(lines))

    def show_yandex_cadastre_object_on_map(self, result):
        payload = {
            "query": result.query,
            "cadastral_number": result.cadastral_number,
            "object_type": result.object_type,
            "title": result.title,
            "address": result.address,
            "lat": result.lat,
            "lon": result.lon,
            "geometry": result.geometry_json,
            "has_boundary": result.has_boundary,
            "source": result.source,
            "confidence": result.confidence,
            "info": result.info,
        }
        js = f"window.showCadastreObject({json.dumps(payload, ensure_ascii=False)});"
        self.yandex_map_view.page().runJavaScript(js)

    def toggle_yandex_cad_layer_from_gui(self):
        visible = self.yandex_cad_layer_checkbox.isChecked()
        js = f"window.setCadLayerVisible({str(visible).lower()});"
        self.yandex_map_view.page().runJavaScript(js)

def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
